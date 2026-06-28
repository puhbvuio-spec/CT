# -*- coding: utf-8 -*-
"""Excel 数据文件合并与规范化清洗模块。

本模块提供将多个零散的同结构 Excel 数据文件合并为单个文件的工作，
支持依据关键字过滤文件名、统一“序号”字段对齐行号，并实现数据清洗防护。
"""

from __future__ import annotations

import argparse
import time
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.core import build_output_path, sanitize_xlsx_cell


# 平台别名至归一化前缀的映射映射表
PLATFORM_PREFIX = {
    "youtube": "youtube",
    "tiktok": "tiktok",
    "x": "x",
    "twitter": "x",
}


def normalize_platform(platform: str) -> str:
    """归一化平台标识名称。

    Args:
        platform: 原始平台名称。

    Returns:
        str: 映射后的标准前缀（例如 "x"），默认返回 "merged"。
    """
    value = (platform or "").strip().lower()
    if value in PLATFORM_PREFIX:
        return PLATFORM_PREFIX[value]
    return value or "merged"


def find_xlsx_files(folder: str | Path, keyword: str, output_file: str | Path | None = None) -> list[Path]:
    """扫描指定目录下符合名称关键字条件的所有 Excel (.xlsx) 文件，并过滤掉临时文件及输出目标自身。

    Args:
        folder: 待扫描的目录路径。
        keyword: 文件名过滤关键字（不区分大小写）。
        output_file: 正在输出的目标文件路径，若存在则将其排除，防止自循环读写。

    Returns:
        list[Path]: 排序后的符合条件的 Excel 文件 Path 列表。
    """
    folder_path = Path(folder)
    keyword = (keyword or "").strip().lower()
    output_name = Path(output_file).name.lower() if output_file else ""
    files: list[Path] = []
    for path in sorted(folder_path.glob("*.xlsx")):
        name = path.name.lower()
        # 1. 过滤掉与输出文件名相同的文件，防止产生重复吞噬
        if output_name and name == output_name:
            continue
        # 2. 过滤掉 Windows/Excel 临时自动保存文件（如 ~$data.xlsx）
        if path.name.startswith("~$"):
            continue
        # 3. 过滤掉文件名不含过滤关键字的文件
        if keyword and keyword not in name:
            continue
        files.append(path)
    return files


def _normalize_headers(raw_headers) -> list[str]:
    """表头行数据清洗与规范化，去除两端空白，丢弃尾部空单元格。

    Args:
        raw_headers: 原始表头数据元组或列表。

    Returns:
        list[str]: 规范化处理后的表头字符串列表。
    """
    headers = [str(value).strip() if value is not None else "" for value in raw_headers]
    # 从末尾移除可能存在的空字符表头
    while headers and not headers[-1]:
        headers.pop()
    return headers


def merge_xlsx_files(
    folder: str | Path,
    keyword: str = "keyword",
    platform: str = "merged",
    output_file: str | Path | None = None,
) -> tuple[str, int, int]:
    """合并指定文件夹下的多个 Excel 文件，对齐表头，维护统一自增序号，防止注入。

    支持对不规则表头或缺失“序号”的原始文件进行补全及位置对齐，
    在写入合并表前对单元格进行安全清洁处理。

    Args:
        folder: 源 Excel 文件所在的目录。
        keyword: 文件名包含的关键字，用于定位要合并的子集。
        platform: 对应的平台前缀，作为自动命名时子文件夹与前缀标识。
        output_file: 可选。指定合并后导出的绝对路径。若空则自动生成。

    Returns:
        tuple[str, int, int]: (最终保存的文件绝对路径, 成功合并的文件总数, 累计合并的有效数据行数)。
    """
    platform_prefix = normalize_platform(platform)
    output_path = Path(output_file) if output_file else Path(
        build_output_path(platform_prefix, f"{platform_prefix}_merge_{time.strftime('%Y%m%d_%H%M%S')}.xlsx", channel="merge")
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)

    files = find_xlsx_files(folder, keyword, output_path)
    if not files:
        if keyword:
            raise FileNotFoundError(f"没有找到文件名包含“{keyword}”的 .xlsx 文件")
        raise FileNotFoundError("没有找到可合并的 .xlsx 文件")

    output_wb = Workbook()
    output_ws = output_wb.active
    output_ws.title = "合并数据"

    headers: list[str] | None = None
    serial_col_index: int | None = None
    current_no = 1
    merged_rows = 0
    merged_files = 0

    for file_path in files:
        wb = None
        try:
            # 开启 read_only=True 与 data_only=True 以最小内存只读方式加速加载公式计算结果值
            wb = load_workbook(file_path, read_only=True, data_only=True)
            file_row_count = 0

            for ws in wb.worksheets:
                row_iter = ws.iter_rows(values_only=True)
                source_headers = _normalize_headers(next(row_iter, []))
                if not source_headers or all(not value for value in source_headers):
                    continue

                # 初始化全局对齐表头
                if headers is None:
                    headers = list(source_headers)
                    if "序号" in headers:
                        serial_col_index = headers.index("序号")
                    else:
                        # 若原始文件无“序号”列，动态在最前方补位加入
                        headers.insert(0, "序号")
                        serial_col_index = 0
                    output_ws.append(headers)

                # 构建源文件表头的名称到列索引的哈希映射映射表，以防止列乱序错位
                source_index = {name: index for index, name in enumerate(source_headers)}
                for row_values in row_iter:
                    # 排除全空行
                    if not row_values or all(value is None or str(value).strip() == "" for value in row_values):
                        continue
                    output_row = []
                    for column_index, header in enumerate(headers):
                        if column_index == serial_col_index:
                            # 填入全局自增行号，覆盖原始不连续的局部序号
                            output_row.append(current_no)
                        else:
                            source_pos = source_index.get(header)
                            # 根据表头名字动态定位数据，若原文件缺少该列字段，自动补位空字串
                            value = row_values[source_pos] if source_pos is not None and source_pos < len(row_values) else ""
                            # 调用 sanitize_xlsx_cell 滤除以 '=', '+', '-' 等起始的恶意公式注入输入
                            output_row.append(sanitize_xlsx_cell(value))
                    output_ws.append(output_row)
                    current_no += 1
                    merged_rows += 1
                    file_row_count += 1

            if file_row_count:
                merged_files += 1
                output_wb.save(output_path)
        except Exception as exc:
            print(f"警告：跳过文件 {file_path.name}（{exc}）")
        finally:
            if wb is not None:
                wb.close()

    if merged_rows <= 0:
        raise ValueError("没有成功读取任何有效数据，请检查文件、关键词或表头")

    output_wb.save(output_path)
    return str(output_path), merged_files, merged_rows


def main(argv=None):
    """合并脚本入口函数，解析命令行参数并执行。"""
    parser = argparse.ArgumentParser(description="合并多个 xlsx 文件")
    parser.add_argument("folder", help="包含 xlsx 文件的文件夹")
    parser.add_argument("--keyword", default="keyword", help="文件名包含的关键词，留空则合并所有 xlsx")
    parser.add_argument("--platform", default="merged", help="平台前缀，例如 youtube/tiktok/x")
    parser.add_argument("--output", default="", help="输出 xlsx 路径，不填则自动写入 output/<platform>")
    args = parser.parse_args(argv)
    merge_xlsx_files(args.folder, args.keyword, args.platform, args.output or None)


if __name__ == "__main__":
    main()
