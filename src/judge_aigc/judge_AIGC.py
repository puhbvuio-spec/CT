# -*- coding: utf-8 -*-
"""AIGC 判断与语种识别核心逻辑模块。

本模块通过本地启发式规则（正则匹配与字符统计）和 LLM（大语言模型）结合的方式，
对推文或视频的标题文本进行 AIGC（AI 生成内容）标记判定以及主要语种识别，并输出结果。
"""

import json
import re
import os
import openpyxl
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import TypedDict, List, Any, Dict, Set, Callable

from .config import config
from src.core import interruptible_sleep, wait_if_paused


class State(TypedDict):
    """LangGraph 工作流的状态定义。

    Attributes:
        data: 输入的标题列表，每一项为 [序号, 标题]。
        ai_response: AI 模型返回的原始文本内容。
    """
    data: List[List[Any]]
    ai_response: str


# 本地快速判定 AIGC 内容的关键词库（不区分大小写）
AIGC_KEYWORDS = (
    "ai生成",
    "ai动画",
    "ai视频",
    "ai cover",
    "ai art",
    "ai generated",
    "stable diffusion",
    "midjourney",
    "sora",
    "runway",
    "comfyui",
    "lora",
    "aigc",
    "ai実写化",
    "aiアニメ",
    "ai作成",
    "ai動画",
    "ai実写",
    "aiアレンジ",
    "ai live action",
    "ai animation",
    "aiart",
)

# 系统支持的主要语言选项集合（标准名称）
LANGUAGE_OPTIONS = {
    "中文",
    "英语",
    "日语",
    "韩语",
    "西班牙语",
    "泰语",
    "俄语",
    "阿拉伯语",
    "印地语",
    "越南语",
    "法语",
    "德语",
    "葡萄牙语",
    "印尼语",
    "意大利语",
    "土耳其语",
    "混合语言",
    "未知",
}

# 常见语言别名至标准名称的映射字典
LANGUAGE_ALIASES = {
    "中文": "中文",
    "汉语": "中文",
    "英文": "英语",
    "英语": "英语",
    "日文": "日语",
    "日语": "日语",
    "韩文": "韩语",
    "韩语": "韩语",
    "西语": "西班牙语",
    "西班牙文": "西班牙语",
    "西班牙语": "西班牙语",
    "泰文": "泰语",
    "泰语": "泰语",
    "俄文": "俄语",
    "俄语": "俄语",
    "阿拉伯文": "阿拉伯语",
    "阿拉伯语": "阿拉伯语",
    "印地文": "印地语",
    "印地语": "印地语",
    "越南文": "越南语",
    "越南语": "越南语",
    "法文": "法语",
    "法语": "法语",
    "德文": "德语",
    "德语": "德语",
    "葡语": "葡萄牙语",
    "葡萄牙文": "葡萄牙语",
    "葡萄牙语": "葡萄牙语",
    "印尼文": "印尼语",
    "印尼语": "印尼语",
    "印度尼西亚语": "印尼语",
    "意大利文": "意大利语",
    "意大利语": "意大利语",
    "土耳其文": "土耳其语",
    "土耳其语": "土耳其语",
    "混合": "混合语言",
    "混合语": "混合语言",
    "混合语言": "混合语言",
    "未知": "未知",
}


def start_node(state: State) -> State:
    """LangGraph 工作流的入口节点。

    直接透传状态，不做任何处理。
    """
    return state


def clean_json_text(text: str) -> str:
    """清理 AI 返回的内容，防止其输出带有 ```json 代码块包裹。

    Args:
        text: 待清理的原始文本。

    Returns:
        清理后的 JSON 格式文本字符串。
    """
    text = text.strip()

    if text.startswith("```json"):
        text = text.replace("```json", "", 1).strip()

    if text.startswith("```"):
        text = text.replace("```", "", 1).strip()

    if text.endswith("```"):
        text = text[:-3].strip()

    return text.strip()


def extract_ai_content(response: dict) -> str:
    """从 LangChain agent 返回结果中解析并提取最后一条 AI 消息内容。

    Args:
        response: 代理返回的原始响应字典。

    Returns:
        AI 返回的文本内容。
    """
    messages = response.get("messages", [])

    if not messages:
        raise ValueError("AI 没有返回 messages。")

    last_message = messages[-1]

    if hasattr(last_message, "content"):
        return last_message.content

    if isinstance(last_message, dict):
        return last_message.get("content", "")

    raise TypeError(f"无法识别的消息类型：{type(last_message)}")


def generate_node(state: State) -> State:
    """LangGraph 的生成节点，调用大模型分析标题数据。

    Args:
        state: 当前工作流状态。

    Returns:
        更新了 ai_response 的状态字典。
    """
    data_input = state["data"]

    user_prompt = f"""
请根据下面每条标题，判断标题所描述的视频或推文内容是否是 AI 生成内容，并判断标题文本的主要语言。

输入数据：
{json.dumps(data_input, ensure_ascii=False)}

请严格按照系统提示词要求返回 JSON。
"""

    from .agent_create import get_agent

    response = get_agent().invoke(
        {
            "messages": [
                {
                    "role": "user",
                    "content": user_prompt,
                }
            ]
        }
    )

    ai_content = extract_ai_content(response)
    state["ai_response"] = ai_content

    return state


def end_node(state: State) -> State:
    """LangGraph 工作流的终点节点。

    直接透传状态，不做额外处理。
    """
    return state


def create_graph():
    """编译并生成 LangGraph 的状态图工作流。

    Returns:
        编译后的 StateGraph 实例。
    """
    from langgraph.graph import StateGraph, END

    builder = StateGraph(State)

    builder.add_node("start", start_node)
    builder.add_node("generate", generate_node)
    builder.add_node("end", end_node)

    builder.set_entry_point("start")
    builder.add_edge("start", "generate")
    builder.add_edge("generate", "end")
    builder.add_edge("end", END)

    return builder.compile()


def read_txt(file_path: str, row_limit: int):
    """分批读取包含序号和标题的 TXT 输入文件。

    每行支持以下格式：
    - "1   视频标题" (使用空格分隔)
    - "2\t视频标题" (使用 Tab 分隔)
    - 或者是被自动合并到上一行的多行标题内容。

    Args:
        file_path: 输入 TXT 文件的绝对路径。
        row_limit: 每一批（Batch）的最大行数限制。

    Yields:
        List[List[Any]]: 一个批次的数据，每项格式为 [序号(int), 标题(str)]。
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"找不到输入文件：{file_path}")

    with open(file_path, "r", encoding="utf-8") as f:
        alist = []
        last_item = None

        for line in f:
            line = line.strip()

            if not line:
                continue

            # 使用正则匹配行首的数字序号，后续跟空格/Tab，再跟标题文本
            match = re.match(r"^(\d+)[ \t]+(.+)$", line)

            if not match:
                # 若无法解析出序号，则将当前行追加入上一个解析出的标题中（处理换行折行标题）
                if last_item is not None:
                    last_item[1] = f"{last_item[1]} {line}"
                else:
                    print(f"跳过无法解析的行：{line}")
                continue

            number = int(match.group(1))
            title = match.group(2).strip()

            if len(alist) >= row_limit:
                yield alist
                alist = []

            last_item = [number, title]
            alist.append(last_item)

        if alist:
            yield alist


def parse_ai_json(json_text: str, batch_index: int):
    """解析 AI 模型返回的 JSON 文本数据，失败时保存问题内容以供调试。

    Args:
        json_text: 模型返回的包含 JSON 的文本。
        batch_index: 当前处理的批次序号。

    Returns:
        List[Dict/List]: 反序列化后的 JSON 数组数据。
    """
    json_text = clean_json_text(json_text)

    try:
        data = json.loads(json_text)
    except json.JSONDecodeError as e:
        # 当 JSON 解析失败，将错误内容保存到当前目录，防止数据丢失且便于排查
        bad_file = f"bad_response_batch_{batch_index}.txt"

        try:
            with open(bad_file, "w", encoding="utf-8") as f:
                f.write(json_text)
        except OSError:
            print(f"无法写入错误文件，原始返回：{json_text[:500]}...")

        print("=" * 80)
        print(f"第 {batch_index} 批 JSON 解析失败。")
        print(f"错误信息：{e}")
        print(f"AI 原始返回已保存到：{bad_file}")
        print("建议把 ROW_LIMIT 再调小，比如 50。")
        print("=" * 80)

        raise

    if not isinstance(data, list):
        raise TypeError("AI 返回结果不是 JSON 数组。")

    return data


def detect_aigc_locally(title: str) -> str:
    """本地进行快速的 AIGC 关键词匹配判定。

    不区分大小写匹配内置的 AIGC_KEYWORDS 关键词列表。

    Args:
        title: 视频或推文标题。

    Returns:
        "是" 或 "否"。
    """
    title_lower = title.lower()
    return "是" if any(keyword.lower() in title_lower for keyword in AIGC_KEYWORDS) else "否"


def normalize_language(language: Any) -> str:
    """将 AI 返回的语言别名归一化为标准的系统语种名称。

    Args:
        language: AI 返回的任意格式的语言标识符。

    Returns:
        str: 归一化后的标准语种名称（如 "中文", "英语" 等）。
    """
    return LANGUAGE_ALIASES.get(str(language).strip(), str(language).strip())


def detect_language_locally(title: str) -> str:
    """本地启发式语种识别算法。

    通过特定语种的特殊变音符号（Markers）及独有高频虚词（Words）判定拉丁字母系语种；
    通过 Unicode 字符集编码范围判定中日韩、泰、俄（西里尔）、阿拉伯、印地等非拉丁系语种；
    如果多种语种并存且权重相当，判定为混合语言。

    Args:
        title: 待识别的标题文本。

    Returns:
        str: 标准的语种名称。
    """
    # 1. 各拉丁目系语种的特异性标记及高频特征词定义
    spanish_markers = ("ñ", "á", "é", "í", "ó", "ú", "¿", "¡")
    spanish_words = (" el ", " la ", " los ", " las ", " del ", " que ", " una ", " por ")
    vietnamese_markers = ("ă", "đ", "ơ", "ư")
    vietnamese_words = (" xin ", " chào ", " và ", " của ", " với ", " cho ", " không ", " trong ")
    french_markers = ("à", "â", "æ", "ç", "é", "è", "ê", "ë", "î", "ï", "ô", "œ", "ù", "û", "ü", "ÿ")
    french_words = (" bonjour ", " le ", " les ", " des ", " avec ", " pour ", " une ")
    german_markers = ("ä", "ö", "ü", "ß")
    german_words = (" der ", " die ", " das ", " und ", " mit ", " nicht ")
    portuguese_markers = ("ã", "õ", "á", "â", "ê", "í", "ó", "ô", "ú", "ç")
    portuguese_words = (" você ", " uma ", " para ", " com ", " dos ", " das ", " não ")
    italian_words = (" il ", " gli ", " della ", " delle ", " che ", " per ")
    indonesian_words = (" yang ", " dan ", " dengan ", " untuk ", " ini ", " dari ")
    turkish_markers = ("ğ", "ı", "İ", "ş", "ç", "ö", "ü")
    turkish_words = (" bir ", " ve ", " için ", " ile ", " değil ")

    # 2. 通过 Unicode 编码区间统计各类特殊字符个数
    # 平假名/片假名区间
    hira_kata = sum(
        1
        for char in title
        if "\u3041" <= char <= "\u30ff" or "\u31f0" <= char <= "\u31ff"
    )
    # 朝鲜谚文（韩语文字）区间
    hangul = sum(1 for char in title if "\uac00" <= char <= "\ud7af")
    # 泰文字符区间
    thai = sum(1 for char in title if "\u0e00" <= char <= "\u0e7f")
    # 西里尔字母（俄语）区间
    cyrillic = sum(1 for char in title if "\u0400" <= char <= "\u04ff")
    # 阿拉伯文字区间
    arabic = sum(1 for char in title if "\u0600" <= char <= "\u06ff")
    # 天城文（印地语）区间
    devanagari = sum(1 for char in title if "\u0900" <= char <= "\u097f")
    # 汉字（CJK 统一汉字）区间
    cjk = sum(1 for char in title if "\u4e00" <= char <= "\u9fff")
    # 拉丁字母（ASCII 纯字母）
    latin = sum(1 for char in title if char.isascii() and char.isalpha())

    # 为词汇边界匹配首尾补空格
    lowered = f" {title.lower()} "
    detected = []

    # 3. 统计并添加非拉丁目系语种的得分候选项
    if hangul:
        detected.append(("韩语", hangul))

    if thai:
        detected.append(("泰语", thai))

    if cyrillic:
        detected.append(("俄语", cyrillic))

    if arabic:
        detected.append(("阿拉伯语", arabic))

    if devanagari:
        detected.append(("印地语", devanagari))

    if hira_kata:
        # 日语包含平假名/片假名，其汉字部分（CJK）一同累计作为日语权重
        detected.append(("日语", hira_kata + cjk))
    elif cjk:
        detected.append(("中文", cjk))

    # 4. 统计并计算拉丁目系语种的启发式得分
    latin_candidates = [
        ("越南语", vietnamese_markers, vietnamese_words),
        ("德语", german_markers, german_words),
        ("土耳其语", turkish_markers, turkish_words),
        ("葡萄牙语", portuguese_markers, portuguese_words),
        ("法语", french_markers, french_words),
        ("西班牙语", spanish_markers, spanish_words),
        ("意大利语", (), italian_words),
        ("印尼语", (), indonesian_words),
    ]
    latin_detected = []

    for language, markers, words in latin_candidates:
        # 每个变音特殊符号得 2 分
        marker_score = sum(2 for marker in markers if marker in lowered)
        # 每个特征高频词得 1 分
        word_score = sum(1 for word in words if word in lowered)
        score = marker_score + word_score
        if score:
            # 候选得分取拉丁字母总数与特征得分的最大值，确保特异识别
            latin_detected.append((language, max(latin, score)))

    if latin_detected:
        latin_detected.sort(key=lambda item: item[1], reverse=True)
        detected.append(latin_detected[0])
    elif latin >= 2:
        detected.append(("英语", latin))

    if not detected:
        return "未知"

    detected.sort(key=lambda item: item[1], reverse=True)

    if len(detected) > 1:
        top_language, top_score = detected[0]
        second_score = detected[1][1]
        if top_score >= 4 and second_score >= 4 and second_score / top_score >= 0.8:
            return "混合语言"
        return top_language

    return detected[0][0]


def classify_locally(lines: List[List[Any]]):
    rows = []
    unresolved = []

    for number, title in lines:
        language = detect_language_locally(title)
        is_aigc = detect_aigc_locally(title)
        row = [number, title, is_aigc, language]
        rows.append(row)

        should_ask_ai = language == "未知" or (
            is_aigc == "否" and not config.TRUST_LOCAL_NEGATIVE_AIGC
        )

        if should_ask_ai:
            unresolved.append([number, title])

    return rows, unresolved


def validate_result_rows(
    input_rows: List[List[Any]],
    result_rows: List[List[Any]],
    batch_index: int,
):
    """校验 AI 模型或本地处理返回的行数据，确保序号和标题完全对应，数据项符合规范。

    Args:
        input_rows: 送入处理的原始行列表，结构为 [[序号, 标题], ...]。
        result_rows: 处理返回的结果行列表，结构为 [[序号, 标题, AIGC判定, 语种], ...]。
        batch_index: 当前处理的批次标识符。

    Returns:
        List[List[Any]]: 校验并规范化之后的行数据列表。
    """
    if len(result_rows) != len(input_rows):
        raise ValueError(
            f"第 {batch_index} 批返回数量不一致：输入 {len(input_rows)} 行，返回 {len(result_rows)} 行"
        )

    validated = []

    for expected, actual in zip(input_rows, result_rows):
        if not isinstance(actual, list) or len(actual) != 4:
            raise ValueError(f"第 {batch_index} 批返回行格式异常：{actual}")

        expected_number, expected_title = expected
        actual_number, actual_title, is_aigc, language = actual
        is_aigc = str(is_aigc).strip()
        language = normalize_language(language)

        try:
            actual_number = int(actual_number)
        except (TypeError, ValueError) as exc:
            raise ValueError(f"第 {batch_index} 批序号不是数字：{actual}") from exc

        # 校验序号顺序和标题文本是否完全吻合，避免 LLM 幻觉产生错位
        if actual_number != expected_number:
            raise ValueError(
                f"第 {batch_index} 批序号顺序不一致：期望 {expected_number}，实际 {actual_number}"
            )

        if str(actual_title).strip() != str(expected_title).strip():
            raise ValueError(f"第 {batch_index} 批标题不一致：序号 {expected_number}")

        if is_aigc not in {"是", "否"}:
            raise ValueError(f"第 {batch_index} 批 AIGC 值异常：{actual}")

        if language not in LANGUAGE_OPTIONS:
            raise ValueError(f"第 {batch_index} 批语言值异常：{actual}")

        validated.append([actual_number, actual_title, is_aigc, language])

    return validated


def merge_ai_rows(local_rows: List[List[Any]], ai_rows: List[List[Any]]):
    """将大模型判定后的疑难结果行合并覆盖回本地预分类结果中。

    Args:
        local_rows: 本地快速判定得出的初始结果行列表。
        ai_rows: 调用大模型补充判定得出的结果行列表。

    Returns:
        List[List[Any]]: 合并更新后的最终结果行列表。
    """
    # 建立序号对行的映射，方便快速覆盖
    ai_by_number: Dict[int, List[Any]] = {int(row[0]): row for row in ai_rows}
    merged = []

    for row in local_rows:
        number = int(row[0])
        # 如果 AI 判定的映射中有当前序号，则使用 AI 判定的数据覆盖，否则保留本地预判定
        merged.append(ai_by_number.get(number, row))

    return merged


def get_existing_numbers(ws) -> Set[int]:
    """读取已存在的 Excel 列表，提取所有已处理的序号集合，用于断点续传。

    Args:
        ws: openpyxl Worksheet 实例。

    Returns:
        Set[int]: 所有已写入过的序号集合。
    """
    numbers = set()

    # 从第二行开始逐行扫描首列的序号
    for row in ws.iter_rows(min_row=2, values_only=True):
        number = row[0]
        if number is None:
            continue

        try:
            numbers.add(int(number))
        except (TypeError, ValueError):
            continue

    return numbers


def create_or_load_excel(save_path: str):
    """创建或加载 Excel 工作簿。若已存在则直接加载，若不存在则新建并初始化表头。

    Args:
        save_path: 目标 Excel 保存路径。

    Returns:
        tuple[Workbook, Worksheet]: openpyxl Workbook 与活动的 Worksheet 对象。
    """
    output_dir = os.path.dirname(os.path.abspath(save_path))
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    if os.path.exists(save_path):
        wb = openpyxl.load_workbook(save_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = config.SHEET_NAME
        ws.append(config.HEADERS)

    return wb, ws


def split_evenly(values: List[List[Any]], group_count: int):
    """将列表数据尽可能均匀地切分成指定组数，用于线程池多任务并发。

    Args:
        values: 待切分的行列表。
        group_count: 期望切分成的组数。

    Returns:
        List[List[List[Any]]]: 切分后的二维列表。
    """
    group_count = max(1, min(group_count, len(values)))
    chunk_size = (len(values) + group_count - 1) // group_count
    return [values[index:index + chunk_size] for index in range(0, len(values), chunk_size)]


def ask_ai_for_lines(lines: List[List[Any]], batch_index: Any):
    """构建 LangGraph 状态并调用 AI 代理对指定行进行识别。

    Args:
        lines: 待识别的标题子集，格式为 [[序号, 标题], ...]。
        batch_index: 批次索引标识（可能是浮点表示的子块如 1.2）。

    Returns:
        List[List[Any]]: 大模型分析并完成严格校验后的行列表。
    """
    state: State = {
        "data": lines,
        "ai_response": "",
    }
    graph = create_graph()
    result_state = graph.invoke(state)
    ai_response = parse_ai_json(
        result_state["ai_response"],
        batch_index=batch_index,
    )
    return validate_result_rows(
        lines,
        ai_response,
        batch_index=batch_index,
    )


def process_batch(batch_index: int, lines: List[List[Any]], max_workers: int = 1):
    """处理单批次数据：本地预分类 -> 提取疑难行 -> 并发调用 AI -> 合并结果并校验。

    Args:
        batch_index: 批次索引。
        lines: 本批原始行数据。
        max_workers: AI 并发并发数（即线程池最大工作线程数）。

    Returns:
        tuple[List[List[Any]], int]: 最终的判定数据列表，以及调用大模型的行数。
    """
    local_rows, unresolved_lines = classify_locally(lines)
    ai_response = []

    if unresolved_lines:
        worker_count = max(1, min(int(max_workers or 1), len(unresolved_lines)))
        if worker_count == 1:
            # 单线程调用，直接请求 AI
            ai_response = ask_ai_for_lines(unresolved_lines, batch_index)
        else:
            # 多线程并发：将疑难行分割成多组，并发请求以提升接口响应效率
            chunks = split_evenly(unresolved_lines, worker_count)
            indexed_results: Dict[int, List[List[Any]]] = {}
            with ThreadPoolExecutor(max_workers=worker_count) as executor:
                futures = {
                    executor.submit(ask_ai_for_lines, chunk, f"{batch_index}.{chunk_index}"): chunk_index
                    for chunk_index, chunk in enumerate(chunks, start=1)
                }
                for future in as_completed(futures):
                    try:
                        # 限制单任务最大超时为 300 秒，防挂死
                        result = future.result(timeout=300)
                    except TimeoutError:
                        import logging
                        logging.getLogger(__name__).error("Batch chunk timed out after 300s")
                        result = None
                    indexed_results[futures[future]] = result
            # 按切分顺序重新组装多线程结果
            for chunk_index in sorted(indexed_results):
                chunk_result = indexed_results[chunk_index]
                if chunk_result is not None:
                    ai_response.extend(chunk_result)

    result_rows = merge_ai_rows(local_rows, ai_response)
    return validate_result_rows(
        lines,
        result_rows,
        batch_index=batch_index,
    ), len(unresolved_lines)


def run_judge(
    input_txt_path: str | None = None,
    output_excel_path: str | None = None,
    row_limit: int | None = None,
    max_workers: int = 1,
    save_every_batches: int | None = None,
    log_callback: Callable[[str], None] | None = None,
    stop_event=None,
    pause_event=None,
):
    """运行 AIGC 判定与语种识别核心任务的主驱动函数。

    分批读取文本，比对已处理结果实现增量断点续答，调度本地与并发 AI 计算，
    定期刷盘保存到 Excel 文件中。

    Args:
        input_txt_path: 输入 TXT 路径，若空则使用配置默认值。
        output_excel_path: 输出 Excel 路径，若空则使用配置默认值。
        row_limit: 每批处理的标题条数。
        max_workers: 并发调用大模型的最大并发线程数。
        save_every_batches: 每处理多少批次执行一次文件保存，防断电丢失。
        log_callback: 日志记录回调函数。
        stop_event: 任务停止事件（线程安全）。
        pause_event: 任务暂停事件（线程安全）。

    Returns:
        str: 最终保存的 Excel 物理路径。
    """
    input_txt_path = input_txt_path or config.INPUT_TXT_PATH
    output_excel_path = output_excel_path or config.OUTPUT_EXCEL_PATH
    row_limit = max(1, int(row_limit or config.ROW_LIMIT))
    max_workers = max(1, int(max_workers or 1))
    default_save_every = getattr(config, "SAVE_EVERY_BATCHES", 1)
    save_every_batches = max(1, int(save_every_batches or default_save_every or 1))
    log = log_callback or print

    log("判断 AIGC 及主要语言中...")
    log(f"输入文件：{input_txt_path}")
    log(f"输出文件：{output_excel_path}")
    log(f"每批行数：{row_limit}，当前批 AI 并发数：{max_workers}")
    log(f"保存频率：每 {save_every_batches} 批保存一次")

    wb, ws = create_or_load_excel(output_excel_path)
    # 获取 Excel 中已存在的序号，用于增量断点续传跳过
    existing_numbers = get_existing_numbers(ws)

    total_count = 0
    written_count = 0
    batch_count = 0
    batches_since_save = 0
    def stopped() -> bool:
        return bool(stop_event and stop_event.is_set())

    def write_rows(result_rows: List[List[Any]]) -> int:
        nonlocal written_count
        batch_written = 0
        for row in result_rows:
            number = int(row[0])
            if number in existing_numbers:
                continue
            ws.append(row)
            existing_numbers.add(number)
            written_count += 1
            batch_written += 1
        return batch_written

    try:
        for lines in read_txt(file_path=input_txt_path, row_limit=row_limit):
            if stopped():
                log("任务已停止。")
                break
            if wait_if_paused(pause_event, stop_event):
                break

            batch_count += 1
            total_count += len(lines)
            # 过滤掉当前批次中已经在 Excel 中存在的数据行
            pending_lines = [line for line in lines if int(line[0]) not in existing_numbers]

            if not pending_lines:
                log(f"第 {batch_count} 批已全部存在，跳过 {len(lines)} 行。")
                continue

            _, unresolved_lines = classify_locally(pending_lines)
            log(
                f"正在处理第 {batch_count} 批：本批 {len(lines)} 行，"
                f"待写入 {len(pending_lines)} 行，需调用 AI {len(unresolved_lines)} 行。"
            )

            try:
                result_rows, unresolved_count = process_batch(batch_count, pending_lines, max_workers=max_workers)
            except Exception as exc:
                wb.save(output_excel_path)
                raise RuntimeError(f"第 {batch_count} 批处理失败：{exc}") from exc

            batch_written = write_rows(result_rows)
            batches_since_save += 1
            saved_this_batch = False
            # 定时定量写入磁盘，降低频繁 IO 损耗并保证安全性
            if batches_since_save >= save_every_batches:
                wb.save(output_excel_path)
                batches_since_save = 0
                saved_this_batch = True

            log(
                f"第 {batch_count} 批完成：本批 {len(lines)} 行，"
                f"调用 AI {unresolved_count} 行，写入 {batch_written} 行"
                f"{'，已保存。' if saved_this_batch else '。'}"
            )

            if config.SLEEP_SECONDS > 0:
                interruptible_sleep(config.SLEEP_SECONDS, stop_event)

        wb.save(output_excel_path)
    except Exception:
        wb.save(output_excel_path)
        raise

    log(f"全部处理完成：读取 {total_count} 行，本次运行写入 {written_count} 行。")
    log(f"已保存到：{output_excel_path}")
    return output_excel_path


def run_judge_legacy():
    """向后兼容的旧版运行接口。"""
    return run_judge()
