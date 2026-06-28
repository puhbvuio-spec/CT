#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""关键词可观察搜索覆盖实验工具。"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import json
import logging
import os
import re
import sys
import time
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from urllib.parse import parse_qs, urlparse

import openpyxl

try:
    from src.platforms.tiktok.keyword import run_tiktok_spider
except ModuleNotFoundError:
    run_tiktok_spider = None

try:
    from src.platforms.x_twitter.keyword import run_x_spider
except ModuleNotFoundError:
    run_x_spider = None

try:
    from src.platforms.youtube.keyword import run_youtube_spider
except ModuleNotFoundError:
    run_youtube_spider = None
from src.version import __version__

VALID_PLATFORMS = ("youtube", "tiktok", "x_twitter")
SUCCESSFUL_RUN_STATUSES = {"SUCCESS", "EMPTY_RESULT"}
REPORT_FAILURE_STATUS = "BASELINE_FAILED"

STATUS_SUCCESS = "SUCCESS"
STATUS_EMPTY_RESULT = "EMPTY_RESULT"
STATUS_AUTH_REQUIRED = "AUTH_REQUIRED"
STATUS_CAPTCHA_OR_RISK = "CAPTCHA_OR_RISK"
STATUS_QUOTA_EXCEEDED = "QUOTA_EXCEEDED"
STATUS_TIMEOUT = "TIMEOUT"
STATUS_OUTPUT_SCHEMA_ERROR = "OUTPUT_SCHEMA_ERROR"
STATUS_UNKNOWN_PLATFORM = "UNKNOWN_PLATFORM"
STATUS_FAILED = "FAILED"

REPORT_INTRO_LINES = [
    "# 关键词可观察搜索覆盖实验报告",
    "",
    "## 口径说明",
    "",
    "本报告评估的是在指定平台、指定账号或环境、指定时间窗口、指定搜索入口、指定排序方式和指定采集深度下，不同关键词组对可观察搜索结果的召回、重叠和增量影响。",
    "",
    "本报告不代表平台全量内容覆盖率。",
]

CSV_HEADERS = [
    "Game",
    "Platform",
    "Baseline Query",
    "Baseline Status",
    "Baseline Unique ID Count",
    "Baseline Raw Link Count",
    "Keyword Group",
    "Group Status",
    "Result Count",
    "Raw Link Count",
    "Baseline Intersection Count",
    "Relative Result Volume (%)",
    "Baseline Overlap Rate (%)",
    "Unique Result Count",
    "Incremental Gain (%)",
    "Jaccard Similarity (%)",
    "Error Message",
]

MARKDOWN_TABLE_HEADERS = [
    "Group Index",
    "Keyword Combination",
    "Status",
    "Result Count",
    "Raw Link Count",
    "Baseline Intersection Count",
    "Relative Result Volume",
    "Baseline Overlap Rate",
    "Unique Result Count",
    "Incremental Gain",
    "Jaccard Similarity",
    "Error Message",
]

_LINE_BREAK_RE = re.compile(r"[\r\n\u2028\u2029]+")


@dataclass
class SpiderRunResult:
    platform: str
    keyword: str
    status: str
    ids: set[str]
    links: set[str]
    output_path: str | None
    error_message: str | None
    started_at: str
    finished_at: str
    scanned_count: int | None = None
    written_count: int | None = None
    hit_limit: bool = False

    def to_snapshot(self) -> dict[str, Any]:
        return {
            "platform": self.platform,
            "keyword": self.keyword,
            "status": self.status,
            "ids": sorted(self.ids),
            "links": sorted(self.links),
            "output_path": self.output_path,
            "error_message": self.error_message,
            "started_at": self.started_at,
            "finished_at": self.finished_at,
            "scanned_count": self.scanned_count,
            "written_count": self.written_count,
            "hit_limit": self.hit_limit,
        }


def now_str() -> str:
    return dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def should_stop(stop_event=None) -> bool:
    return bool(stop_event and stop_event.is_set())


def wait_if_paused(pause_event=None, stop_event=None) -> bool:
    while pause_event and pause_event.is_set():
        if should_stop(stop_event):
            return True
        time.sleep(0.1)
    return should_stop(stop_event)


def sanitize_csv_cell(value: Any) -> Any:
    if value is None:
        return ""
    if not isinstance(value, str):
        return value
    return _LINE_BREAK_RE.sub(" ", value).strip()


def workspace_root() -> Path:
    return Path(__file__).resolve().parents[2]


def parse_platforms(platforms_cfg: Any) -> list[str]:
    if platforms_cfg is None:
        return list(VALID_PLATFORMS)

    if isinstance(platforms_cfg, list):
        raw_platforms = [item for item in platforms_cfg if isinstance(item, str)]
    elif isinstance(platforms_cfg, str):
        raw_platforms = platforms_cfg.split(",")
    else:
        raw_platforms = []

    normalized: list[str] = []
    seen: set[str] = set()
    for raw in raw_platforms:
        name = raw.strip().lower()
        if not name or name in seen:
            continue
        seen.add(name)
        normalized.append(name)

    return normalized or list(VALID_PLATFORMS)


def invalid_platforms(platforms: list[str]) -> list[str]:
    return [platform for platform in platforms if platform not in VALID_PLATFORMS]


def parse_keyword_groups_text(raw_text: str) -> list[list[str]]:
    keyword_groups: list[list[str]] = []
    for raw_line in (raw_text or "").splitlines():
        line = raw_line.strip()
        if not line:
            continue
        keywords = [item.strip() for item in re.split(r"[,，]", line) if item.strip()]
        if keywords:
            keyword_groups.append(keywords)
    return keyword_groups


def format_keyword_groups_text(keyword_groups: list[list[str]]) -> str:
    lines: list[str] = []
    for group in keyword_groups or []:
        keywords = [str(keyword).strip() for keyword in group if str(keyword).strip()]
        if keywords:
            lines.append(", ".join(keywords))
    return "\n".join(lines)


def normalize_games_config(games: list[dict[str, Any]]) -> list[dict[str, Any]]:
    normalized_games: list[dict[str, Any]] = []
    for index, game in enumerate(games, 1):
        if not isinstance(game, dict):
            raise ValueError(f"第 {index} 个游戏配置必须是对象。")

        name = str(game.get("name", "")).strip()
        baseline_query = str(game.get("baseline_query", "")).strip()
        raw_groups = game.get("keyword_groups", [])

        if not name:
            raise ValueError(f"第 {index} 个游戏缺少名称。")
        if not baseline_query:
            raise ValueError(f"第 {index} 个游戏缺少基准词。")
        if not isinstance(raw_groups, list) or not raw_groups:
            raise ValueError(f"第 {index} 个游戏至少需要一个测试词组。")

        keyword_groups: list[list[str]] = []
        for group_index, group in enumerate(raw_groups, 1):
            if not isinstance(group, list):
                raise ValueError(f"第 {index} 个游戏的第 {group_index} 个词组必须是数组。")
            keywords = [str(keyword).strip() for keyword in group if str(keyword).strip()]
            if not keywords:
                raise ValueError(f"第 {index} 个游戏的第 {group_index} 个词组不能为空。")
            keyword_groups.append(keywords)

        normalized_games.append(
            {
                "name": name,
                "baseline_query": baseline_query,
                "keyword_groups": keyword_groups,
            }
        )

    if not normalized_games:
        raise ValueError("请至少配置一个游戏。")
    return normalized_games


def parse_games_definition(raw_definition: str) -> list[dict[str, Any]]:
    text = (raw_definition or "").strip()
    if not text:
        raise ValueError("请至少配置一个游戏。")

    if text.startswith("["):
        try:
            payload = json.loads(text)
        except json.JSONDecodeError as exc:
            raise ValueError(f"JSON 游戏配置解析失败: {exc}") from exc
        if not isinstance(payload, list):
            raise ValueError("JSON 游戏配置必须是数组。")
        return normalize_games_config(payload)

    blocks = re.split(r"\n\s*\n+", text)
    games: list[dict[str, Any]] = []
    for block_index, block in enumerate(blocks, 1):
        lines = []
        for raw_line in block.splitlines():
            line = raw_line.strip()
            if not line or line.startswith("#"):
                continue
            lines.append(line)

        if not lines:
            continue

        header = lines[0]
        if "|" not in header:
            raise ValueError(f"第 {block_index} 个游戏配置的首行必须写成“游戏名称 | 基准词”。")

        name, baseline_query = [part.strip() for part in header.split("|", 1)]
        if not name or not baseline_query:
            raise ValueError(f"第 {block_index} 个游戏配置的名称和基准词都不能为空。")

        keyword_groups = parse_keyword_groups_text("\n".join(lines[1:]))

        games.append(
            {
                "name": name,
                "baseline_query": baseline_query,
                "keyword_groups": keyword_groups,
            }
        )

    return normalize_games_config(games)


def extract_id_from_link(link: str, platform: str) -> str:
    if not link or not isinstance(link, str):
        return ""

    link = link.strip()
    try:
        if platform == "youtube":
            match = re.search(r"(?:v=|/shorts/|/embed/|/live/|/v/|youtu\.be/)([a-zA-Z0-9_-]{11})", link)
            if match:
                return match.group(1)
            parsed = urlparse(link)
            if parsed.netloc and parsed.query:
                qs = parse_qs(parsed.query)
                if qs.get("v"):
                    return qs["v"][0]
            path_parts = [part for part in parsed.path.split("/") if part]
            if path_parts:
                return path_parts[-1]

        if platform == "tiktok":
            match = re.search(r"/video/(\d+)", link)
            if match:
                return match.group(1)
            match = re.search(r"/v/(\d+)(?:\.html)?", link)
            if match:
                return match.group(1)
            parsed = urlparse(link)
            path_parts = [part for part in parsed.path.split("/") if part]
            if path_parts:
                return path_parts[-1]

        if platform == "x_twitter":
            match = re.search(r"/status/(\d+)", link)
            if match:
                return match.group(1)
            parsed = urlparse(link)
            path_parts = [part for part in parsed.path.split("/") if part]
            if path_parts:
                return path_parts[-1]
    except Exception:
        return ""

    return link


def load_config(config_path: str) -> dict:
    if not os.path.exists(config_path):
        raise FileNotFoundError(f"Configuration file not found: {config_path}")

    with open(config_path, "r", encoding="utf-8") as file:
        config = json.load(file)

    if "games" not in config or not isinstance(config["games"], list):
        raise ValueError("Configuration must contain a 'games' list.")

    return config


def extract_links_from_excel(file_path: str, platform: str) -> set[str]:
    links: set[str] = set()
    if not file_path or not os.path.exists(file_path):
        return links

    workbook = None
    try:
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        sheet_name = None
        if platform == "x_twitter":
            if "数据" in workbook.sheetnames:
                sheet_name = "数据"
            elif "推文信息" in workbook.sheetnames:
                sheet_name = "推文信息"
        else:
            if "视频信息" in workbook.sheetnames:
                sheet_name = "视频信息"
            elif "数据" in workbook.sheetnames:
                sheet_name = "数据"

        sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook.active
        target_col = "推文链接" if platform == "x_twitter" else "视频链接"

        try:
            headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        except StopIteration:
            headers = []

        if target_col not in headers:
            return links

        col_idx = headers.index(target_col) + 1
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) < col_idx:
                continue
            value = row[col_idx - 1]
            if value is None:
                continue
            text = str(value).strip()
            if text:
                links.add(text)
    except Exception as exc:
        logging.exception("读取 Excel 文件失败 (%s, platform=%s): %s", file_path, platform, exc)
        raise
    finally:
        if workbook is not None:
            workbook.close()

    return links


def classify_error_status(message: str | None) -> str:
    lowered = (message or "").lower()
    if any(token in lowered for token in ("quota", "rate limit", "too many requests")):
        return STATUS_QUOTA_EXCEEDED
    if any(token in lowered for token in ("captcha", "verify", "risk", "风控")):
        return STATUS_CAPTCHA_OR_RISK
    if any(token in lowered for token in ("login", "sign in", "signin", "auth", "unauthorized", "forbidden", "permission")):
        return STATUS_AUTH_REQUIRED
    if any(token in lowered for token in ("timeout", "timed out", "超时")):
        return STATUS_TIMEOUT
    return STATUS_FAILED


def is_successful_run(result: SpiderRunResult) -> bool:
    return result.status in SUCCESSFUL_RUN_STATUSES


def pct(numerator: int, denominator: int) -> float | None:
    if denominator <= 0:
        return None
    return round((numerator / denominator) * 100.0, 2)


def format_percent(value: float | None) -> str:
    return "" if value is None else f"{value:.2f}%"


def format_csv_percent(value: float | None) -> str | float:
    return "" if value is None else value


def calculate_group_metrics(group_ids: set[str], baseline_ids: set[str]) -> dict[str, Any]:
    union_ids = group_ids | baseline_ids
    baseline_intersection_count = len(group_ids & baseline_ids)
    unique_result_count = len(group_ids - baseline_ids)
    return {
        "result_count": len(group_ids),
        "baseline_intersection_count": baseline_intersection_count,
        "relative_result_volume": pct(len(group_ids), len(baseline_ids)),
        "baseline_overlap_rate": pct(baseline_intersection_count, len(baseline_ids)),
        "unique_result_count": unique_result_count,
        "incremental_gain": pct(unique_result_count, len(union_ids)),
        "jaccard_similarity": pct(baseline_intersection_count, len(union_ids)),
    }


def resolve_output_base(output_path: str, log_callback=None) -> Path:
    if not output_path:
        base_path = workspace_root() / "output"
    else:
        raw_path = Path(output_path)
        if not raw_path.is_absolute():
            raw_path = workspace_root() / raw_path
        if raw_path.suffix:
            if log_callback:
                log_callback(f"Legacy report path detected, using parent directory as output root: {raw_path.parent}")
            base_path = raw_path.parent
        else:
            base_path = raw_path

    if base_path.name != "calibration":
        base_path = base_path / "calibration"
    return base_path


def create_run_directory(output_path: str, log_callback=None) -> tuple[str, Path]:
    base_path = resolve_output_base(output_path, log_callback=log_callback)
    run_id = dt.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    run_dir = base_path / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    (run_dir / "raw").mkdir(exist_ok=True)
    (run_dir / "reports").mkdir(exist_ok=True)
    return run_id, run_dir


def write_json(path: Path, payload: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def build_environment_snapshot(
    *,
    run_id: str,
    run_started_at: str,
    run_finished_at: str,
    start_date: str,
    end_date: str,
    days: int,
    platforms: list[str],
    x_search_tab: str,
) -> dict[str, Any]:
    return {
        "run_id": run_id,
        "tool_version": __version__,
        "run_started_at": run_started_at,
        "run_finished_at": run_finished_at,
        "time_window": {
            "start_date": start_date,
            "end_date": end_date,
            "days": days,
        },
        "platforms": platforms,
        "x_search_tab": x_search_tab,
    }


def raw_platform_dir_name(platform: str) -> str:
    normalized = re.sub(r"[^a-z0-9_-]+", "_", (platform or "").strip().lower()).strip("_")
    return normalized or "unknown_platform"


def build_group_snapshot(
    *,
    platform: str,
    group_key: str,
    keywords: list[str],
    status: str,
    ids: set[str],
    links: set[str],
    error_message: str,
    started_at: str | None,
    finished_at: str | None,
    scanned_count: int | None,
    written_count: int | None,
    hit_limit: bool,
    output_paths: list[str],
    keyword_runs: list[SpiderRunResult] | None = None,
) -> dict[str, Any]:
    payload = {
        "platform": platform,
        "group_key": group_key,
        "keywords": keywords,
        "status": status,
        "ids": sorted(ids),
        "links": sorted(links),
        "output_path": output_paths[0] if output_paths else None,
        "output_paths": output_paths,
        "error_message": error_message,
        "started_at": started_at,
        "finished_at": finished_at,
        "scanned_count": scanned_count,
        "written_count": written_count,
        "hit_limit": hit_limit,
    }
    if keyword_runs is not None:
        payload["keyword_runs"] = [run.to_snapshot() for run in keyword_runs]
    return payload


def derive_group_status(keyword_runs: list[SpiderRunResult], group_ids: set[str]) -> tuple[str, str]:
    failures = [run for run in keyword_runs if run.status not in SUCCESSFUL_RUN_STATUSES]
    if failures:
        messages = [f"Keyword '{run.keyword}' failed: {run.error_message or run.status}" for run in failures]
        return STATUS_FAILED, "; ".join(messages)
    if group_ids:
        return STATUS_SUCCESS, ""
    if any(run.status == STATUS_EMPTY_RESULT for run in keyword_runs):
        return STATUS_EMPTY_RESULT, ""
    return STATUS_FAILED, ""


def summarize_counts(runs: list[SpiderRunResult], group_links: set[str]) -> tuple[int | None, int | None, bool]:
    scanned_values = [run.scanned_count for run in runs if run.scanned_count is not None]
    written_values = [run.written_count for run in runs if run.written_count is not None]
    return (
        sum(scanned_values) if scanned_values else None,
        sum(written_values) if written_values else len(group_links),
        any(run.hit_limit for run in runs),
    )


def select_x_search_tab(platform_config: dict[str, Any]) -> str:
    search_tab = str(platform_config.get("x_search_tab", "latest")).strip().lower()
    if search_tab not in {"latest", "top"}:
        return "latest"
    return search_tab


def run_platform_spider(
    platform: str,
    keyword: str,
    start_date: str,
    end_date: str,
    platform_config: dict[str, Any],
    days: int,
    stop_event=None,
    pause_event=None,
) -> SpiderRunResult:
    retrieved_path: str | None = None
    started_at = now_str()

    def finish_callback(path):
        nonlocal retrieved_path
        retrieved_path = path

    def log_callback(message: str):
        logging.debug("[calibration] %s", message)

    try:
        if platform == "youtube":
            if run_youtube_spider is None:
                raise ModuleNotFoundError("google-api-python-client is required for YouTube keyword scraping")
            run_youtube_spider(
                api_keys=platform_config.get("api_keys", []),
                keywords_list=[keyword],
                max_results=platform_config.get("max_results", 10),
                limit_time_str="是",
                start_date=start_date,
                end_date=end_date,
                get_comments_str="否",
                max_comments=0,
                log_callback=log_callback,
                finish_callback=finish_callback,
                stop_event=stop_event,
                pause_event=pause_event,
                config={"youtube_search_method": "仅API（消耗配额）"},
            )
        elif platform == "tiktok":
            if run_tiktok_spider is None:
                raise ModuleNotFoundError("playwright is required for TikTok keyword scraping")
            max_videos = int(platform_config.get("max_videos", 10))
            run_tiktok_spider(
                keywords_list=[keyword],
                max_videos=max_videos,
                max_candidates=max(int(platform_config.get("max_candidates", max_videos * 3)), max_videos),
                limit_time_str="是",
                start_date=start_date,
                end_date=end_date,
                get_comments_str="否",
                max_comments=0,
                cdp_port_or_url=platform_config.get("cdp_url", "http://localhost:9222"),
                log_callback=log_callback,
                finish_callback=finish_callback,
                stop_event=stop_event,
                pause_event=pause_event,
            )
        elif platform == "x_twitter":
            if run_x_spider is None:
                raise ModuleNotFoundError("playwright is required for X keyword scraping")
            adv_params = {
                "limit_time": "是",
                "start_date": start_date,
                "end_date": end_date,
                "get_comments": "否",
                "max_comments": 0,
                "lang": "any",
                "search_tab": select_x_search_tab(platform_config),
            }
            run_x_spider(
                keywords_list=[keyword],
                adv_params=adv_params,
                port=platform_config.get("cdp_url", "http://localhost:9222"),
                log_callback=log_callback,
                finish_callback=finish_callback,
                stop_event=stop_event,
                pause_event=pause_event,
                config={
                    "max_scrolls": int(platform_config.get("max_scrolls", 2)),
                    "cooldown_min": 2.0,
                    "cooldown_max": 4.0,
                    "no_new_scroll_limit": 2,
                    "slice_days": days,
                    "max_parallel_tabs": 1,
                },
            )
        else:
            finished_at = now_str()
            return SpiderRunResult(
                platform=platform,
                keyword=keyword,
                status=STATUS_UNKNOWN_PLATFORM,
                ids=set(),
                links=set(),
                output_path=None,
                error_message=f"Unknown platform: {platform}",
                started_at=started_at,
                finished_at=finished_at,
            )
    except Exception as exc:
        finished_at = now_str()
        error_message = str(exc)
        logging.exception("爬虫异常 (platform=%s, keyword=%s)", platform, keyword)
        return SpiderRunResult(
            platform=platform,
            keyword=keyword,
            status=classify_error_status(error_message),
            ids=set(),
            links=set(),
            output_path=retrieved_path,
            error_message=error_message,
            started_at=started_at,
            finished_at=finished_at,
        )

    finished_at = now_str()
    if not retrieved_path:
        return SpiderRunResult(
            platform=platform,
            keyword=keyword,
            status=STATUS_FAILED,
            ids=set(),
            links=set(),
            output_path=None,
            error_message="No Excel path returned from spider",
            started_at=started_at,
            finished_at=finished_at,
        )

    if not os.path.exists(retrieved_path):
        return SpiderRunResult(
            platform=platform,
            keyword=keyword,
            status=STATUS_OUTPUT_SCHEMA_ERROR,
            ids=set(),
            links=set(),
            output_path=retrieved_path,
            error_message=f"Excel file not found at {retrieved_path}",
            started_at=started_at,
            finished_at=finished_at,
        )

    try:
        links = extract_links_from_excel(retrieved_path, platform)
    except Exception as exc:
        return SpiderRunResult(
            platform=platform,
            keyword=keyword,
            status=STATUS_OUTPUT_SCHEMA_ERROR,
            ids=set(),
            links=set(),
            output_path=retrieved_path,
            error_message=f"Failed to parse Excel: {exc}",
            started_at=started_at,
            finished_at=finished_at,
        )

    ids = {extract_id_from_link(link, platform) for link in links if link}
    ids = {content_id for content_id in ids if content_id}
    status = STATUS_SUCCESS if ids else STATUS_EMPTY_RESULT
    return SpiderRunResult(
        platform=platform,
        keyword=keyword,
        status=status,
        ids=ids,
        links=links,
        output_path=retrieved_path,
        error_message="",
        started_at=started_at,
        finished_at=finished_at,
        written_count=len(links),
    )


def write_csv_report(games: dict[str, Any], csv_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with open(csv_path, "w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(CSV_HEADERS)

        for game_name, game_data in games.items():
            baseline_query = game_data["baseline_query"]
            for platform, platform_data in game_data["platforms"].items():
                for group in platform_data["groups"]:
                    writer.writerow(
                        [
                            sanitize_csv_cell(game_name),
                            sanitize_csv_cell(platform),
                            sanitize_csv_cell(baseline_query),
                            sanitize_csv_cell(platform_data["baseline_status"]),
                            platform_data["baseline_result_count"],
                            platform_data["baseline_raw_link_count"],
                            sanitize_csv_cell(", ".join(group["keywords"])),
                            sanitize_csv_cell(group["status"]),
                            group["result_count"],
                            group["raw_link_count"],
                            group["baseline_intersection_count"],
                            format_csv_percent(group["relative_result_volume"]),
                            format_csv_percent(group["baseline_overlap_rate"]),
                            group["unique_result_count"],
                            format_csv_percent(group["incremental_gain"]),
                            format_csv_percent(group["jaccard_similarity"]),
                            sanitize_csv_cell(group["error_message"]),
                        ]
                    )


def write_markdown_report(context: dict[str, Any], markdown_path: Path) -> None:
    lines = list(REPORT_INTRO_LINES)
    lines.extend(
        [
            "",
            f"Generated on: {context['generated_at']}",
            "",
            f"- Run ID: `{context['run_id']}`",
            f"- Time Window: `{context['start_date']}` to `{context['end_date']}`",
            f"- Platforms: `{', '.join(context['platforms'])}`",
            f"- X Search Tab: `{context['x_search_tab']}`",
            "",
        ]
    )

    for game_name, game_data in context["games"].items():
        lines.append(f"## Game: {game_name}")
        lines.append(f"- **Baseline Query**: `{game_data['baseline_query']}`")
        lines.append("")

        for platform, platform_data in game_data["platforms"].items():
            lines.append(f"### Platform: {platform.upper()}")
            lines.append(f"- **Baseline Status**: `{platform_data['baseline_status']}`")
            if platform_data["baseline_error"]:
                lines.append(f"- **Baseline Error**: {platform_data['baseline_error']}")
            lines.append(f"- **Baseline Result Count**: {platform_data['baseline_result_count']}")
            lines.append(f"- **Baseline Raw Link Count**: {platform_data['baseline_raw_link_count']}")
            lines.append("")
            lines.append("| " + " | ".join(MARKDOWN_TABLE_HEADERS) + " |")
            lines.append("|:---:|:---|:---:|---:|---:|---:|---:|---:|---:|---:|---:|:---|")

            for idx, group in enumerate(platform_data["groups"], 1):
                escaped_keywords = [keyword.replace("|", "\\|") for keyword in group["keywords"]]
                keyword_text = ", ".join(f"`{keyword}`" for keyword in escaped_keywords)
                lines.append(
                    "| "
                    + " | ".join(
                        [
                            str(idx),
                            keyword_text,
                            f"`{group['status']}`",
                            str(group["result_count"]),
                            str(group["raw_link_count"]),
                            str(group["baseline_intersection_count"]),
                            format_percent(group["relative_result_volume"]),
                            format_percent(group["baseline_overlap_rate"]),
                            str(group["unique_result_count"]),
                            format_percent(group["incremental_gain"]),
                            format_percent(group["jaccard_similarity"]),
                            group["error_message"].replace("|", "\\|"),
                        ]
                    )
                    + " |"
                )
            lines.append("")
        lines.append("---")
        lines.append("")

    markdown_path.parent.mkdir(parents=True, exist_ok=True)
    markdown_path.write_text("\n".join(lines), encoding="utf-8")


def generate_reports(context: dict[str, Any], run_dir: str | Path) -> dict[str, str]:
    run_dir = Path(run_dir)
    reports_dir = run_dir / "reports"
    markdown_path = reports_dir / "calibration_report.md"
    csv_path = reports_dir / "calibration_report.csv"
    write_markdown_report(context, markdown_path)
    write_csv_report(context["games"], csv_path)
    return {
        "markdown": str(markdown_path),
        "csv": str(csv_path),
    }


def run_calibration_task(config: dict, output_path: str, log_callback=None, stop_event=None, pause_event=None):
    time_period = config.get("time_period", {})
    days_raw = time_period.get("days", 7)
    try:
        days = int(days_raw)
    except (ValueError, TypeError) as exc:
        if log_callback:
            log_callback(f"Invalid days value in config (must be integer): {exc}")
        raise ValueError(f"Invalid days value: {exc}") from exc

    if "start_date" in time_period and "end_date" in time_period:
        start_date_str = time_period["start_date"]
        end_date_str = time_period["end_date"]
        try:
            start_dt = dt.datetime.strptime(start_date_str, "%Y-%m-%d")
            end_dt = dt.datetime.strptime(end_date_str, "%Y-%m-%d")
            days = max(1, (end_dt - start_dt).days)
        except (ValueError, TypeError):
            pass
    else:
        end_dt = dt.datetime.now()
        start_dt = end_dt - dt.timedelta(days=days)
        start_date_str = start_dt.strftime("%Y-%m-%d")
        end_date_str = end_dt.strftime("%Y-%m-%d")

    platforms = parse_platforms(config.get("platforms"))
    run_started_at = now_str()
    run_id, run_dir = create_run_directory(output_path, log_callback=log_callback)
    x_search_tab = select_x_search_tab(config.get("x_twitter", {}))

    message = f"Calibration period: {start_date_str} to {end_date_str} ({days} days)"
    print(message)
    if log_callback:
        log_callback(message)
        log_callback(f"Run directory: {run_dir}")

    config_snapshot_path = run_dir / "config_snapshot.json"
    write_json(config_snapshot_path, config)

    games: dict[str, Any] = {}

    for game in config.get("games", []):
        if should_stop(stop_event):
            break
        if wait_if_paused(pause_event, stop_event):
            break

        game_name = game["name"]
        baseline_query = game["baseline_query"]
        keyword_groups = game.get("keyword_groups", [])

        if log_callback:
            log_callback(f"\nProcessing game: {game_name}")
        games[game_name] = {"baseline_query": baseline_query, "platforms": {}}

        for platform in platforms:
            if should_stop(stop_event):
                break
            if wait_if_paused(pause_event, stop_event):
                break

            platform_config = dict(config.get(platform, {}))
            if platform == "x_twitter":
                platform_config["x_search_tab"] = x_search_tab

            if log_callback:
                log_callback(f"  Running searches on platform: {platform}")

            baseline_result = run_platform_spider(
                platform=platform,
                keyword=baseline_query,
                start_date=start_date_str,
                end_date=end_date_str,
                platform_config=platform_config,
                days=days,
                stop_event=stop_event,
                pause_event=pause_event,
            )

            write_json(
                run_dir / "raw" / raw_platform_dir_name(platform) / "baseline.json",
                build_group_snapshot(
                    platform=platform,
                    group_key="baseline",
                    keywords=[baseline_query],
                    status=baseline_result.status,
                    ids=baseline_result.ids,
                    links=baseline_result.links,
                    error_message=baseline_result.error_message or "",
                    started_at=baseline_result.started_at,
                    finished_at=baseline_result.finished_at,
                    scanned_count=baseline_result.scanned_count,
                    written_count=baseline_result.written_count,
                    hit_limit=baseline_result.hit_limit,
                    output_paths=[baseline_result.output_path] if baseline_result.output_path else [],
                ),
            )

            platform_entry = {
                "baseline_status": baseline_result.status,
                "baseline_error": baseline_result.error_message or "",
                "baseline_result_count": len(baseline_result.ids),
                "baseline_raw_link_count": len(baseline_result.links),
                "groups": [],
            }
            games[game_name]["platforms"][platform] = platform_entry

            if not is_successful_run(baseline_result):
                for index, keywords in enumerate(keyword_groups, 1):
                    platform_entry["groups"].append(
                        {
                            "keywords": keywords,
                            "status": REPORT_FAILURE_STATUS,
                            "error_message": f"Baseline query failed: {baseline_result.error_message or baseline_result.status}",
                            "result_count": 0,
                            "raw_link_count": 0,
                            "baseline_intersection_count": 0,
                            "relative_result_volume": None,
                            "baseline_overlap_rate": None,
                            "unique_result_count": 0,
                            "incremental_gain": None,
                            "jaccard_similarity": None,
                        }
                    )
                    write_json(
                        run_dir / "raw" / raw_platform_dir_name(platform) / f"group_{index:02d}.json",
                        build_group_snapshot(
                            platform=platform,
                            group_key=f"group_{index:02d}",
                            keywords=keywords,
                            status=REPORT_FAILURE_STATUS,
                            ids=set(),
                            links=set(),
                            error_message=f"Baseline query failed: {baseline_result.error_message or baseline_result.status}",
                            started_at=None,
                            finished_at=None,
                            scanned_count=None,
                            written_count=None,
                            hit_limit=False,
                            output_paths=[],
                        ),
                    )
                continue

            for index, keywords in enumerate(keyword_groups, 1):
                if should_stop(stop_event):
                    break
                if wait_if_paused(pause_event, stop_event):
                    break

                keyword_runs: list[SpiderRunResult] = []
                group_ids: set[str] = set()
                group_links: set[str] = set()

                for keyword in keywords:
                    if should_stop(stop_event):
                        break
                    if wait_if_paused(pause_event, stop_event):
                        break

                    result = run_platform_spider(
                        platform=platform,
                        keyword=keyword,
                        start_date=start_date_str,
                        end_date=end_date_str,
                        platform_config=platform_config,
                        days=days,
                        stop_event=stop_event,
                        pause_event=pause_event,
                    )
                    keyword_runs.append(result)
                    group_ids.update(result.ids)
                    group_links.update(result.links)

                group_status, group_error = derive_group_status(keyword_runs, group_ids)
                metrics = calculate_group_metrics(group_ids, baseline_result.ids)
                scanned_count, written_count, hit_limit = summarize_counts(keyword_runs, group_links)
                output_paths = [result.output_path for result in keyword_runs if result.output_path]
                started_at = min((result.started_at for result in keyword_runs), default=None)
                finished_at = max((result.finished_at for result in keyword_runs), default=None)

                group_entry = {
                    "keywords": keywords,
                    "status": group_status,
                    "error_message": group_error,
                    "result_count": metrics["result_count"],
                    "raw_link_count": len(group_links),
                    "baseline_intersection_count": metrics["baseline_intersection_count"],
                    "relative_result_volume": metrics["relative_result_volume"],
                    "baseline_overlap_rate": metrics["baseline_overlap_rate"],
                    "unique_result_count": metrics["unique_result_count"],
                    "incremental_gain": metrics["incremental_gain"],
                    "jaccard_similarity": metrics["jaccard_similarity"],
                }
                platform_entry["groups"].append(group_entry)

                write_json(
                    run_dir / "raw" / raw_platform_dir_name(platform) / f"group_{index:02d}.json",
                    build_group_snapshot(
                        platform=platform,
                        group_key=f"group_{index:02d}",
                        keywords=keywords,
                        status=group_status,
                        ids=group_ids,
                        links=group_links,
                        error_message=group_error,
                        started_at=started_at,
                        finished_at=finished_at,
                        scanned_count=scanned_count,
                        written_count=written_count,
                        hit_limit=hit_limit,
                        output_paths=output_paths,
                        keyword_runs=keyword_runs,
                    ),
                )

    run_finished_at = now_str()
    environment_snapshot = build_environment_snapshot(
        run_id=run_id,
        run_started_at=run_started_at,
        run_finished_at=run_finished_at,
        start_date=start_date_str,
        end_date=end_date_str,
        days=days,
        platforms=platforms,
        x_search_tab=x_search_tab,
    )
    write_json(run_dir / "environment_snapshot.json", environment_snapshot)

    if not should_stop(stop_event):
        report_context = {
            "run_id": run_id,
            "generated_at": run_finished_at,
            "start_date": start_date_str,
            "end_date": end_date_str,
            "platforms": platforms,
            "x_search_tab": x_search_tab,
            "games": games,
        }
        report_paths = generate_reports(report_context, run_dir)
        if log_callback:
            log_callback(f"Markdown report: {report_paths['markdown']}")
            log_callback(f"CSV report: {report_paths['csv']}")
            log_callback(f"Run completed: {run_dir}")

    return str(run_dir)


def main():
    parser = argparse.ArgumentParser(description="Observable Search Coverage Calibration Tool")
    parser.add_argument("--config", type=str, default="config/calibration_config.json", help="Path to configuration file")
    parser.add_argument("--output", type=str, default="output/calibration_report.md", help="Output root or legacy report file path")
    args = parser.parse_args()

    try:
        config = load_config(args.config)
    except Exception as exc:
        print(f"Failed to load config: {exc}", file=sys.stderr)
        sys.exit(1)

    try:
        run_calibration_task(config, args.output)
    except Exception as exc:
        print(f"Calibration failed: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
