from __future__ import annotations

import queue
import random
import re
import threading
import time
import urllib.parse
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

try:
    from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError, Error as PlaywrightError
    PLAYWRIGHT_IMPORT_ERROR = None
except ModuleNotFoundError as exc:  # pragma: no cover - exercised via import-time fallback
    sync_playwright = None
    PLAYWRIGHT_IMPORT_ERROR = exc

    class PlaywrightTimeoutError(Exception):
        pass

    class PlaywrightError(Exception):
        pass

from src.core import (
    XlsxRowWriter,
    MultiSheetXlsxWriter,
    build_output_path,
    connect_existing_chromium,
    ensure_chrome_for_cdp,
    expand_compact_number,
    interruptible_sleep,
    log_error,
    log_line,
    make_keyword_log,
    random_cooldown,
    sanitize_csv_row,
    sanitize_csv_rows,
    should_stop,
    wait_if_paused,
)
from src.platforms.x_twitter.comments import extract_comments
from src.platforms.x_twitter.profiles import (
    extract_author_from_article as extract_x_author_from_article,
    extract_profile_record as extract_x_profile_record,
)

MAX_SEARCH_SCROLLS = 200
STATUS_PATH_RE = re.compile(r"(/[^/]+/status/\d+)")

CSV_FIELDS = [
    "原始搜索词",
    "完整搜索语法",
    "序号",
    "推文内容",
    "浏览量",
    "点赞量",
    "转发量",
    "评论数",
    "发帖时间",
    "推文链接",
    "作者主页链接",
    "作者的名称",
    "账号ID",
    "粉丝数",
    "简介",
    "标签",
]


def count_x_keyword_output_rows(path: str | Path | None) -> int:
    """
    统计 X 关键词输出文件中的主数据行数（不含表头）。

    只把“推文信息”或“数据”sheet 作为关键词成功依据，不把“评论信息”计入。
    文件不存在、读取失败或只有表头时返回 0。
    """
    if not path:
        return 0
    file_path = Path(path)
    if not file_path.exists():
        return 0
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
    except Exception:
        return 0
    try:
        target_sheets = []
        for name in ("推文信息", "数据"):
            if name in wb.sheetnames:
                target_sheets.append(wb[name])
        if not target_sheets:
            # 兜底：没有标准 sheet 名时，只取第一个非“评论信息”的 sheet
            target_sheets = [ws for ws in wb.worksheets if ws.title != "评论信息"][:1]

        total = 0
        for ws in target_sheets:
            first = True
            for row in ws.iter_rows(values_only=True):
                if first:
                    first = False
                    continue
                if row and any(value is not None and str(value).strip() != "" for value in row):
                    total += 1
        return total
    finally:
        try:
            wb.close()
        except Exception:
            pass


def is_x_keyword_output_empty(path: str | Path | None) -> bool:
    """
    判断 X 关键词输出文件是否为空（没有真实推文数据行）。
    """
    return count_x_keyword_output_rows(path) <= 0


def normalize_status_url(url: str) -> str:
    if not url:
        return ""
    normalized = url.strip().replace("twitter.com", "x.com")
    normalized = normalized.split("?")[0].split("#")[0]
    if normalized.startswith("//"):
        normalized = "https:" + normalized
    if normalized.startswith("/"):
        normalized = "https://x.com" + normalized
    if normalized and not normalized.startswith("http"):
        normalized = "https://" + normalized
    return normalized

def safe_text(locator, default: str = "") -> str:
    try:
        if locator.count() <= 0:
            return default
        return locator.first.inner_text(timeout=1500).strip() or default
    except (PlaywrightTimeoutError, PlaywrightError, ValueError):
        return default

def safe_attr(locator, attr: str, default: str = "") -> str:
    try:
        if locator.count() <= 0:
            return default
        return locator.first.get_attribute(attr, timeout=1500) or default
    except (PlaywrightTimeoutError, PlaywrightError, ValueError, KeyError):
        return default

def collect_status_urls(article) -> list[str]:
    urls: list[str] = []
    seen = set()
    try:
        anchors = article.locator('a[href*="/status/"]').all()
    except (PlaywrightTimeoutError, PlaywrightError):
        return urls

    for anchor in anchors:
        try:
            href = anchor.get_attribute("href") or ""
        except (PlaywrightTimeoutError, PlaywrightError, KeyError):
            continue
        match = STATUS_PATH_RE.search(href)
        if not match:
            continue
        normalized = normalize_status_url(match.group(1))
        if normalized and normalized not in seen:
            urls.append(normalized)
            seen.add(normalized)
    return urls

def is_repost_context(text: str) -> bool:
    lowered = (text or "").lower()
    return any(
        token in lowered
        for token in [
            "reposted", "repost", "retweeted", "retweet",
            "republished", "reposted by", "quoted", "quote",
            "转推", "转发", "引用", "リポスト",
            "リツイート", "再投稿", "已轉推",
        ]
    )

def get_social_context(article) -> str:
    return safe_text(article.locator('[data-testid="socialContext"]'))

def article_contains_nested_tweet(article) -> bool:
    status_urls = collect_status_urls(article)
    if len(status_urls) > 1:
        return True
    try:
        nested_articles = article.locator('article[data-testid="tweet"]').count()
        return nested_articles > 0
    except (PlaywrightTimeoutError, PlaywrightError):
        return False

def get_tweet_url(article) -> str:
    status_urls = collect_status_urls(article)
    return status_urls[0] if status_urls else ""

def get_tweet_text(article, stop_event=None) -> str:
    try:
        revert_locator = article.locator("text='view original', text='查看原文', text='原文を表示', text='show original', text='原文を見る'").first
        revert_locator.click(timeout=500)
    except (PlaywrightTimeoutError, PlaywrightError):
        pass

    try:
        expand_locator = article.locator("text='show more', text='show more...', text='もっと見る', text='더 보기', text='显示更多'").first
        expand_locator.click(timeout=500)
    except (PlaywrightTimeoutError, PlaywrightError):
        pass

    interruptible_sleep(0.3, stop_event)
    return safe_text(article.locator('[data-testid="tweetText"]'), default="无文字内容")

def get_tweet_time(article) -> str:
    raw_time = safe_attr(article.locator("time"), "datetime", default="")
    if not raw_time:
        return ""
    try:
        dt_obj = datetime.fromisoformat(raw_time.replace("Z", "+00:00"))
        return dt_obj.strftime("%Y-%m-%d %H:%M:%S")
    except (ValueError, TypeError):
        return raw_time

def extract_metric_value(locator, default: str = "未知") -> str:
    try:
        if locator.count() <= 0:
            return default
        node = locator.first
        raw_text = node.inner_text(timeout=1500).strip()
        if raw_text:
            return expand_compact_number(raw_text)
        aria = node.get_attribute("aria-label", timeout=1500) or ""
        match = re.search(r"(\d[\d,.]*[KMBkmb千万萬亿億]?)", aria)
        if match:
            return expand_compact_number(match.group(1))
        if aria:
            return "0"
    except (PlaywrightTimeoutError, PlaywrightError, ValueError, AttributeError):
        pass
    return default

def extract_metric_from_article(article, selectors, default: str = "未知") -> str:
    if isinstance(selectors, str):
        selectors = [selectors]
    for selector in selectors:
        value = extract_metric_value(article.locator(selector), default="")
        if value:
            return value
    return default

def _x_media_tag(media_label: str) -> str:
    """Convert get_media_label output to numeric tag.
    0=图片+视频, 1=图片, 2=视频, 3=纯文本, 4=其它
    """
    has_video = "视频" in media_label
    has_photo = "图片" in media_label
    if has_video and has_photo:
        return "0"
    if has_photo:
        return "1"
    if has_video:
        return "2"
    if media_label:
        return "4"
    return "3"


def get_media_label(article) -> str:
    """Detect media type from the article's own content.

    Avoids false positives:
    - Video thumbnails won't be counted as photos (video takes priority).
    - Media inside embedded/quoted tweets is ignored — we only look at
      the article's direct children, not nested articles.
    """
    # If this article embeds another tweet (repost / quote), only look at
    # media that belongs to the outer article itself by excluding the
    # nested article subtree.
    embedded_root = _find_embedded_tweet_root(article)

    labels: list[str] = []
    video_selectors = [
        "video",
        '[data-testid="videoPlayer"]',
        '[aria-label*="Play"]',
        '[aria-label*="play"]',
        '[aria-label*="播放"]',
        '[aria-label*="再生"]',
    ]
    photo_selectors = [
        '[data-testid="tweetPhoto"]',
        'a[href*="/photo/"]',
    ]

    def _element_is_inside_embedded(el_handle) -> bool:
        if embedded_root is None:
            return False
        try:
            return bool(article.evaluate(
                """([el, root]) => {
                    let node = el;
                    while (node && node !== root && node !== document.body) {
                        node = node.parentElement;
                    }
                    return node === root;
                }""",
                [el_handle, embedded_root],
            ))
        except Exception:
            return False

    has_video = False
    for selector in video_selectors:
        try:
            elements = article.locator(selector).all()
        except Exception:
            continue
        for el in elements:
            try:
                if not _element_is_inside_embedded(el):
                    has_video = True
                    break
            except Exception:
                continue
        if has_video:
            break

    has_photo = False
    # Only check photos if no video detected (avoid thumbnail false positives)
    if not has_video:
        for selector in photo_selectors:
            try:
                elements = article.locator(selector).all()
            except Exception:
                continue
            for el in elements:
                try:
                    if not _element_is_inside_embedded(el):
                        has_photo = True
                        break
                except Exception:
                    continue
            if has_photo:
                break

    if has_video:
        labels.append("视频")
    if has_photo:
        labels.append("图片")

    if not has_video and not has_photo:
        first_line = (article.inner_text() or "").split("\n")[0].strip().lower()
        if first_line == "gif":
            labels.append("GIF")

    return f"[{' + '.join(labels)}]" if labels else ""


def _find_embedded_tweet_root(article):
    """Return the root element of a nested/embedded tweet inside *article*, or None."""
    try:
        return article.evaluate("""el => {
            const nested = el.querySelector(
                'article[data-testid="tweet"]:not([data-testid="tweet"] [data-testid="tweet"])'
            );
            if (nested && nested !== el) {
                // Walk up one level to capture the quote/repost container
                let container = nested.closest('[role="link"]');
                if (!container) container = nested.parentElement;
                return container || nested;
            }
            return null;
        }""")
    except Exception:
        return None


def should_keep_article(article) -> bool:
    if is_repost_context(get_social_context(article)):
        return False
    if article_contains_nested_tweet(article):
        return False
    return True

def append_rows(writer, rows: list[dict], sheet_name: str = "推文信息"):
    if not rows:
        return
    sanitized = sanitize_csv_rows(rows)
    if hasattr(writer, "writerow") and hasattr(writer, "worksheets"):
        for row in sanitized:
            writer.writerow(sheet_name, row)
    else:
        writer.writerows(sanitized)

def build_search_query(base_keyword: str, adv_params: dict, since: str, until: str) -> str:
    query_parts = [base_keyword]
    if adv_params.get("lang", "any") != "any":
        query_parts.append(f"lang:{adv_params['lang']}")
    if since and until:
        query_parts.append(f"since:{since}")
        query_parts.append(f"until:{until}")
    return " ".join(query_parts)


def resolve_search_tab_filter(search_tab: str | None) -> str:
    mapping = {
        "top": "top",
        "latest": "live",
        "media": "media",
        "people": "user",
    }
    normalized = (search_tab or "top").strip().lower()
    return mapping.get(normalized, "top")


def build_search_url(final_query: str, search_tab: str | None = "top") -> str:
    tab_filter = resolve_search_tab_filter(search_tab)
    return f"https://x.com/search?q={urllib.parse.quote(final_query)}&src=typed_query&f={tab_filter}"


def ensure_playwright_available():
    if sync_playwright is None:
        raise ModuleNotFoundError("playwright is required for X keyword scraping") from PLAYWRIGHT_IMPORT_ERROR

def _make_keyword_log_callback(base_log_callback, keyword: str):
    """Wrap log_callback to prefix messages with [keyword] for disambiguation."""
    return make_keyword_log(base_log_callback, keyword)


def _try_reload_if_empty(page, page_timeout, refresh_count, refresh_interval, log, stop_event, label="页面"):
    """After goto, reload the page if no tweet articles appear."""
    for attempt in range(refresh_count + 1):
        if should_stop(stop_event):
            return
        try:
            page.wait_for_selector('article[data-testid="tweet"]', state="attached", timeout=15000)
            return
        except Exception:
            if attempt < refresh_count:
                log(f"  {label}未加载内容，第 {attempt + 1}/{refresh_count} 次刷新...")
                try:
                    page.reload(wait_until="domcontentloaded", timeout=page_timeout)
                except Exception:
                    pass
                if interruptible_sleep(refresh_interval, stop_event):
                    return


def _x_comment_consumer(keyword, queue_obj, cdp_port_or_url, writer, writer_lock,
                       log_callback, stop_event, pause_event, max_comments,
                       consumers_ready=None, page_timeout=30000,
                       comment_no_new_scroll_limit=5,
                       comment_refresh_count=3, comment_refresh_interval=5.0,
                       browser_choice=None):
    """Consumer thread: creates its own Playwright connection + page, pops from queue."""
    log = _make_keyword_log_callback(log_callback, keyword)
    comments_page = None
    try:
        with sync_playwright() as p:
            try:
                _, context = connect_existing_chromium(p, cdp_port_or_url, browser=browser_choice)
                comments_page = context.new_page()
            except Exception as exc:
                log(f"    评论线程连接浏览器失败: {exc}")
                return
            if consumers_ready is not None:
                consumers_ready.set()
            while True:
                try:
                    item = queue_obj.get(timeout=3)
                except Exception:
                    # queue.get can raise on timeout; check if we should keep waiting
                    if should_stop(stop_event):
                        break
                    if wait_if_paused(pause_event, stop_event):
                        break
                    continue
                if item is None:
                    break
                if should_stop(stop_event):
                    break
                if wait_if_paused(pause_event, stop_event):
                    break
                serial_number, tweet_url, max_scan = item
                try:
                    comments_page.goto(tweet_url, wait_until="domcontentloaded", timeout=page_timeout)
                    interruptible_sleep(random.uniform(2, 3), stop_event)
                    _try_reload_if_empty(comments_page, page_timeout, comment_refresh_count, comment_refresh_interval, log, stop_event, "评论页")
                    interruptible_sleep(random.uniform(3, 5), stop_event)
                    comments = extract_comments(comments_page, tweet_url, max_scan, log,
                                                stop_event, pause_event=pause_event,
                                                no_new_scroll_limit=comment_no_new_scroll_limit)
                    with writer_lock:
                        for comment in comments:
                            comment_row = {
                                "序号": str(serial_number),
                                "推文链接": tweet_url,
                                "评论的点赞量": comment.get("likes", ""),
                                "评论内容": comment.get("content", ""),
                                "评论发布时间": comment.get("time", ""),
                            }
                            writer.writerow("评论信息", sanitize_csv_row(comment_row))
                        writer.save()
                except Exception as exc:
                    log(f"    提取评论失败：{exc}")
    except Exception as exc:
        log_error(log, f"评论线程异常: {exc}")
    finally:
        if comments_page is not None:
            try:
                if not comments_page.is_closed():
                    comments_page.close()
            except Exception:
                pass


RECOMMENDATION_MARKERS = (
    "discover more", "find more", "发现更多", "更多了解",
    "もっと見る", "더 보기", "encontrar más", "descubre más",
    "weiter entdecken", "scopri di più",
)


def _find_recommendation_boundary_index(page) -> int:
    """Return the index of the first article after the recommendation divider, or -1 if none found."""
    try:
        return page.evaluate("""markers => {
            const articles = Array.from(document.querySelectorAll('article[data-testid="tweet"]'));
            if (articles.length === 0) return -1;
            // Search for a cellInnerDiv that contains marker text but no tweet article.
            const cells = document.querySelectorAll('[data-testid="cellInnerDiv"]');
            for (const cell of cells) {
                if (cell.querySelector('article[data-testid="tweet"]')) continue;
                const text = (cell.textContent || '').trim().toLowerCase();
                if (markers.some(m => text.includes(m))) {
                    for (let i = 0; i < articles.length; i++) {
                        if (cell.compareDocumentPosition(articles[i]) & 2) return i;
                    }
                    return articles.length;
                }
            }
            // Fallback: search for any heading/span outside articles
            for (const heading of document.querySelectorAll('[role="heading"], h1, h2, h3')) {
                const text = (heading.textContent || '').trim().toLowerCase();
                if (markers.some(m => text.includes(m))) {
                    for (let i = 0; i < articles.length; i++) {
                        if (heading.compareDocumentPosition(articles[i]) & 2) return i;
                    }
                    return articles.length;
                }
            }
            return -1;
        }""", list(RECOMMENDATION_MARKERS))
    except Exception:
        return -1


def _x_profile_cache_key(profile_url: str, account_id: str = "") -> str:
    if account_id:
        return account_id.lower().lstrip("@")
    match = re.search(r"x\.com/([^/?#]+)/?$", normalize_status_url(profile_url) or "")
    return match.group(1).lower() if match else (profile_url or "").lower()


def _fallback_x_author_record(author: dict) -> dict[str, str]:
    return {
        "作者主页链接": author.get("profile_url", ""),
        "作者的名称": author.get("author_name", ""),
        "账号ID": author.get("account_id", ""),
        "粉丝数": "",
        "简介": "",
    }


def enrich_x_author_profile(profile_page, author: dict, cache: dict[str, dict[str, str]], log, page_timeout=None, stop_event=None) -> dict[str, str]:
    """
    进入 X 作者主页补齐名称、粉丝数和简介；同一作者只抓一次。
    """
    if not author:
        return _fallback_x_author_record({})
    profile_url = author.get("profile_url", "")
    account_id = author.get("account_id", "")
    if not profile_url and account_id:
        profile_url = f"https://x.com/{account_id}"
    if not profile_url:
        return _fallback_x_author_record(author)

    cache_key = _x_profile_cache_key(profile_url, account_id)
    if cache_key in cache:
        cached = cache[cache_key]
    else:
        try:
            cached = extract_x_profile_record(
                profile_page,
                profile_url,
                log,
                page_timeout=page_timeout,
                stop_event=stop_event,
            ) or {}
        except Exception as exc:
            log(f"    作者主页补充失败，保留搜索页字段：{profile_url}，{exc}")
            cached = {}
        fallback = _fallback_x_author_record({**author, "profile_url": profile_url})
        for key, value in fallback.items():
            cached.setdefault(key, value)
        cache[cache_key] = cached
    return {
        "作者主页链接": cached.get("作者主页链接") or profile_url,
        "作者的名称": cached.get("作者的名称") or author.get("author_name", ""),
        "账号ID": cached.get("账号ID") or account_id,
        "粉丝数": cached.get("粉丝数", ""),
        "简介": cached.get("简介", ""),
    }


def _scrape_single_x_keyword(base_keyword, adv_params, port,
                             log_callback, stop_event, pause_event,
                             search_page_timeout, scroll_cooldown_min, scroll_cooldown_max,
                             no_change_threshold, max_search_scrolls, slice_days,
                             max_comment_tabs, max_queue_size,
                             comment_no_new_scroll_limit=5,
                             search_refresh_count=3, search_refresh_interval=5.0,
                             comment_refresh_count=3, comment_refresh_interval=5.0,
                             browser_choice=None):
    """Scrape a single X keyword in this thread. Spawns comment consumer threads if needed."""
    log = _make_keyword_log_callback(log_callback, base_keyword)
    output_path = None
    writer = None
    writer_lock = None
    comment_queue = None
    comment_threads: list[threading.Thread] = []
    search_page = None
    profile_page = None
    author_profile_cache: dict[str, dict[str, str]] = {}
    try:
        if should_stop(stop_event):
            log("任务已停止。")
            return None
        if wait_if_paused(pause_event, stop_event):
            log("任务已停止。")
            return None

        limit_time_bool = adv_params.get("limit_time") == "是"
        get_comments_bool = adv_params.get("get_comments") == "是"
        max_comments = int(adv_params.get("max_comments", 500))

        safe_fn = re.sub(r'[\\/*?:"<>|]', "", base_keyword)
        output_path = build_output_path("x", f"x_keyword_{safe_fn}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx", channel="keyword")

        log(f"\n{'=' * 50}")
        log(f"开始关键词：{base_keyword}")
        log(f"输出文件：{output_path}")

        if limit_time_bool:
            try:
                start_dt = datetime.strptime(adv_params["start_date"], "%Y-%m-%d")
                end_dt = datetime.strptime(adv_params["end_date"], "%Y-%m-%d") + timedelta(days=1)
            except ValueError:
                log("日期或切片格式错误：日期必须是 YYYY-MM-DD，切片天数必须是整数。")
                return None
            if start_dt >= end_dt:
                log("起始日期必须早于结束日期。")
                return None
        else:
            start_dt = datetime.now()
            end_dt = datetime.now()

        with sync_playwright() as p:
            _, context = connect_existing_chromium(p, port, browser=browser_choice)
            search_page = context.new_page()
            profile_page = context.new_page()

            if get_comments_bool:
                comment_fields = ["序号", "推文链接", "评论的点赞量", "评论内容", "评论发布时间"]
                writer = MultiSheetXlsxWriter(output_path, {"推文信息": CSV_FIELDS, "评论信息": comment_fields}, autosave_every=10)
                writer_lock = threading.Lock()
                comment_queue = queue.Queue(maxsize=max_queue_size)
                consumers_ready = threading.Event()
                for _ in range(max_comment_tabs):
                    t = threading.Thread(
                        target=_x_comment_consumer,
                        args=(base_keyword, comment_queue, port, writer, writer_lock,
                              log_callback, stop_event, pause_event, max_comments,
                              consumers_ready, search_page_timeout,
                              comment_no_new_scroll_limit,
                              comment_refresh_count, comment_refresh_interval,
                              browser_choice),
                        daemon=True,
                    )
                    t.start()
                    comment_threads.append(t)
            else:
                writer = XlsxRowWriter(output_path, CSV_FIELDS, autosave_every=10)

            seen_urls = set()
            total_count = 0
            current_end_dt = end_dt
            slice_index = 1

            while (limit_time_bool and current_end_dt > start_dt) or (not limit_time_bool and slice_index == 1):
                if should_stop(stop_event):
                    log("已请求停止，结束当前关键词。")
                    break
                if wait_if_paused(pause_event, stop_event):
                    break

                if limit_time_bool:
                    current_start_dt = max(start_dt, current_end_dt - timedelta(days=slice_days))
                    since = current_start_dt.strftime("%Y-%m-%d")
                    until = current_end_dt.strftime("%Y-%m-%d")
                    log(f"\n[切片 {slice_index}] {since} 至 {until}")
                else:
                    since = ""
                    until = ""
                    log("\n[搜索] 不限时间")

                final_query = build_search_query(base_keyword, adv_params, since, until)
                search_url = build_search_url(final_query, adv_params.get("search_tab", "top"))
                log(f"搜索语法：{final_query}")

                try:
                    search_page.goto(search_url, wait_until="domcontentloaded", timeout=search_page_timeout)
                except Exception:
                    log("页面加载超时，继续尝试提取当前已加载内容。")

                if interruptible_sleep(random.uniform(4, 6), stop_event):
                    break
                _try_reload_if_empty(search_page, search_page_timeout, search_refresh_count, search_refresh_interval, log, stop_event, "搜索页")
                slice_count = 0
                previous_count = -1
                no_change_strikes = 0
                buffer_rows: list[dict] = []

                for scroll_index in range(max_search_scrolls):
                    if should_stop(stop_event):
                        break
                    if wait_if_paused(pause_event, stop_event):
                        break
                    try:
                        retry_btn = search_page.locator(
                            "button:has-text('Retry'), button:has-text('重试'), button:has-text('再試行')"
                        ).first
                        if retry_btn.count() > 0:
                            retry_btn.click(force=True)
                            if interruptible_sleep(3, stop_event):
                                break
                    except Exception:
                        pass

                    stop_outer = False
                    all_articles = search_page.locator('article[data-testid="tweet"]').all()
                    boundary_idx = _find_recommendation_boundary_index(search_page)
                    if boundary_idx >= 0:
                        all_articles = all_articles[:boundary_idx]
                    for article in all_articles:
                        if should_stop(stop_event):
                            break
                        if wait_if_paused(pause_event, stop_event):
                            break
                        try:
                            if not should_keep_article(article):
                                continue

                            tweet_url = get_tweet_url(article)
                            if not tweet_url or tweet_url in seen_urls:
                                continue
                            seen_urls.add(tweet_url)

                            media_label = get_media_label(article)
                            author = extract_x_author_from_article(article)
                            author_record = enrich_x_author_profile(
                                profile_page,
                                author,
                                author_profile_cache,
                                log,
                                page_timeout=search_page_timeout,
                                stop_event=stop_event,
                            )
                            row = {
                                "原始搜索词": base_keyword,
                                "完整搜索语法": final_query,
                                "序号": str(total_count + 1),
                                "推文内容": get_tweet_text(article, stop_event=stop_event) + media_label,
                                "浏览量": extract_metric_from_article(article, [
                                    'a[href*="/analytics"]',
                                    'div[data-testid="postViewCount"]',
                                    '[aria-label*="Views"]',
                                    '[aria-label*="views"]',
                                    '[aria-label*="浏览"]',
                                ]),
                                "点赞量": extract_metric_from_article(article, '[data-testid="like"], [data-testid="unlike"]'),
                                "转发量": extract_metric_from_article(article, '[data-testid="retweet"], [data-testid="unretweet"]'),
                                "评论数": extract_metric_from_article(article, '[data-testid="reply"]'),
                                "发帖时间": get_tweet_time(article),
                                "推文链接": tweet_url,
                                "作者主页链接": author_record.get("作者主页链接", ""),
                                "作者的名称": author_record.get("作者的名称", ""),
                                "账号ID": author_record.get("账号ID", ""),
                                "粉丝数": author_record.get("粉丝数", ""),
                                "简介": author_record.get("简介", ""),
                                "标签": _x_media_tag(media_label),
                            }
                            buffer_rows.append(row)
                            total_count += 1
                            slice_count += 1
                            log(f"    [{total_count}] {tweet_url}")

                            if get_comments_bool:
                                comment_str = row.get("评论数", "0")
                                if comment_str not in ("0", "未知", ""):
                                    if consumers_ready.wait(timeout=0.5):
                                        try:
                                            comment_queue.put(
                                                (row["序号"], tweet_url, max_comments),
                                                block=True,
                                                timeout=15,
                                            )
                                        except Exception:
                                            log("    评论队列已满或消费线程异常，跳过本条评论采集。")
                                    else:
                                        log("    跳过评论采集：评论消费线程连接失败。")

                            if len(buffer_rows) >= 10:
                                if writer_lock:
                                    with writer_lock:
                                        append_rows(writer, buffer_rows)
                                else:
                                    append_rows(writer, buffer_rows)
                                log(f"  自动保存：累计 {total_count} 条含媒体原创推文。")
                                buffer_rows.clear()
                                if total_count and total_count % 20 == 0:
                                    if random_cooldown(log, stop_event, 3.0, 8.0):
                                        stop_outer = True
                                        break
                        except Exception as e:
                            log(f"  单条推文提取失败，已跳过：{e}")

                    if buffer_rows:
                        if writer_lock:
                            with writer_lock:
                                append_rows(writer, buffer_rows)
                        else:
                            append_rows(writer, buffer_rows)
                        buffer_rows.clear()

                    if slice_count == previous_count:
                        no_change_strikes += 1
                        if no_change_strikes >= no_change_threshold:
                            log(f"  连续 {no_change_threshold} 次滚动无新增，停止当前切片。")
                            break
                    else:
                        no_change_strikes = 0
                        log(f"  第 {scroll_index + 1} 次滚动：累计 {total_count} 条。")
                    previous_count = slice_count

                    if not stop_outer:
                        search_page.mouse.wheel(delta_x=0, delta_y=random.randint(900, 1400))
                    if stop_outer or interruptible_sleep(random.uniform(scroll_cooldown_min, scroll_cooldown_max), stop_event):
                        break

                log(f"当前切片捕获 {slice_count} 条含媒体原创推文。")
                if limit_time_bool:
                    current_end_dt = current_start_dt
                slice_index += 1

            log(f"关键词完成：{base_keyword}，累计 {total_count} 条。")
            if comment_threads and comment_queue is not None:
                for _ in comment_threads:
                    comment_queue.put(None)
                for t in comment_threads:
                    t.join(timeout=120)

            if writer_lock:
                with writer_lock:
                    writer.save()
            else:
                writer.save()
            return output_path

    except Exception as exc:
        log_error(log, f"发生致命错误：{exc}")
        if writer is not None:
            try:
                if writer_lock:
                    with writer_lock:
                        writer.save()
                else:
                    writer.save()
            except Exception:
                pass
        return None
    finally:
        if comment_threads and comment_queue is not None:
            try:
                for _ in comment_threads:
                    comment_queue.put(None)
            except Exception:
                pass
            for t in comment_threads:
                if t.is_alive():
                    t.join(timeout=10)
        if search_page is not None and not search_page.is_closed():
            try:
                search_page.close()
            except Exception:
                pass
        if profile_page is not None and not profile_page.is_closed():
            try:
                profile_page.close()
            except Exception:
                pass


def run_x_spider(keywords_list, adv_params, port, log_callback, finish_callback, stop_event=None, config=None, pause_event=None):
    ensure_playwright_available()
    if config is None:
        config = {}
    search_page_timeout = int(config.get("search_page_timeout", 40000))
    scroll_cooldown_min = float(config.get("cooldown_min", 5.0))
    scroll_cooldown_max = float(config.get("cooldown_max", 7.0))
    no_change_threshold = int(config.get("no_new_scroll_limit", 5))
    max_search_scrolls = int(config.get("max_scrolls", MAX_SEARCH_SCROLLS))
    slice_days = int(config.get("slice_days", 7))
    max_parallel_tabs = max(1, min(3, int(config.get("max_parallel_tabs", 1))))
    max_comment_tabs = max(1, min(3, int(config.get("max_comment_tabs", 1))))
    max_queue_size = max(10, min(10000, int(config.get("max_queue_size", 5000))))
    comment_no_new_scroll_limit = int(config.get("comment_no_new_scroll_limit", 5))
    search_refresh_count = int(config.get("search_refresh_count", 3))
    search_refresh_interval = float(config.get("search_refresh_interval", 5.0))
    comment_refresh_count = int(config.get("comment_refresh_count", 3))
    comment_refresh_interval = float(config.get("comment_refresh_interval", 5.0))
    empty_retry_rounds = max(0, min(10, int(config.get("empty_retry_rounds", 2))))
    empty_retry_cooldown_min = float(config.get("empty_retry_cooldown_min", 15.0))
    empty_retry_cooldown_max = float(config.get("empty_retry_cooldown_max", 30.0))
    browser_choice = config.get("browser")

    def run_keyword_batch(batch_keywords: list[str], round_label: str) -> dict[str, str]:
        """
        执行一批关键词，返回 {keyword: output_path}。顺序/并行都完整收集路径。
        """
        batch_paths: dict[str, str] = {}
        if not batch_keywords:
            return batch_paths

        if max_parallel_tabs <= 1 or len(batch_keywords) <= 1:
            for keyword_index, base_keyword in enumerate(batch_keywords, 1):
                if should_stop(stop_event):
                    log_line(log_callback, "任务已停止。")
                    break
                if wait_if_paused(pause_event, stop_event):
                    break
                log_line(log_callback, f"[{round_label} {keyword_index}/{len(batch_keywords)}] 开始关键词：{base_keyword}")
                path = _scrape_single_x_keyword(
                    base_keyword, adv_params, port,
                    log_callback, stop_event, pause_event,
                    search_page_timeout, scroll_cooldown_min, scroll_cooldown_max,
                    no_change_threshold, max_search_scrolls, slice_days,
                    max_comment_tabs, max_queue_size,
                    comment_no_new_scroll_limit,
                    search_refresh_count, search_refresh_interval,
                    comment_refresh_count, comment_refresh_interval,
                    browser_choice,
                )
                if path:
                    batch_paths[base_keyword] = path
            return batch_paths

        with ThreadPoolExecutor(max_workers=max_parallel_tabs) as executor:
            future_to_keyword = {}
            for keyword_index, base_keyword in enumerate(batch_keywords, 1):
                if should_stop(stop_event):
                    break
                if wait_if_paused(pause_event, stop_event):
                    break
                log_line(log_callback, f"[{round_label} {keyword_index}/{len(batch_keywords)}] 开始关键词：{base_keyword}")
                future = executor.submit(
                    _scrape_single_x_keyword,
                    base_keyword, adv_params, port,
                    log_callback, stop_event, pause_event,
                    search_page_timeout, scroll_cooldown_min, scroll_cooldown_max,
                    no_change_threshold, max_search_scrolls, slice_days,
                    max_comment_tabs, max_queue_size,
                    comment_no_new_scroll_limit,
                    search_refresh_count, search_refresh_interval,
                    comment_refresh_count, comment_refresh_interval,
                    browser_choice,
                )
                future_to_keyword[future] = base_keyword

            for future in as_completed(future_to_keyword):
                keyword = future_to_keyword[future]
                try:
                    path = future.result()
                    if path:
                        batch_paths[keyword] = path
                except Exception as exc:
                    log_error(log_callback, f"[{keyword}] 线程异常: {exc}")
        return batch_paths

    try:
        # pre-launch browser once before fanning out to threads
        ensure_chrome_for_cdp(port, log_callback=log_callback, browser=browser_choice)

        get_comments_bool = adv_params.get("get_comments") == "是"
        if not get_comments_bool:
            log_line(log_callback, "过滤规则：跳过转推、跳过引用/嵌套推文。\n")

        latest_path_by_keyword: dict[str, str] = {}

        # 初轮：跑全部关键词
        latest_path_by_keyword.update(run_keyword_batch(list(keywords_list), "初轮"))

        def empty_keywords() -> list[str]:
            empties = []
            for kw in keywords_list:
                path = latest_path_by_keyword.get(kw)
                if is_x_keyword_output_empty(path):
                    empties.append(kw)
            return empties

        pending_empty = empty_keywords()
        if pending_empty:
            log_line(log_callback, f"检测到 {len(pending_empty)} 个关键词结果为空。")

        # 补爬空关键词，最多 empty_retry_rounds 轮
        for retry_round in range(1, empty_retry_rounds + 1):
            if not pending_empty:
                break
            if should_stop(stop_event):
                log_line(log_callback, "任务已停止，跳过空关键词补爬。")
                break
            if wait_if_paused(pause_event, stop_event):
                break

            log_line(log_callback, f"开始第 {retry_round}/{empty_retry_rounds} 轮空关键词补爬：{len(pending_empty)} 个关键词。")
            cooldown = random.uniform(empty_retry_cooldown_min, empty_retry_cooldown_max)
            if cooldown > 0:
                log_line(log_callback, f"补爬前冷却 {cooldown:.1f} 秒，降低继续触发风控概率。")
                if interruptible_sleep(cooldown, stop_event):
                    break

            retry_paths = run_keyword_batch(pending_empty, f"补爬{retry_round}")
            latest_path_by_keyword.update(retry_paths)
            pending_empty = empty_keywords()
            if pending_empty:
                log_line(log_callback, f"第 {retry_round} 轮补爬后仍有 {len(pending_empty)} 个关键词为空。")
            else:
                log_line(log_callback, "空关键词补爬完成：所有关键词均已有数据。")

        # 最终只把非空输出传给 UI 汇总，避免 summary 被空表污染
        nonempty_paths = []
        still_empty = []
        for kw in keywords_list:
            path = latest_path_by_keyword.get(kw)
            if path and not is_x_keyword_output_empty(path):
                nonempty_paths.append(path)
            else:
                still_empty.append(kw)

        log_line(log_callback, f"\nX 关键词媒体推文搜索任务结束。有效关键词 {len(nonempty_paths)}/{len(keywords_list)} 个。")
        if still_empty:
            log_line(log_callback, f"仍为空的关键词：{', '.join(still_empty[:20])}{' ...' if len(still_empty) > 20 else ''}")
        for p in nonempty_paths:
            log_line(log_callback, f"  {p}")
        finish_callback(nonempty_paths if nonempty_paths else None)

    except Exception as e:
        log_error(log_callback, f"发生致命错误：{e}")
        finish_callback()
