# -*- coding: utf-8 -*-
"""YouTube 关键词搜索结果快照工具。

读取已有关键词搜索生成的 Excel 文件，提取其中的视频 ID，
重新查询最新的热度数据（播放、点赞、评论），并将新数据以"快照天数"为后缀追加到原表中保存。
"""

from __future__ import annotations

import os
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

from src.core import build_output_path, log_error, log_line, log_warn, should_stop, wait_if_paused
from src.platforms.youtube.keyword import YouTubeClientPool, _api_call_with_rotation, chunked
from src.platforms.youtube.comments import extract_video_id


def run_youtube_snapshot(api_keys: list[str], xlsx_path: str, snapshot_days: str, log_callback, finish_callback, stop_event=None, config=None, pause_event=None):
    """运行快照更新的主流程。"""
    try:
        if not os.path.exists(xlsx_path):
            log_error(log_callback, f"找不到文件: {xlsx_path}")
            return

        suffix = f"_{snapshot_days}"
        log_line(log_callback, f"=== 开始 YouTube 视频热度快照 ({snapshot_days}) ===")
        log_line(log_callback, f"目标文件: {xlsx_path}")

        # 使用 openpyxl 读取并解析
        try:
            wb = load_workbook(xlsx_path)
            # 找到含有 "视频信息" 或者是 default sheet 的表格
            sheet_name = "视频信息" if "视频信息" in wb.sheetnames else wb.sheetnames[0]
            ws = wb[sheet_name]
        except Exception as e:
            log_error(log_callback, f"读取 Excel 失败，请检查文件格式: {e}")
            return

        # 解析表头
        headers = [cell.value for cell in ws[1]]
        if "视频链接" not in headers:
            log_error(log_callback, "Excel 中未找到「视频链接」列，无法提取视频 ID。")
            return
            
        link_col_idx = headers.index("视频链接")

        # 检查是否已存在当前天数的快照列，如果存在则获取索引，不存在则追加表头
        target_fields = ["播放量", "点赞数", "评论数", "查询时间"]
        snapshot_headers = [f"{field}{suffix}" for field in target_fields]
        
        # 扩展表头
        header_indices = {}
        for sh in snapshot_headers:
            if sh in headers:
                header_indices[sh] = headers.index(sh) + 1  # 1-based for openpyxl
            else:
                ws.cell(row=1, column=len(headers) + 1, value=sh)
                headers.append(sh)
                header_indices[sh] = len(headers)

        # 提取所有视频链接
        row_video_map = {}  # {row_idx: video_id}
        for row_idx in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row_idx, column=link_col_idx + 1).value
            if cell_val:
                vid = extract_video_id(str(cell_val))
                if vid:
                    row_video_map[row_idx] = vid

        if not row_video_map:
            log_warn(log_callback, "未能在表格中提取到任何有效的视频 ID。")
            return

        log_line(log_callback, f"成功解析 {len(row_video_map)} 个视频，开始获取最新数据...")

        client_pool = YouTubeClientPool(api_keys)
        unique_vids = list(set(row_video_map.values()))
        metrics_map = {}  # {video_id: {viewCount, likeCount, commentCount}}
        batch_size = 50

        # 分批查询最新热度
        written_count = 0
        for batch_ids in chunked(unique_vids, batch_size):
            if should_stop(stop_event) or wait_if_paused(pause_event, stop_event):
                break

            try:
                response = _api_call_with_rotation(
                    client_pool,
                    lambda ids=batch_ids: client_pool.client.videos().list(
                        part="statistics",
                        id=",".join(ids),
                        maxResults=batch_size,
                        fields="items(id,statistics(viewCount,likeCount,commentCount))"
                    ),
                    log_callback,
                    stop_event,
                )
            except Exception as e:
                import googleapiclient.errors
                if isinstance(e, googleapiclient.errors.HttpError) and e.resp.status in [403, 429]:
                    log_warn(log_callback, f"API 配额耗尽或受限: {e}")
                    break
                log_warn(log_callback, f"API 请求失败: {e}")
                continue

            for item in response.get("items", []):
                vid = item.get("id")
                stats = item.get("statistics", {})
                if vid:
                    metrics_map[vid] = {
                        "播放量": stats.get("viewCount", ""),
                        "点赞数": stats.get("likeCount", ""),
                        "评论数": stats.get("commentCount", ""),
                    }
            
            written_count += len(batch_ids)
            log_line(log_callback, f"  已获取 {written_count}/{len(unique_vids)} 个视频数据")

        # 将数据写回表格
        current_time = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        for row_idx, vid in row_video_map.items():
            metrics = metrics_map.get(vid)
            if metrics:
                ws.cell(row=row_idx, column=header_indices[f"播放量{suffix}"], value=metrics["播放量"])
                ws.cell(row=row_idx, column=header_indices[f"点赞数{suffix}"], value=metrics["点赞数"])
                ws.cell(row=row_idx, column=header_indices[f"评论数{suffix}"], value=metrics["评论数"])
                ws.cell(row=row_idx, column=header_indices[f"查询时间{suffix}"], value=current_time)

        # 另存为新文件（不覆盖原文件）
        orig_path = Path(xlsx_path)
        stem = orig_path.stem
        # 如果原始文件名已经包含了后缀，防止重复叠加
        if not stem.endswith(f"_snapshot{suffix}"):
            new_stem = f"{stem}_snapshot{suffix}"
        else:
            new_stem = stem
        
        new_filename = f"{new_stem}{orig_path.suffix}"
        # 保存到当前任务的独立输出路径
        output_path = build_output_path("youtube", new_filename, channel="snapshot")
        
        # 保存原子化
        temp_path = f"{output_path}.tmp"
        wb.save(temp_path)
        try:
            os.replace(temp_path, output_path)
        except OSError:
            wb.save(output_path)
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except OSError:
                    pass
            
        log_line(log_callback, f"快照更新完成！另存为新文件：\n  {output_path}")

    except Exception as exc:
        log_error(log_callback, f"运行失败：{exc}")
        output_path = None
        raise
    finally:
        finish_callback(output_path if 'output_path' in locals() else None)
