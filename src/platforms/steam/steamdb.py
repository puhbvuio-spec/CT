# -*- coding: utf-8 -*-
"""Dynamic-window SteamDB visible-page collector.

This module intentionally uses a user-controlled Chromium window through CDP.
It does not request SteamDB pages with requests/http clients and does not try to
bypass browser checks. If SteamDB shows a browser check or block page, the task
logs an instruction and pauses so the user can handle it in the opened browser.
"""

from __future__ import annotations

import json
import re
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus

from src.core import (
    build_output_path,
    cdp_url_for_browser,
    connect_existing_chromium,
    interruptible_sleep,
    log_error,
    log_line,
    log_warn,
    should_stop,
    wait_if_paused,
)
from src.core.task_checkpoint import open_checkpointed_multi_sheet_writer, open_task_checkpoint
from src.platforms.steam.api import discover_apps_for_keyword, normalize_keywords, parse_app_ids

BASE_URL = "https://steamdb.info"

STEAMDB_STATUS_FIELDS = [
    "来源类型",
    "搜索词",
    "AppID",
    "页面类型",
    "SteamDB URL",
    "最终URL",
    "页面标题",
    "H1",
    "状态",
    "备注",
    "可见表格数",
    "可见行数",
    "查询时间",
]

STEAMDB_OVERVIEW_FIELDS = [
    "来源类型",
    "搜索词",
    "AppID",
    "页面标题",
    "H1",
    "游戏名",
    "发行日期",
    "开发商",
    "发行商",
    "当前在线",
    "24小时峰值",
    "历史峰值",
    "评分/评价",
    "标签/类型",
    "摘要文本",
    "SteamDB URL",
    "最终URL",
    "状态",
    "备注",
    "查询时间",
]

STEAMDB_CHARTS_FIELDS = [
    "来源类型",
    "搜索词",
    "AppID",
    "页面标题",
    "图表/表格",
    "行号",
    "排名",
    "名称/时间",
    "数值1",
    "数值2",
    "数值3",
    "可见指标文本",
    "实体链接",
    "SteamDB URL",
    "状态",
    "备注",
    "查询时间",
]

STEAMDB_PACKAGE_FIELDS = [
    "来源类型",
    "搜索词",
    "AppID",
    "PackageID",
    "套餐名称",
    "价格/区域",
    "折扣/备注",
    "可见指标文本",
    "套餐链接",
    "SteamDB URL",
    "状态",
    "备注",
    "查询时间",
]

STEAMDB_DLC_FIELDS = [
    "来源类型",
    "搜索词",
    "AppID",
    "DLC AppID",
    "DLC名称",
    "发行日期",
    "价格/备注",
    "可见指标文本",
    "DLC链接",
    "SteamDB URL",
    "状态",
    "备注",
    "查询时间",
]

STEAMDB_DEPOT_FIELDS = [
    "来源类型",
    "搜索词",
    "AppID",
    "DepotID",
    "Depot名称",
    "大小/系统",
    "分支/备注",
    "可见指标文本",
    "Depot链接",
    "SteamDB URL",
    "状态",
    "备注",
    "查询时间",
]

STEAMDB_HISTORY_FIELDS = [
    "来源类型",
    "搜索词",
    "AppID",
    "变更时间",
    "变更ID",
    "变更类型",
    "摘要",
    "变更链接",
    "SteamDB URL",
    "状态",
    "备注",
    "查询时间",
]

STEAMDB_RAW_TABLE_FIELDS = [
    "来源类型",
    "搜索词",
    "AppID",
    "页面类型",
    "SteamDB URL",
    "表格序号",
    "表格标题",
    "行号",
    "表头",
    "单元格JSON",
    "链接JSON",
    "可见文本",
    "状态",
    "备注",
    "查询时间",
]

STEAMDB_SHEETS_FIELDS = {
    "采集状态": STEAMDB_STATUS_FIELDS,
    "概览": STEAMDB_OVERVIEW_FIELDS,
    "Charts": STEAMDB_CHARTS_FIELDS,
    "Packages": STEAMDB_PACKAGE_FIELDS,
    "DLCs": STEAMDB_DLC_FIELDS,
    "Depots": STEAMDB_DEPOT_FIELDS,
    "History": STEAMDB_HISTORY_FIELDS,
    "RawTables": STEAMDB_RAW_TABLE_FIELDS,
}

PAGE_LABELS = {
    "overview": "概览页",
    "charts": "Charts页",
    "packages": "Packages页",
    "dlcs": "DLCs页",
    "depots": "Depots页",
    "history": "History页",
}

BLOCK_PATTERNS = [
    "checking your browser",
    "please wait while we check your browser",
    "browser check",
    "cf-browser-verification",
    "cloudflare ray id",
    "ray id",
    "attention required",
    "access denied",
    "forbidden",
    "stop. do not make any further requests",
    "do not make any further requests",
]


@dataclass(frozen=True)
class SteamDbWorkItem:
    appid: int
    source_type: str
    source: str = ""
    seed_name: str = ""

    @property
    def checkpoint_key(self) -> str:
        return str(self.appid).lower()


@dataclass
class SteamDbPageSnapshot:
    item: SteamDbWorkItem
    page_type: str
    url: str
    final_url: str = ""
    title: str = ""
    h1: str = ""
    body_text: str = ""
    tables: list[dict[str, Any]] = field(default_factory=list)
    status: str = "ok"
    note: str = ""
    query_time: str = ""


@dataclass
class SteamDbBundle:
    status_rows: list[dict[str, Any]] = field(default_factory=list)
    overview_rows: list[dict[str, Any]] = field(default_factory=list)
    charts_rows: list[dict[str, Any]] = field(default_factory=list)
    package_rows: list[dict[str, Any]] = field(default_factory=list)
    dlc_rows: list[dict[str, Any]] = field(default_factory=list)
    depot_rows: list[dict[str, Any]] = field(default_factory=list)
    history_rows: list[dict[str, Any]] = field(default_factory=list)
    raw_rows: list[dict[str, Any]] = field(default_factory=list)

    @property
    def meta(self) -> dict[str, int]:
        return {
            "status_rows": len(self.status_rows),
            "overview_rows": len(self.overview_rows),
            "charts_rows": len(self.charts_rows),
            "package_rows": len(self.package_rows),
            "dlc_rows": len(self.dlc_rows),
            "depot_rows": len(self.depot_rows),
            "history_rows": len(self.history_rows),
            "raw_rows": len(self.raw_rows),
        }


def now_text() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


def clean_text(value: Any, *, max_length: int | None = None) -> str:
    text = str(value or "").replace("\r", "\n")
    text = re.sub(r"[ \t\f\v]+", " ", text)
    text = re.sub(r"\n\s*\n+", "\n", text)
    text = text.strip()
    if max_length and len(text) > max_length:
        return text[:max_length].rstrip()
    return text


def compact_text(value: Any, *, max_length: int | None = None) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if max_length and len(text) > max_length:
        return text[:max_length].rstrip()
    return text


def parse_steamdb_app_ids(value: str | list[str] | tuple[str, ...]) -> list[int]:
    """Extract Steam AppIDs from SteamDB URLs, Steam store URLs, or bare numbers."""
    if isinstance(value, str):
        lines = value.splitlines()
    else:
        lines = list(value or [])
    appids: list[int] = []
    seen: set[int] = set()
    for raw in lines:
        text = str(raw or "").strip()
        if not text:
            continue
        matches = re.findall(r"(?:steamdb\.info/app/|store\.steampowered\.com/app/|/app/|^)(\d{2,12})(?:\D|$)", text, flags=re.I)
        if not matches and text.isdigit():
            matches = [text]
        for value_text in matches:
            try:
                appid = int(value_text)
            except ValueError:
                continue
            if appid not in seen:
                seen.add(appid)
                appids.append(appid)
    return appids


def build_steamdb_app_url(appid: int, page_type: str = "overview") -> str:
    appid = int(appid)
    page = str(page_type or "overview").strip().lower()
    if page == "overview":
        return f"{BASE_URL}/app/{appid}/"
    if page == "charts":
        return f"{BASE_URL}/app/{appid}/charts/"
    if page in {"packages", "subs"}:
        return f"{BASE_URL}/app/{appid}/subs/"
    if page in {"dlcs", "dlc"}:
        return f"{BASE_URL}/app/{appid}/dlc/"
    if page == "depots":
        # Do not add branch query parameters; /app/*/depots/?branch=* is avoided.
        return f"{BASE_URL}/app/{appid}/depots/"
    if page == "history":
        # Do not add changeid query parameters; /app/*/history/?changeid=* is avoided.
        return f"{BASE_URL}/app/{appid}/history/"
    raise ValueError(f"Unsupported SteamDB page type: {page_type}")


def detect_steamdb_block(title: str = "", body_text: str = "") -> str:
    haystack = f"{title}\n{body_text}".lower()
    for pattern in BLOCK_PATTERNS:
        if pattern in haystack:
            return pattern
    if "steamdb" in haystack and "stop" in haystack and "requests" in haystack:
        return "stop/requests"
    return ""


def read_app_ids_from_xlsx(path: str, limit: int = 0) -> list[int]:
    """Read AppID-like columns from an existing workbook without scraping SteamDB."""
    xlsx_path = str(path or "").strip()
    if not xlsx_path:
        return []
    workbook_path = Path(xlsx_path)
    if not workbook_path.exists():
        raise ValueError(f"Excel 文件不存在：{xlsx_path}")

    from openpyxl import load_workbook

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    preferred_headers = {"appid", "app id", "应用id", "目标appid", "游戏appid", "dlc appid"}
    appids: list[int] = []
    seen: set[int] = set()
    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        try:
            header = [compact_text(value).lower() for value in next(rows)]
        except StopIteration:
            continue
        indexes = [idx for idx, name in enumerate(header) if name in preferred_headers or name.replace(" ", "") == "appid"]
        if not indexes:
            continue
        for row in rows:
            for idx in indexes:
                if idx >= len(row):
                    continue
                for appid in parse_steamdb_app_ids(str(row[idx] or "")):
                    if appid in seen:
                        continue
                    seen.add(appid)
                    appids.append(appid)
                    if limit and len(appids) >= limit:
                        return appids
    return appids


def _absolute_steamdb_url(href: str) -> str:
    text = str(href or "").strip()
    if not text:
        return ""
    if text.startswith("http://") or text.startswith("https://"):
        return text
    if text.startswith("//"):
        return f"https:{text}"
    if text.startswith("/"):
        return f"{BASE_URL}{text}"
    return f"{BASE_URL}/{text.lstrip('/')}"


def _first_link_for(row: dict[str, Any], pattern: str) -> str:
    regex = re.compile(pattern, flags=re.I)
    for link in row.get("links", []) or []:
        href = str((link or {}).get("href") or "")
        if regex.search(href):
            return _absolute_steamdb_url(href)
    return ""


def _first_id_from_row(row: dict[str, Any], pattern: str) -> str:
    regex = re.compile(pattern, flags=re.I)
    text = " ".join(str(cell or "") for cell in row.get("cells", []) or [])
    for link in row.get("links", []) or []:
        text += " " + str((link or {}).get("href") or "")
    match = regex.search(text)
    return match.group(1) if match else ""


def _first_nonempty(values: list[Any]) -> str:
    for value in values:
        text = compact_text(value)
        if text:
            return text
    return ""


def _cells(row: dict[str, Any]) -> list[str]:
    return [compact_text(cell) for cell in row.get("cells", []) or []]


def _row_visible_text(row: dict[str, Any], *, max_length: int = 4000) -> str:
    return compact_text(" | ".join(_cells(row)), max_length=max_length)


def _rows_from_tables(snapshot: SteamDbPageSnapshot) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for table in snapshot.tables:
        for row in table.get("rows", []) or []:
            rows.append({"table": table, "row": row})
    return rows


def _extract_tables_from_page(page, max_rows_per_table: int = 100) -> list[dict[str, Any]]:
    return page.evaluate(
        """
        (maxRowsPerTable) => {
            const visible = (el) => {
                if (!el) return false;
                const style = window.getComputedStyle(el);
                if (style && (style.display === 'none' || style.visibility === 'hidden')) return false;
                const rects = el.getClientRects();
                return rects && rects.length > 0;
            };
            const clean = (text) => (text || '').replace(/\\s+/g, ' ').trim();
            const tableTitle = (table, index) => {
                const caption = table.querySelector('caption');
                if (caption && clean(caption.innerText || caption.textContent)) return clean(caption.innerText || caption.textContent);
                let node = table.previousElementSibling;
                for (let i = 0; node && i < 4; i++, node = node.previousElementSibling) {
                    const tag = (node.tagName || '').toLowerCase();
                    if (/^h[1-6]$/.test(tag) || node.className) {
                        const text = clean(node.innerText || node.textContent);
                        if (text && text.length <= 160) return text;
                    }
                }
                return `table_${index + 1}`;
            };
            const tables = Array.from(document.querySelectorAll('table')).filter(visible);
            return tables.map((table, tableIndex) => {
                const headers = Array.from(table.querySelectorAll('thead th')).map(th => clean(th.innerText || th.textContent)).filter(Boolean);
                const rowNodes = Array.from(table.querySelectorAll('tbody tr'));
                const fallbackRows = rowNodes.length ? rowNodes : Array.from(table.querySelectorAll('tr')).slice(headers.length ? 1 : 0);
                const rows = [];
                for (const tr of fallbackRows) {
                    if (!visible(tr)) continue;
                    const cellNodes = Array.from(tr.querySelectorAll('th,td'));
                    const cells = cellNodes.map(td => clean(td.innerText || td.textContent)).filter(Boolean);
                    if (!cells.length) continue;
                    const links = Array.from(tr.querySelectorAll('a[href]')).map(a => ({
                        text: clean(a.innerText || a.textContent),
                        href: a.getAttribute('href') || '',
                        absolute: a.href || ''
                    })).filter(x => x.href || x.absolute);
                    rows.push({ row_index: rows.length + 1, cells, links });
                    if (rows.length >= maxRowsPerTable) break;
                }
                return { table_index: tableIndex + 1, title: tableTitle(table, tableIndex), headers, rows };
            }).filter(table => table.rows.length > 0);
        }
        """,
        max(1, int(max_rows_per_table or 1)),
    ) or []


def _extract_page_text(page) -> tuple[str, str, str]:
    try:
        title = page.title()
    except Exception:
        title = ""
    try:
        h1 = page.locator("h1").first.inner_text(timeout=2000)
    except Exception:
        h1 = ""
    try:
        body = page.locator("body").inner_text(timeout=5000)
    except Exception:
        body = ""
    return compact_text(title, max_length=300), compact_text(h1, max_length=300), clean_text(body, max_length=30000)


def _scroll_visible_page(page, max_scrolls: int, stop_event=None, pause_event=None) -> None:
    for _ in range(max(0, int(max_scrolls or 0))):
        if wait_if_paused(pause_event, stop_event) or should_stop(stop_event):
            return
        try:
            page.mouse.wheel(0, 900)
        except Exception:
            try:
                page.evaluate("window.scrollBy(0, 900)")
            except Exception:
                pass
        if interruptible_sleep(0.5, stop_event, pause_event=pause_event):
            return


def _pause_for_block(pattern: str, url: str, log_callback=None, stop_event=None, pause_event=None) -> None:
    message = (
        f"SteamDB 页面触发浏览器检查/阻止（{pattern}）：{url}\n"
        "已按配置暂停。请在打开的浏览器窗口中完成验证或确认页面状态，"
        "然后在工具窗口点击「暂停/继续」按钮一次恢复；不会自动绕过保护。"
    )
    log_warn(log_callback, message)
    if pause_event is None:
        return
    pause_event.set()
    while pause_event.is_set() and not should_stop(stop_event):
        time.sleep(0.5)


def load_steamdb_page(
    page,
    item: SteamDbWorkItem,
    page_type: str,
    *,
    page_timeout: int = 30000,
    after_load_wait: float = 2.0,
    max_scrolls: int = 1,
    max_table_rows_per_page: int = 100,
    block_handling: str = "只暂停提示",
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> SteamDbPageSnapshot:
    url = build_steamdb_app_url(item.appid, page_type)
    snapshot = SteamDbPageSnapshot(item=item, page_type=page_type, url=url, query_time=now_text())
    try:
        if wait_if_paused(pause_event, stop_event) or should_stop(stop_event):
            snapshot.status = "stopped"
            snapshot.note = "任务停止或暂停中断。"
            return snapshot
        page.goto(url, wait_until="domcontentloaded", timeout=max(1000, int(page_timeout or 30000)))
        if after_load_wait > 0:
            interruptible_sleep(float(after_load_wait), stop_event, pause_event=pause_event)
        try:
            page.wait_for_load_state("networkidle", timeout=min(10000, max(1000, int((page_timeout or 30000) / 2))))
        except Exception:
            pass
        snapshot.final_url = str(getattr(page, "url", "") or "")
        title, h1, body = _extract_page_text(page)
        block_pattern = detect_steamdb_block(title, body)
        if block_pattern and str(block_handling or "只暂停提示") == "只暂停提示":
            _pause_for_block(block_pattern, url, log_callback=log_callback, stop_event=stop_event, pause_event=pause_event)
            if should_stop(stop_event):
                snapshot.status = "stopped"
                snapshot.note = "阻止页暂停期间任务被停止。"
                return snapshot
            title, h1, body = _extract_page_text(page)
            block_pattern = detect_steamdb_block(title, body)
        snapshot.title = title
        snapshot.h1 = h1
        snapshot.body_text = body
        snapshot.final_url = str(getattr(page, "url", "") or snapshot.final_url or url)
        if block_pattern:
            snapshot.status = "blocked"
            snapshot.note = f"SteamDB 显示浏览器检查/阻止：{block_pattern}。已暂停提示，未绕过。"
            return snapshot
        if "404" in title.lower() or "not found" in body[:1200].lower():
            snapshot.status = "not_found"
            snapshot.note = "页面显示未找到。"
            return snapshot
        _scroll_visible_page(page, max_scrolls, stop_event=stop_event, pause_event=pause_event)
        snapshot.tables = _extract_tables_from_page(page, max_rows_per_table=max_table_rows_per_page)
        if not snapshot.tables and not snapshot.h1:
            snapshot.status = "empty"
            snapshot.note = "当前动态窗口未识别到标题或可见表格。"
        else:
            snapshot.status = "ok"
            snapshot.note = "仅采集当前动态窗口可见 DOM；未请求 SteamDB 原始接口，未做全站枚举。"
    except Exception as exc:
        snapshot.final_url = str(getattr(page, "url", "") or "")
        try:
            snapshot.title, snapshot.h1, snapshot.body_text = _extract_page_text(page)
        except Exception:
            pass
        snapshot.status = "error"
        snapshot.note = str(exc)[:500]
    return snapshot


def build_status_row(snapshot: SteamDbPageSnapshot) -> dict[str, Any]:
    row_count = sum(len(table.get("rows", []) or []) for table in snapshot.tables)
    return {
        "来源类型": snapshot.item.source_type,
        "搜索词": snapshot.item.source,
        "AppID": snapshot.item.appid,
        "页面类型": PAGE_LABELS.get(snapshot.page_type, snapshot.page_type),
        "SteamDB URL": snapshot.url,
        "最终URL": snapshot.final_url,
        "页面标题": snapshot.title,
        "H1": snapshot.h1,
        "状态": snapshot.status,
        "备注": snapshot.note,
        "可见表格数": len(snapshot.tables),
        "可见行数": row_count,
        "查询时间": snapshot.query_time,
    }


def _base_snapshot_row(snapshot: SteamDbPageSnapshot) -> dict[str, Any]:
    return {
        "来源类型": snapshot.item.source_type,
        "搜索词": snapshot.item.source,
        "AppID": snapshot.item.appid,
        "SteamDB URL": snapshot.url,
        "状态": snapshot.status,
        "备注": snapshot.note,
        "查询时间": snapshot.query_time,
    }


def _first_match(text: str, patterns: list[str]) -> str:
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I | re.S)
        if match:
            return compact_text(match.group(1), max_length=300)
    return ""


def _summary_lines(text: str, *, limit: int = 80) -> str:
    lines: list[str] = []
    seen: set[str] = set()
    for raw in clean_text(text).splitlines():
        line = compact_text(raw)
        if not line or line.lower() in seen:
            continue
        if len(line) < 2:
            continue
        seen.add(line.lower())
        lines.append(line)
        if len(lines) >= limit:
            break
    return clean_text("\n".join(lines), max_length=8000)


def build_overview_row(snapshot: SteamDbPageSnapshot) -> dict[str, Any]:
    row = _base_snapshot_row(snapshot)
    text = snapshot.body_text
    title_name = snapshot.h1 or _first_match(snapshot.title, [r"^(.*?)\s+SteamDB", r"^(.*?)\s+on SteamDB"])
    row.update(
        {
            "页面标题": snapshot.title,
            "H1": snapshot.h1,
            "游戏名": title_name,
            "发行日期": _first_match(text, [r"Release Date\s*[:\n ]\s*([^\n]+)", r"Released\s*[:\n ]\s*([^\n]+)"]),
            "开发商": _first_match(text, [r"Developer\s*[:\n ]\s*([^\n]+)", r"Developers\s*[:\n ]\s*([^\n]+)"]),
            "发行商": _first_match(text, [r"Publisher\s*[:\n ]\s*([^\n]+)", r"Publishers\s*[:\n ]\s*([^\n]+)"]),
            "当前在线": _first_match(text, [r"([0-9][0-9,]*)\s+players?\s+right now", r"Current Players\s*[:\n ]\s*([0-9][0-9,]*)"]),
            "24小时峰值": _first_match(text, [r"([0-9][0-9,]*)\s+24-hour peak", r"24-hour peak\s*[:\n ]\s*([0-9][0-9,]*)"]),
            "历史峰值": _first_match(text, [r"([0-9][0-9,]*)\s+all-time peak", r"All-time peak\s*[:\n ]\s*([0-9][0-9,]*)"]),
            "评分/评价": _first_match(text, [r"Reviews?\s*[:\n ]\s*([^\n]+)", r"([0-9]{1,3}(?:\.[0-9]+)?%[^\n]{0,120})"]),
            "标签/类型": _first_match(text, [r"Genres?\s*[:\n ]\s*([^\n]+)", r"Tags\s*[:\n ]\s*([^\n]+)"]),
            "摘要文本": _summary_lines(text),
            "最终URL": snapshot.final_url,
        }
    )
    return row


def build_raw_table_rows(snapshot: SteamDbPageSnapshot) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for table in snapshot.tables:
        headers = [compact_text(header) for header in table.get("headers", []) or []]
        for row in table.get("rows", []) or []:
            cells = _cells(row)
            links = row.get("links", []) or []
            rows.append(
                {
                    "来源类型": snapshot.item.source_type,
                    "搜索词": snapshot.item.source,
                    "AppID": snapshot.item.appid,
                    "页面类型": PAGE_LABELS.get(snapshot.page_type, snapshot.page_type),
                    "SteamDB URL": snapshot.url,
                    "表格序号": table.get("table_index", ""),
                    "表格标题": compact_text(table.get("title")),
                    "行号": row.get("row_index", ""),
                    "表头": " | ".join(headers),
                    "单元格JSON": json.dumps(cells, ensure_ascii=False),
                    "链接JSON": json.dumps(links, ensure_ascii=False),
                    "可见文本": _row_visible_text(row, max_length=4000),
                    "状态": snapshot.status,
                    "备注": snapshot.note,
                    "查询时间": snapshot.query_time,
                }
            )
    return rows


def build_chart_rows(snapshot: SteamDbPageSnapshot) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for entry in _rows_from_tables(snapshot):
        table = entry["table"]
        row = entry["row"]
        cells = _cells(row)
        rank = cells[0] if cells and re.match(r"^#?\d+", cells[0]) else ""
        name = cells[1] if rank and len(cells) > 1 else (cells[0] if cells else "")
        values = cells[2:] if rank else cells[1:]
        out.append(
            {
                "来源类型": snapshot.item.source_type,
                "搜索词": snapshot.item.source,
                "AppID": snapshot.item.appid,
                "页面标题": snapshot.title,
                "图表/表格": compact_text(table.get("title")),
                "行号": row.get("row_index", ""),
                "排名": rank,
                "名称/时间": name,
                "数值1": values[0] if len(values) > 0 else "",
                "数值2": values[1] if len(values) > 1 else "",
                "数值3": values[2] if len(values) > 2 else "",
                "可见指标文本": _row_visible_text(row),
                "实体链接": _first_nonempty([_absolute_steamdb_url((link or {}).get("href", "")) for link in row.get("links", []) or []]),
                "SteamDB URL": snapshot.url,
                "状态": snapshot.status,
                "备注": snapshot.note,
                "查询时间": snapshot.query_time,
            }
        )
    if not out:
        out.append({**_base_snapshot_row(snapshot), "页面标题": snapshot.title, "图表/表格": "", "行号": "", "排名": "", "名称/时间": snapshot.h1, "数值1": "", "数值2": "", "数值3": "", "可见指标文本": _summary_lines(snapshot.body_text, limit=20), "实体链接": ""})
    return out


def build_package_rows(snapshot: SteamDbPageSnapshot) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for entry in _rows_from_tables(snapshot):
        row = entry["row"]
        cells = _cells(row)
        package_id = _first_id_from_row(row, r"/sub/(\d+)/") or (cells[0] if cells and cells[0].isdigit() else "")
        name_index = 1 if package_id and len(cells) > 1 else 0
        out.append(
            {
                "来源类型": snapshot.item.source_type,
                "搜索词": snapshot.item.source,
                "AppID": snapshot.item.appid,
                "PackageID": package_id,
                "套餐名称": cells[name_index] if len(cells) > name_index else "",
                "价格/区域": cells[name_index + 1] if len(cells) > name_index + 1 else "",
                "折扣/备注": cells[name_index + 2] if len(cells) > name_index + 2 else "",
                "可见指标文本": _row_visible_text(row),
                "套餐链接": _first_link_for(row, r"/sub/\d+"),
                "SteamDB URL": snapshot.url,
                "状态": snapshot.status,
                "备注": snapshot.note,
                "查询时间": snapshot.query_time,
            }
        )
    return out


def build_dlc_rows(snapshot: SteamDbPageSnapshot) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for entry in _rows_from_tables(snapshot):
        row = entry["row"]
        cells = _cells(row)
        dlc_appid = _first_id_from_row(row, r"/app/(\d+)/") or (cells[0] if cells and cells[0].isdigit() else "")
        name_index = 1 if dlc_appid and len(cells) > 1 else 0
        out.append(
            {
                "来源类型": snapshot.item.source_type,
                "搜索词": snapshot.item.source,
                "AppID": snapshot.item.appid,
                "DLC AppID": dlc_appid,
                "DLC名称": cells[name_index] if len(cells) > name_index else "",
                "发行日期": cells[name_index + 1] if len(cells) > name_index + 1 else "",
                "价格/备注": cells[name_index + 2] if len(cells) > name_index + 2 else "",
                "可见指标文本": _row_visible_text(row),
                "DLC链接": _first_link_for(row, r"/app/\d+"),
                "SteamDB URL": snapshot.url,
                "状态": snapshot.status,
                "备注": snapshot.note,
                "查询时间": snapshot.query_time,
            }
        )
    return out


def build_depot_rows(snapshot: SteamDbPageSnapshot) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for entry in _rows_from_tables(snapshot):
        row = entry["row"]
        cells = _cells(row)
        depot_id = _first_id_from_row(row, r"/depot/(\d+)/") or (cells[0] if cells and cells[0].isdigit() else "")
        name_index = 1 if depot_id and len(cells) > 1 else 0
        out.append(
            {
                "来源类型": snapshot.item.source_type,
                "搜索词": snapshot.item.source,
                "AppID": snapshot.item.appid,
                "DepotID": depot_id,
                "Depot名称": cells[name_index] if len(cells) > name_index else "",
                "大小/系统": cells[name_index + 1] if len(cells) > name_index + 1 else "",
                "分支/备注": cells[name_index + 2] if len(cells) > name_index + 2 else "",
                "可见指标文本": _row_visible_text(row),
                "Depot链接": _first_link_for(row, r"/depot/\d+"),
                "SteamDB URL": snapshot.url,
                "状态": snapshot.status,
                "备注": snapshot.note,
                "查询时间": snapshot.query_time,
            }
        )
    return out


def build_history_rows(snapshot: SteamDbPageSnapshot) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for entry in _rows_from_tables(snapshot):
        row = entry["row"]
        cells = _cells(row)
        change_id = _first_id_from_row(row, r"changeid=(\d+)") or _first_id_from_row(row, r"/history/(\d+)")
        change_link = _first_link_for(row, r"/history|changeid=")
        out.append(
            {
                "来源类型": snapshot.item.source_type,
                "搜索词": snapshot.item.source,
                "AppID": snapshot.item.appid,
                "变更时间": cells[0] if cells else "",
                "变更ID": change_id,
                "变更类型": cells[1] if len(cells) > 1 else "",
                "摘要": " | ".join(cells[2:]) if len(cells) > 2 else _row_visible_text(row),
                "变更链接": change_link,
                "SteamDB URL": snapshot.url,
                "状态": snapshot.status,
                "备注": snapshot.note,
                "查询时间": snapshot.query_time,
            }
        )
    return out


def collect_steamdb_app_bundle(
    page,
    item: SteamDbWorkItem,
    *,
    page_types: list[str],
    page_timeout: int,
    after_load_wait: float,
    page_delay: float,
    max_scrolls: int,
    max_table_rows_per_page: int,
    block_handling: str,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> SteamDbBundle:
    bundle = SteamDbBundle()
    for index, page_type in enumerate(page_types):
        if wait_if_paused(pause_event, stop_event) or should_stop(stop_event):
            break
        snapshot = load_steamdb_page(
            page,
            item,
            page_type,
            page_timeout=page_timeout,
            after_load_wait=after_load_wait,
            max_scrolls=max_scrolls,
            max_table_rows_per_page=max_table_rows_per_page,
            block_handling=block_handling,
            log_callback=log_callback,
            stop_event=stop_event,
            pause_event=pause_event,
        )
        bundle.status_rows.append(build_status_row(snapshot))
        bundle.raw_rows.extend(build_raw_table_rows(snapshot))
        if page_type == "overview":
            bundle.overview_rows.append(build_overview_row(snapshot))
        elif page_type == "charts":
            bundle.charts_rows.extend(build_chart_rows(snapshot))
        elif page_type == "packages":
            bundle.package_rows.extend(build_package_rows(snapshot))
        elif page_type == "dlcs":
            bundle.dlc_rows.extend(build_dlc_rows(snapshot))
        elif page_type == "depots":
            bundle.depot_rows.extend(build_depot_rows(snapshot))
        elif page_type == "history":
            bundle.history_rows.extend(build_history_rows(snapshot))
        if snapshot.status == "blocked":
            log_warn(log_callback, f"SteamDB {PAGE_LABELS.get(page_type, page_type)} 仍处于阻止页：AppID={item.appid}")
        if index < len(page_types) - 1 and page_delay > 0:
            if interruptible_sleep(float(page_delay), stop_event, pause_event=pause_event):
                break
    return bundle


def _yes(value: Any, default: str = "否") -> bool:
    return str(value if value is not None else default).strip() == "是"


def _dedupe_items(items: list[SteamDbWorkItem], limit: int = 0) -> list[SteamDbWorkItem]:
    deduped: list[SteamDbWorkItem] = []
    seen: set[int] = set()
    for item in items:
        if not item.appid or item.appid in seen:
            continue
        seen.add(item.appid)
        deduped.append(item)
        if limit and len(deduped) >= limit:
            break
    return deduped


def prepare_steamdb_work_items(
    app_ids_or_urls: str | list[str],
    keywords: str | list[str],
    xlsx_path: str = "",
    *,
    max_apps_per_keyword: int = 20,
    max_apps: int = 100,
    language: str = "english",
    country: str = "US",
    config: dict[str, Any] | None = None,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> list[SteamDbWorkItem]:
    config = config or {}
    items: list[SteamDbWorkItem] = []
    for appid in parse_steamdb_app_ids(app_ids_or_urls):
        items.append(SteamDbWorkItem(appid=appid, source_type="直接输入"))
    if xlsx_path:
        for appid in read_app_ids_from_xlsx(xlsx_path):
            items.append(SteamDbWorkItem(appid=appid, source_type="Excel", source=Path(str(xlsx_path)).name))
    for keyword in normalize_keywords(keywords):
        if wait_if_paused(pause_event, stop_event) or should_stop(stop_event):
            break
        try:
            log_line(log_callback, f"SteamDB 关键词先用 Steam 商店公开接口发现 AppID：{keyword}")
            found = discover_apps_for_keyword(
                keyword,
                max_apps=max(1, int(max_apps_per_keyword or 1)),
                language=language,
                country=country,
                config=config,
                log_callback=log_callback,
                stop_event=stop_event,
                pause_event=pause_event,
            )
            for work in found:
                items.append(SteamDbWorkItem(appid=work.appid, source_type="关键词", source=keyword, seed_name=work.seed_name))
        except Exception as exc:
            log_error(log_callback, f"SteamDB 关键词发现失败：{keyword}: {exc}")
    return _dedupe_items(items, max(0, int(max_apps or 0)))


def _selected_page_types(config: dict[str, Any]) -> list[str]:
    page_types: list[str] = []
    if _yes(config.get("collect_overview"), "是"):
        page_types.append("overview")
    if _yes(config.get("collect_charts"), "是"):
        page_types.append("charts")
    if _yes(config.get("collect_packages"), "是"):
        page_types.append("packages")
    if _yes(config.get("collect_dlcs"), "是"):
        page_types.append("dlcs")
    if _yes(config.get("collect_depots"), "是"):
        page_types.append("depots")
    if _yes(config.get("collect_history"), "是"):
        page_types.append("history")
    return page_types or ["overview"]


def _write_bundle(writer, bundle: SteamDbBundle) -> None:
    for row in bundle.status_rows:
        writer.writerow("采集状态", row)
    for row in bundle.overview_rows:
        writer.writerow("概览", row)
    for row in bundle.charts_rows:
        writer.writerow("Charts", row)
    for row in bundle.package_rows:
        writer.writerow("Packages", row)
    for row in bundle.dlc_rows:
        writer.writerow("DLCs", row)
    for row in bundle.depot_rows:
        writer.writerow("Depots", row)
    for row in bundle.history_rows:
        writer.writerow("History", row)
    for row in bundle.raw_rows:
        writer.writerow("RawTables", row)


def run_steamdb_dynamic_window_spider(
    app_ids_or_urls: str | list[str],
    keywords: str | list[str],
    xlsx_path: str,
    log_callback,
    finish_callback,
    stop_event,
    *,
    pause_event=None,
    config: dict[str, Any] | None = None,
):
    config = config or {}
    language = str(config.get("language", "english") or "english")
    country = str(config.get("country", "US") or "US").upper()
    max_apps_per_keyword = max(1, int(config.get("max_apps_per_keyword", 20) or 20))
    max_apps = max(1, int(config.get("max_apps", 100) or 100))
    page_types = _selected_page_types(config)
    items = prepare_steamdb_work_items(
        app_ids_or_urls,
        keywords,
        xlsx_path,
        max_apps_per_keyword=max_apps_per_keyword,
        max_apps=max_apps,
        language=language,
        country=country,
        config=config,
        log_callback=log_callback,
        stop_event=stop_event,
        pause_event=pause_event,
    )
    if not items:
        raise ValueError("至少需要输入一个 SteamDB/Steam AppID/链接、关键词，或包含 AppID 的 Excel。")

    scope = {
        "app_ids": [item.appid for item in items],
        "keywords": normalize_keywords(keywords),
        "xlsx_path": str(xlsx_path or ""),
        "page_types": page_types,
        "max_apps_per_keyword": max_apps_per_keyword,
        "max_apps": max_apps,
        "language": language,
        "country": country,
    }
    checkpoint = open_task_checkpoint(
        "steamdb_dynamic_window",
        scope,
        log_callback,
        merge_on_keys=("app_ids", "keywords"),
        merge_keep_keys=("page_types", "language", "country"),
    )
    default_output_path = build_output_path(
        "steam",
        f"steamdb_dynamic_window_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
        channel="steamdb_dynamic_window",
    )
    save_batch_size = max(1, int(config.get("save_batch_size", 5) or 5))
    output_path, writer = open_checkpointed_multi_sheet_writer(
        checkpoint,
        default_output_path,
        STEAMDB_SHEETS_FIELDS,
        log_callback,
        autosave_every=save_batch_size,
    )
    checkpoint.add_output_path(output_path)

    browser_choice = str(config.get("steamdb_browser", config.get("browser", "chrome")) or "chrome")
    page_timeout = max(1000, int(config.get("page_timeout", 30000) or 30000))
    after_load_wait = max(0.0, float(config.get("after_load_wait", 2.0) or 0.0))
    page_delay = max(0.0, float(config.get("page_delay", 5.0) or 0.0))
    max_scrolls = max(0, int(config.get("max_scrolls", 1) or 0))
    max_table_rows_per_page = max(1, int(config.get("max_table_rows_per_page", 100) or 100))
    block_handling = str(config.get("block_handling", "只暂停提示") or "只暂停提示")

    completed_this_run = 0
    try:
        from playwright.sync_api import sync_playwright

        with sync_playwright() as playwright:
            resolved_cdp_url = cdp_url_for_browser(browser_choice)
            _, context = connect_existing_chromium(playwright, resolved_cdp_url, log_callback=log_callback, browser=browser_choice)
            page = context.new_page()
            try:
                log_line(
                    log_callback,
                    f"SteamDB 动态窗口任务候选 {len(items)} 个 App，页面：{', '.join(PAGE_LABELS.get(p, p) for p in page_types)}。",
                )
                for index, item in enumerate(items, start=1):
                    if wait_if_paused(pause_event, stop_event) or should_stop(stop_event):
                        break
                    claimed, claim_status = checkpoint.claim_item(item.checkpoint_key)
                    if not claimed:
                        if claim_status == "completed":
                            log_line(log_callback, f"断点跳过已完成 SteamDB AppID：{item.appid}")
                        elif claim_status == "active":
                            log_line(log_callback, f"分流跳过正在采集的 SteamDB AppID：{item.appid}")
                        continue
                    log_line(log_callback, f"[{index}/{len(items)}] SteamDB 动态采集 AppID={item.appid}")
                    try:
                        bundle = collect_steamdb_app_bundle(
                            page,
                            item,
                            page_types=page_types,
                            page_timeout=page_timeout,
                            after_load_wait=after_load_wait,
                            page_delay=page_delay,
                            max_scrolls=max_scrolls,
                            max_table_rows_per_page=max_table_rows_per_page,
                            block_handling=block_handling,
                            log_callback=log_callback,
                            stop_event=stop_event,
                            pause_event=pause_event,
                        )
                    except Exception as exc:
                        checkpoint.release_item(item.checkpoint_key)
                        log_error(log_callback, f"SteamDB 动态采集失败 AppID={item.appid}: {exc}")
                        continue
                    _write_bundle(writer, bundle)
                    if should_stop(stop_event):
                        checkpoint.release_item(item.checkpoint_key)
                        continue
                    checkpoint.mark_completed(item.checkpoint_key, bundle.meta)
                    completed_this_run += 1
                    log_line(
                        log_callback,
                        f"[{index}/{len(items)}] SteamDB 完成 AppID={item.appid}，"
                        f"状态 {len(bundle.status_rows)}，RawTables {len(bundle.raw_rows)} 行。",
                    )
                try:
                    page.close()
                except Exception:
                    pass
            finally:
                pass
    finally:
        try:
            writer.save()
        except Exception:
            pass
        checkpoint.close_run()

    if should_stop(stop_event):
        log_warn(log_callback, f"SteamDB 动态窗口任务已停止，本轮完成 {completed_this_run} 个 App。")
    else:
        log_line(log_callback, f"SteamDB 动态窗口任务完成，本轮完成 {completed_this_run} 个 App。")
    finish_callback(output_path)
    return output_path


__all__ = [
    "STEAMDB_STATUS_FIELDS",
    "STEAMDB_OVERVIEW_FIELDS",
    "STEAMDB_CHARTS_FIELDS",
    "STEAMDB_PACKAGE_FIELDS",
    "STEAMDB_DLC_FIELDS",
    "STEAMDB_DEPOT_FIELDS",
    "STEAMDB_HISTORY_FIELDS",
    "STEAMDB_RAW_TABLE_FIELDS",
    "STEAMDB_SHEETS_FIELDS",
    "SteamDbWorkItem",
    "SteamDbPageSnapshot",
    "SteamDbBundle",
    "parse_steamdb_app_ids",
    "build_steamdb_app_url",
    "detect_steamdb_block",
    "read_app_ids_from_xlsx",
    "build_overview_row",
    "build_package_rows",
    "build_dlc_rows",
    "build_depot_rows",
    "build_history_rows",
    "prepare_steamdb_work_items",
    "run_steamdb_dynamic_window_spider",
]
