# -*- coding: utf-8 -*-
"""Steam API based game research tool."""

from __future__ import annotations

import hashlib
import html
import json
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import requests

from src.core import (
    build_output_path,
    interruptible_sleep,
    log_error,
    log_line,
    log_warn,
    should_stop,
    wait_if_paused,
)
from src.core.app_state import get_app_state_root
from src.core.task_checkpoint import open_checkpointed_multi_sheet_writer, open_task_checkpoint


STORE_SEARCH_URL = "https://store.steampowered.com/api/storesearch/"
STORE_APPDETAILS_URL = "https://store.steampowered.com/api/appdetails"
STORE_REVIEWS_URL = "https://store.steampowered.com/appreviews/{appid}"
STEAM_NEWS_URL = "https://api.steampowered.com/ISteamNews/GetNewsForApp/v2/"
CURRENT_PLAYERS_URL = "https://api.steampowered.com/ISteamUserStats/GetNumberOfCurrentPlayers/v1/"
GLOBAL_ACHIEVEMENTS_URL = "https://api.steampowered.com/ISteamUserStats/GetGlobalAchievementPercentagesForApp/v0002/"
PUBLIC_APP_LIST_URL = "https://api.steampowered.com/ISteamApps/GetAppList/v2/"
ISTORE_APP_LIST_URL = "https://api.steampowered.com/IStoreService/GetAppList/v1/"
PLAYER_SUMMARIES_URL = "https://api.steampowered.com/ISteamUser/GetPlayerSummaries/v0002/"
PLAYER_BANS_URL = "https://api.steampowered.com/ISteamUser/GetPlayerBans/v1/"
FRIEND_LIST_URL = "https://api.steampowered.com/ISteamUser/GetFriendList/v1/"
OWNED_GAMES_URL = "https://api.steampowered.com/IPlayerService/GetOwnedGames/v1/"
RECENT_GAMES_URL = "https://api.steampowered.com/IPlayerService/GetRecentlyPlayedGames/v1/"
STEAM_LEVEL_URL = "https://api.steampowered.com/IPlayerService/GetSteamLevel/v1/"
BADGES_URL = "https://api.steampowered.com/IPlayerService/GetBadges/v1/"
PLAYER_ACHIEVEMENTS_URL = "https://api.steampowered.com/ISteamUserStats/GetPlayerAchievements/v0001/"

DEFAULT_LANGUAGE = "english"
DEFAULT_COUNTRY = "US"

APP_FIELDS = [
    "来源类型",
    "搜索词",
    "AppID",
    "游戏名",
    "状态",
    "类型",
    "是否免费",
    "价格",
    "原价",
    "折扣",
    "发行日期",
    "是否即将发行",
    "开发商",
    "发行商",
    "类型标签",
    "分类标签",
    "支持平台",
    "支持语言",
    "年龄限制",
    "控制器支持",
    "PC最低配置",
    "PC推荐配置",
    "Mac最低配置",
    "Mac推荐配置",
    "Linux最低配置",
    "Linux推荐配置",
    "DLC AppID列表",
    "Demo列表",
    "所属完整游戏",
    "安装包ID列表",
    "套餐/版本列表",
    "内容描述",
    "分级信息",
    "客服网站",
    "客服邮箱",
    "客服电话",
    "第三方账号提示",
    "DRM提示",
    "法律声明",
    "短简介",
    "详细简介",
    "总评描述",
    "总评分数",
    "总评数",
    "好评数",
    "差评数",
    "好评率",
    "推荐数",
    "Metacritic",
    "当前在线",
    "成就数量",
    "成就样本列表",
    "新闻样本数",
    "玩家评论样本数",
    "商店链接",
    "官网",
    "头图",
    "胶囊图",
    "胶囊图V5",
    "背景图",
    "原始背景图",
    "截图列表",
    "视频列表",
    "查询时间",
]

REVIEW_FIELDS = [
    "来源类型",
    "搜索词",
    "AppID",
    "游戏名",
    "评论ID",
    "SteamID",
    "是否推荐",
    "语言",
    "Steam购买",
    "免费获得",
    "抢先体验期间",
    "主要在SteamDeck",
    "拥有游戏数",
    "作者评论数",
    "总游玩小时",
    "近两周游玩小时",
    "评价时游玩小时",
    "最后游玩时间",
    "有价值票数",
    "欢乐票数",
    "加权评分",
    "开发者回复",
    "开发者回复时间",
    "发布时间",
    "最后更新时间",
    "评论链接",
    "评论内容",
    "查询时间",
]

NEWS_FIELDS = [
    "来源类型",
    "搜索词",
    "AppID",
    "游戏名",
    "新闻ID",
    "标题",
    "作者",
    "发布时间",
    "来源名称",
    "来源标签",
    "Feed名称",
    "Feed类型",
    "外部链接",
    "链接",
    "内容摘要",
    "查询时间",
]

PLAYER_PROFILE_FIELDS = [
    "来源类型",
    "搜索词",
    "目标AppID",
    "目标游戏名",
    "SteamID",
    "昵称",
    "主页链接",
    "头像",
    "国家/地区",
    "省州",
    "城市ID",
    "真实名",
    "资料可见性",
    "资料状态",
    "在线状态",
    "最后在线时间",
    "账号创建时间",
    "目标游戏是否推荐",
    "目标游戏总游玩小时",
    "目标游戏近两周小时",
    "目标游戏评价时小时",
    "目标游戏最后游玩时间",
    "Steam等级",
    "徽章数",
    "公开好友数",
    "公开游戏数",
    "公开游戏总小时",
    "近两周游戏数",
    "社区封禁",
    "VAC封禁",
    "VAC封禁数",
    "游戏封禁数",
    "距上次封禁天数",
    "经济封禁",
    "隐私/错误",
    "查询时间",
]

PLAYER_LIBRARY_FIELDS = [
    "来源类型",
    "搜索词",
    "目标AppID",
    "目标游戏名",
    "SteamID",
    "库游戏AppID",
    "库游戏名",
    "总游玩小时",
    "近两周游玩小时",
    "最后游玩时间",
    "有公开社区统计",
    "图标",
    "查询时间",
]

PLAYER_RECENT_FIELDS = [
    "来源类型",
    "搜索词",
    "目标AppID",
    "目标游戏名",
    "SteamID",
    "最近游戏AppID",
    "最近游戏名",
    "近两周游玩小时",
    "总游玩小时",
    "查询时间",
]

PLAYER_ACHIEVEMENT_FIELDS = [
    "来源类型",
    "搜索词",
    "目标AppID",
    "目标游戏名",
    "SteamID",
    "成就API名",
    "成就名称",
    "成就描述",
    "是否解锁",
    "解锁时间",
    "查询时间",
]

PLAYER_BADGE_FIELDS = [
    "来源类型",
    "搜索词",
    "目标AppID",
    "目标游戏名",
    "SteamID",
    "徽章ID",
    "等级",
    "XP",
    "稀有度",
    "完成时间",
    "关联AppID",
    "社区物品ID",
    "边框颜色",
    "查询时间",
]

PLAYER_FRIEND_FIELDS = [
    "来源类型",
    "搜索词",
    "目标AppID",
    "目标游戏名",
    "SteamID",
    "好友SteamID",
    "关系",
    "加好友时间",
    "查询时间",
]

PLAYER_SHEETS_FIELDS = {
    "玩家画像": PLAYER_PROFILE_FIELDS,
    "玩家游戏库": PLAYER_LIBRARY_FIELDS,
    "最近游玩": PLAYER_RECENT_FIELDS,
    "目标游戏成就": PLAYER_ACHIEVEMENT_FIELDS,
    "玩家徽章": PLAYER_BADGE_FIELDS,
    "玩家好友": PLAYER_FRIEND_FIELDS,
}


@dataclass(frozen=True)
class SteamWorkItem:
    appid: int
    source_type: str
    source: str = ""
    seed_name: str = ""

    @property
    def checkpoint_key(self) -> str:
        label = self.source if self.source_type == "关键词" else "direct"
        return f"{label}|{self.appid}".lower()


@dataclass
class SteamAppBundle:
    app_row: dict[str, Any]
    review_rows: list[dict[str, Any]]
    news_rows: list[dict[str, Any]]
    meta: dict[str, Any]


@dataclass(frozen=True)
class SteamPlayerContext:
    steamid: str
    target_appid: int | None = None
    target_game_name: str = ""
    source_type: str = "玩家评论"
    source: str = ""
    voted_up: str = ""
    target_play_hours: str = ""
    target_recent_hours: str = ""
    target_review_hours: str = ""
    target_last_played: str = ""

    @property
    def checkpoint_key(self) -> str:
        target = str(self.target_appid or "none")
        return f"{target}|{self.steamid}".lower()


@dataclass
class SteamPlayerBundle:
    profile_row: dict[str, Any]
    library_rows: list[dict[str, Any]]
    recent_rows: list[dict[str, Any]]
    achievement_rows: list[dict[str, Any]]
    badge_rows: list[dict[str, Any]]
    friend_rows: list[dict[str, Any]]
    meta: dict[str, Any]


def parse_app_ids(value: str | list[str] | tuple[str, ...]) -> list[int]:
    """Extract Steam appids from bare numbers or Steam store URLs."""
    if isinstance(value, str):
        lines = value.splitlines()
    else:
        lines = list(value or [])
    appids: list[int] = []
    seen: set[int] = set()
    for raw in lines:
        text = str(raw or "").strip()
        if not text:
            continue
        match = re.search(r"(?:/app/|^)(\d{2,12})(?:\D|$)", text)
        if not match:
            continue
        appid = int(match.group(1))
        if appid not in seen:
            seen.add(appid)
            appids.append(appid)
    return appids


def parse_steam_ids(value: str | list[str] | tuple[str, ...]) -> list[str]:
    """Extract 64-bit SteamIDs from plain values or Steam profile URLs."""
    if isinstance(value, str):
        lines = value.splitlines()
    else:
        lines = list(value or [])
    steamids: list[str] = []
    seen: set[str] = set()
    for raw in lines:
        text = str(raw or "").strip()
        if not text:
            continue
        matches = re.findall(r"(?<!\d)(7656119\d{10})(?!\d)", text)
        if not matches and text.isdigit() and len(text) >= 16:
            matches = [text]
        for steamid in matches:
            if steamid not in seen:
                seen.add(steamid)
                steamids.append(steamid)
    return steamids


def normalize_keywords(value: str | list[str] | tuple[str, ...]) -> list[str]:
    if isinstance(value, str):
        lines = value.splitlines()
    else:
        lines = list(value or [])
    keywords: list[str] = []
    seen: set[str] = set()
    for raw in lines:
        text = str(raw or "").strip()
        if not text:
            continue
        key = re.sub(r"\s+", " ", text).lower()
        if key not in seen:
            seen.add(key)
            keywords.append(text)
    return keywords


def read_player_contexts_from_xlsx(path: str, limit: int = 0) -> list[SteamPlayerContext]:
    """Read Steam player contexts from an existing workbook, preferably the 玩家评论 sheet."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_name = "玩家评论" if "玩家评论" in workbook.sheetnames else workbook.sheetnames[0]
    worksheet = workbook[sheet_name]
    rows = worksheet.iter_rows(values_only=True)
    try:
        header = [str(value or "").strip() for value in next(rows)]
    except StopIteration:
        return []
    index = {name: pos for pos, name in enumerate(header) if name}
    contexts: list[SteamPlayerContext] = []
    seen: set[str] = set()
    for row in rows:
        steamid = _row_value(row, index, "SteamID")
        if not steamid:
            continue
        appid_text = _row_value(row, index, "AppID") or _row_value(row, index, "目标AppID")
        target_appid = _safe_int(appid_text)
        context = SteamPlayerContext(
            steamid=str(steamid).strip(),
            target_appid=target_appid,
            target_game_name=_row_value(row, index, "游戏名") or _row_value(row, index, "目标游戏名"),
            source_type=_row_value(row, index, "来源类型") or "玩家评论表",
            source=_row_value(row, index, "搜索词"),
            voted_up=_row_value(row, index, "是否推荐"),
            target_play_hours=_row_value(row, index, "总游玩小时") or _row_value(row, index, "目标游戏总游玩小时"),
            target_recent_hours=_row_value(row, index, "近两周游玩小时") or _row_value(row, index, "目标游戏近两周小时"),
            target_review_hours=_row_value(row, index, "评价时游玩小时") or _row_value(row, index, "目标游戏评价时小时"),
            target_last_played=_row_value(row, index, "最后游玩时间") or _row_value(row, index, "目标游戏最后游玩时间"),
        )
        key = context.checkpoint_key
        if key in seen:
            continue
        seen.add(key)
        contexts.append(context)
        if limit and len(contexts) >= limit:
            break
    return contexts


def player_contexts_from_steam_ids(
    steam_ids: str | list[str],
    *,
    target_appid: int | None = None,
    target_game_name: str = "",
) -> list[SteamPlayerContext]:
    return [
        SteamPlayerContext(
            steamid=steamid,
            target_appid=target_appid,
            target_game_name=target_game_name,
            source_type="直接输入",
        )
        for steamid in parse_steam_ids(steam_ids)
    ]


def player_contexts_from_review_rows(rows: list[dict[str, Any]], limit: int = 0) -> list[SteamPlayerContext]:
    contexts: list[SteamPlayerContext] = []
    seen: set[str] = set()
    for row in rows:
        steamid = str(row.get("SteamID") or "").strip()
        if not steamid:
            continue
        target_appid = _safe_int(row.get("AppID"))
        context = SteamPlayerContext(
            steamid=steamid,
            target_appid=target_appid,
            target_game_name=str(row.get("游戏名") or ""),
            source_type=str(row.get("来源类型") or "玩家评论"),
            source=str(row.get("搜索词") or ""),
            voted_up=str(row.get("是否推荐") or ""),
            target_play_hours=str(row.get("总游玩小时") or ""),
            target_recent_hours=str(row.get("近两周游玩小时") or ""),
            target_review_hours=str(row.get("评价时游玩小时") or ""),
            target_last_played=str(row.get("最后游玩时间") or ""),
        )
        key = context.checkpoint_key
        if key in seen:
            continue
        seen.add(key)
        contexts.append(context)
        if limit and len(contexts) >= limit:
            break
    return contexts


def parse_date_range(start_date: str, end_date: str) -> tuple[datetime, datetime]:
    start_dt = datetime.strptime(start_date.strip(), "%Y-%m-%d")
    end_dt = datetime.strptime(end_date.strip(), "%Y-%m-%d")
    if start_dt > end_dt:
        raise ValueError("开始日期不能晚于结束日期")
    return start_dt, end_dt


def _now_text() -> str:
    return time.strftime("%Y-%m-%d %H:%M:%S")


def _cache_dir() -> Path:
    path = get_app_state_root() / "steam_cache"
    path.mkdir(parents=True, exist_ok=True)
    return path


def _cache_key(payload: dict[str, Any]) -> str:
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:20]


def _read_cache(path: Path, max_age_hours: float) -> Any | None:
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        updated_ts = float(data.get("updated_ts", 0) or 0)
        if updated_ts and time.time() - updated_ts <= max(0, max_age_hours) * 3600:
            return data.get("payload")
    except Exception:
        return None
    return None


def _write_cache(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(
        json.dumps({"updated_ts": time.time(), "payload": payload}, ensure_ascii=False),
        encoding="utf-8",
    )
    try:
        tmp.replace(path)
    except OSError:
        path.write_text(tmp.read_text(encoding="utf-8"), encoding="utf-8")
        try:
            tmp.unlink()
        except OSError:
            pass


def _clean_text(value: Any, *, max_length: int | None = None) -> str:
    text = str(value or "")
    if not text:
        return ""
    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(text)
    text = text.replace("\r", "\n")
    text = re.sub(r"[ \t\f\v]+", " ", text)
    text = re.sub(r"\n\s*\n+", "\n", text)
    text = text.strip()
    if max_length and len(text) > max_length:
        return text[:max_length].rstrip()
    return text


def _limit_text(value: Any, max_length: int = 12000) -> str:
    text = str(value or "").strip()
    if len(text) > max_length:
        return text[:max_length].rstrip()
    return text


def _join(values: list[Any]) -> str:
    cleaned = [str(value).strip() for value in values if str(value or "").strip()]
    return _limit_text("\n".join(cleaned))


def _list_desc(values: Any) -> str:
    if not isinstance(values, list):
        return ""
    return " / ".join(str(item.get("description") or item.get("name") or "").strip() for item in values if isinstance(item, dict) and (item.get("description") or item.get("name")))


def _format_timestamp(value: Any) -> str:
    try:
        ts = int(value or 0)
    except (TypeError, ValueError):
        return ""
    if ts <= 0:
        return ""
    return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S")


def _minutes_to_hours(value: Any) -> str:
    try:
        return str(round(float(value or 0) / 60, 1))
    except (TypeError, ValueError):
        return ""


def _yes_no(value: Any) -> str:
    return "是" if bool(value) else "否"


def _requirements_text(value: Any, key: str) -> str:
    if isinstance(value, dict):
        return _clean_text(value.get(key), max_length=3000)
    if key == "minimum":
        return _clean_text(value, max_length=3000)
    return ""


def _appid_name(value: Any) -> str:
    if not isinstance(value, dict):
        return ""
    appid = str(value.get("appid") or "").strip()
    name = str(value.get("name") or "").strip()
    if appid and name:
        return f"{appid} {name}"
    return appid or name


def _demos_text(value: Any) -> str:
    if not isinstance(value, list):
        return ""
    rows = []
    for item in value:
        if isinstance(item, dict):
            appid = str(item.get("appid") or "").strip()
            description = _clean_text(item.get("description"), max_length=120)
            rows.append(f"{appid} {description}".strip())
    return _join(rows)


def _package_groups_text(value: Any) -> str:
    if not isinstance(value, list):
        return ""
    rows: list[str] = []
    for group in value:
        if not isinstance(group, dict):
            continue
        title = _clean_text(group.get("title") or group.get("name"), max_length=120)
        description = _clean_text(group.get("description"), max_length=240)
        subs = []
        for sub in group.get("subs", []) or []:
            if not isinstance(sub, dict):
                continue
            packageid = str(sub.get("packageid") or "").strip()
            option_text = _clean_text(sub.get("option_text"), max_length=160)
            subs.append(f"{packageid} {option_text}".strip())
        rows.append(" | ".join(part for part in [title, description, "; ".join(subs)] if part))
    return _join(rows)


def _content_descriptors_text(value: Any) -> str:
    if not isinstance(value, dict):
        return ""
    ids = value.get("ids", [])
    ids_text = ", ".join(str(item) for item in ids) if isinstance(ids, list) else str(ids or "")
    notes = _clean_text(value.get("notes"), max_length=1000)
    return " | ".join(part for part in [f"ids: {ids_text}" if ids_text else "", notes] if part)


def _ratings_text(value: Any) -> str:
    if not isinstance(value, dict):
        return ""
    rows: list[str] = []
    for rating_name, rating_data in value.items():
        if isinstance(rating_data, dict):
            summary = ", ".join(f"{key}={_clean_text(val, max_length=100)}" for key, val in rating_data.items())
            rows.append(f"{rating_name}: {summary}")
        else:
            rows.append(f"{rating_name}: {_clean_text(rating_data, max_length=200)}")
    return _join(rows)


def _achievement_highlights_text(value: Any) -> str:
    if not isinstance(value, dict):
        return ""
    highlighted = value.get("highlighted", [])
    if not isinstance(highlighted, list):
        return ""
    rows = []
    for item in highlighted:
        if isinstance(item, dict):
            name = _clean_text(item.get("name"), max_length=160)
            path = str(item.get("path") or "").strip()
            rows.append(" | ".join(part for part in [name, path] if part))
    return _join(rows)


def _redact_sensitive_text(value: Any) -> str:
    text = str(value or "")
    if not text:
        return ""
    text = re.sub(r"(?i)([?&](?:key|steam_web_api_key|access_token)=)[^&\s]+", r"\1<redacted>", text)
    text = re.sub(r"(?i)\b(api[_-]?key=)[^&\s]+", r"\1<redacted>", text)
    return text


def _request_json(
    session: requests.Session,
    url: str,
    *,
    params: dict[str, Any] | None = None,
    timeout: float = 30.0,
    max_retries: int = 3,
    request_delay: float = 0.0,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> Any:
    last_exc: Exception | None = None
    headers = {
        "User-Agent": "Mozilla/5.0 social-platform-scraper Steam API collector",
        "Accept": "application/json,text/plain,*/*",
    }
    for attempt in range(max(1, int(max_retries or 1))):
        if wait_if_paused(pause_event, stop_event) or should_stop(stop_event):
            raise InterruptedError("任务已停止")
        try:
            response = session.get(url, params=params, headers=headers, timeout=timeout)
            if response.status_code in {429, 500, 502, 503, 504} and attempt < max_retries - 1:
                wait_seconds = min(30.0, 2.0 ** attempt + 1.0)
                log_warn(log_callback, f"Steam API HTTP {response.status_code}，{wait_seconds:.1f} 秒后重试。")
                interruptible_sleep(wait_seconds, stop_event)
                continue
            response.raise_for_status()
            payload = response.json()
            if request_delay > 0:
                interruptible_sleep(request_delay, stop_event)
            return payload
        except Exception as exc:
            last_exc = exc
            if attempt < max_retries - 1:
                wait_seconds = min(30.0, 2.0 ** attempt + 1.0)
                log_warn(log_callback, f"Steam API 请求失败，{wait_seconds:.1f} 秒后重试：{_redact_sensitive_text(exc)}")
                interruptible_sleep(wait_seconds, stop_event)
                continue
            break
    if last_exc:
        raise RuntimeError(_redact_sensitive_text(last_exc)) from last_exc
    raise RuntimeError("Steam API 请求失败")


def _config_bool(config: dict[str, Any], key: str, default: str = "否") -> bool:
    return str(config.get(key, default)).strip() == "是"


def discover_apps_for_keyword(
    keyword: str,
    *,
    api_key: str = "",
    language: str = DEFAULT_LANGUAGE,
    country: str = DEFAULT_COUNTRY,
    max_apps: int = 100,
    config: dict[str, Any] | None = None,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> list[SteamWorkItem]:
    config = config or {}
    max_apps = max(1, int(max_apps or 1))
    timeout = float(config.get("request_timeout", 30) or 30)
    request_delay = float(config.get("request_delay", 0.2) or 0.0)
    cache_hours = float(config.get("cache_ttl_hours", 168) or 0)
    search_mode = str(config.get("keyword_search_mode", "商店搜索接口（推荐）") or "商店搜索接口（推荐）")

    cache_payload = {
        "keyword": keyword,
        "language": language,
        "country": country,
        "max_apps": max_apps,
        "search_mode": search_mode,
        "api_key_present": bool(api_key),
    }
    cache_path = _cache_dir() / f"keyword_{_cache_key(cache_payload)}.json"
    cached = _read_cache(cache_path, cache_hours)
    if isinstance(cached, list):
        log_line(log_callback, f"关键词缓存命中：{keyword}，{len(cached)} 个 AppID。")
        return [
            SteamWorkItem(appid=int(item["appid"]), source_type="关键词", source=keyword, seed_name=str(item.get("name") or ""))
            for item in cached
            if str(item.get("appid") or "").isdigit()
        ]

    session = requests.Session()
    discovered: list[dict[str, Any]] = []
    if search_mode != "AppList 本地匹配":
        discovered = _search_store_keyword(
            session,
            keyword,
            max_apps=max_apps,
            language=language,
            country=country,
            timeout=timeout,
            request_delay=request_delay,
            log_callback=log_callback,
            stop_event=stop_event,
            pause_event=pause_event,
        )
    if len(discovered) < max_apps and search_mode in {"AppList 本地匹配", "商店搜索后补 AppList"}:
        app_list_matches = _match_app_list_keyword(
            keyword,
            api_key=api_key,
            max_apps=max_apps,
            timeout=timeout,
            cache_hours=cache_hours,
            request_delay=request_delay,
            log_callback=log_callback,
            stop_event=stop_event,
            pause_event=pause_event,
        )
        seen = {int(item["appid"]) for item in discovered if str(item.get("appid") or "").isdigit()}
        for item in app_list_matches:
            appid = int(item["appid"])
            if appid not in seen:
                seen.add(appid)
                discovered.append(item)
            if len(discovered) >= max_apps:
                break

    payload = [{"appid": int(item["appid"]), "name": str(item.get("name") or "")} for item in discovered[:max_apps]]
    _write_cache(cache_path, payload)
    log_line(log_callback, f"关键词发现：{keyword}，找到 {len(payload)} 个候选 App。")
    return [SteamWorkItem(appid=item["appid"], source_type="关键词", source=keyword, seed_name=item.get("name", "")) for item in payload]


def _search_store_keyword(
    session: requests.Session,
    keyword: str,
    *,
    max_apps: int,
    language: str,
    country: str,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    seen: set[int] = set()
    start = 0
    while len(results) < max_apps:
        if wait_if_paused(pause_event, stop_event) or should_stop(stop_event):
            break
        count = min(100, max_apps - len(results))
        payload = _request_json(
            session,
            STORE_SEARCH_URL,
            params={"term": keyword, "l": language, "cc": country, "start": start, "count": count},
            timeout=timeout,
            request_delay=request_delay,
            log_callback=log_callback,
            stop_event=stop_event,
            pause_event=pause_event,
        )
        items = payload.get("items", []) if isinstance(payload, dict) else []
        if not items:
            break
        for item in items:
            if not isinstance(item, dict):
                continue
            appid = item.get("id") or item.get("appid")
            try:
                appid = int(appid)
            except (TypeError, ValueError):
                continue
            if appid in seen:
                continue
            seen.add(appid)
            results.append({"appid": appid, "name": item.get("name", "")})
            if len(results) >= max_apps:
                break
        total = int(payload.get("total", 0) or 0) if isinstance(payload, dict) else 0
        start += len(items)
        if total and start >= total:
            break
        if len(items) < count:
            break
    return results


def _load_app_list(
    *,
    api_key: str,
    timeout: float,
    cache_hours: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> list[dict[str, Any]]:
    cache_path = _cache_dir() / ("app_list_istore.json" if api_key else "app_list_public.json")
    cached = _read_cache(cache_path, cache_hours)
    if isinstance(cached, list):
        return cached

    session = requests.Session()
    apps: list[dict[str, Any]] = []
    if api_key:
        try:
            last_appid = 0
            while True:
                params: dict[str, Any] = {
                    "key": api_key,
                    "include_games": "true",
                    "include_dlc": "false",
                    "include_software": "false",
                    "max_results": 50000,
                }
                if last_appid:
                    params["last_appid"] = last_appid
                payload = _request_json(
                    session,
                    ISTORE_APP_LIST_URL,
                    params=params,
                    timeout=timeout,
                    request_delay=request_delay,
                    log_callback=log_callback,
                    stop_event=stop_event,
                    pause_event=pause_event,
                )
                page_apps = ((payload or {}).get("response") or {}).get("apps", []) if isinstance(payload, dict) else []
                if not page_apps:
                    break
                apps.extend({"appid": int(item.get("appid")), "name": item.get("name", "")} for item in page_apps if str(item.get("appid") or "").isdigit())
                last_appid = int(page_apps[-1].get("appid") or 0)
                if len(page_apps) < 50000:
                    break
        except Exception as exc:
            log_warn(log_callback, f"IStoreService AppList 失败，将使用公开 AppList 兜底：{exc}")
            apps = []

    if not apps:
        payload = _request_json(
            session,
            PUBLIC_APP_LIST_URL,
            timeout=timeout,
            request_delay=request_delay,
            log_callback=log_callback,
            stop_event=stop_event,
            pause_event=pause_event,
        )
        raw_apps = ((payload or {}).get("applist") or {}).get("apps", []) if isinstance(payload, dict) else []
        apps = [{"appid": int(item.get("appid")), "name": item.get("name", "")} for item in raw_apps if str(item.get("appid") or "").isdigit()]

    _write_cache(cache_path, apps)
    return apps


def _match_app_list_keyword(
    keyword: str,
    *,
    api_key: str,
    max_apps: int,
    timeout: float,
    cache_hours: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> list[dict[str, Any]]:
    apps = _load_app_list(
        api_key=api_key,
        timeout=timeout,
        cache_hours=cache_hours,
        request_delay=request_delay,
        log_callback=log_callback,
        stop_event=stop_event,
        pause_event=pause_event,
    )
    keyword_norm = re.sub(r"\s+", " ", keyword.strip().lower())
    matches: list[dict[str, Any]] = []
    for item in apps:
        name = str(item.get("name") or "")
        if keyword_norm and keyword_norm in name.lower():
            matches.append(item)
            if len(matches) >= max_apps:
                break
    return matches


def collect_steam_app_bundle(
    item: SteamWorkItem,
    *,
    api_key: str = "",
    language: str = DEFAULT_LANGUAGE,
    country: str = DEFAULT_COUNTRY,
    collect_reviews: bool = False,
    max_reviews: int = 0,
    collect_news: bool = False,
    max_news: int = 0,
    config: dict[str, Any] | None = None,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> SteamAppBundle:
    config = config or {}
    timeout = float(config.get("request_timeout", 30) or 30)
    request_delay = float(config.get("request_delay", 0.2) or 0.0)
    include_non_games = _config_bool(config, "include_non_games", "否")
    collect_current_players = _config_bool(config, "collect_current_players", "是")
    collect_achievements = _config_bool(config, "collect_achievements", "否")
    review_language = str(config.get("review_language", "all") or "all").strip() or "all"
    reviews_filter = str(config.get("reviews_filter", "all") or "all").strip() or "all"

    session = requests.Session()
    query_time = _now_text()
    details = fetch_app_details(
        session,
        item.appid,
        language=language,
        country=country,
        timeout=timeout,
        request_delay=request_delay,
        log_callback=log_callback,
        stop_event=stop_event,
        pause_event=pause_event,
    )
    data = details.get("data") if details.get("success") else None
    if not isinstance(data, dict):
        row = _base_app_row(item, query_time)
        row.update({"游戏名": item.seed_name, "状态": "not_found", "商店链接": _store_url(item.appid)})
        return SteamAppBundle(row, [], [], {"app_count": 1, "review_count": 0, "news_count": 0})

    app_type = str(data.get("type") or "")
    if app_type and app_type != "game" and not include_non_games:
        row = _build_app_row(item, data, query_time, status="skipped_non_game")
        return SteamAppBundle(row, [], [], {"app_count": 1, "review_count": 0, "news_count": 0, "skipped_non_game": 1})

    review_summary: dict[str, Any] = {}
    review_rows: list[dict[str, Any]] = []
    if collect_reviews:
        review_summary, reviews = fetch_reviews(
            session,
            item.appid,
            max_reviews=max(0, int(max_reviews or 0)),
            language=review_language,
            filter_mode=reviews_filter,
            timeout=timeout,
            request_delay=request_delay,
            log_callback=log_callback,
            stop_event=stop_event,
            pause_event=pause_event,
        )
        review_rows = [_build_review_row(item, data, review, query_time) for review in reviews]

    news_rows: list[dict[str, Any]] = []
    if collect_news and max_news > 0:
        news_items = fetch_news(
            session,
            item.appid,
            max_news=max(0, int(max_news or 0)),
            timeout=timeout,
            request_delay=request_delay,
            log_callback=log_callback,
            stop_event=stop_event,
            pause_event=pause_event,
        )
        news_rows = [_build_news_row(item, data, news, query_time) for news in news_items]

    current_players = ""
    if collect_current_players:
        try:
            current_players = fetch_current_players(
                session,
                item.appid,
                timeout=timeout,
                request_delay=request_delay,
                log_callback=log_callback,
                stop_event=stop_event,
                pause_event=pause_event,
            )
        except Exception as exc:
            log_warn(log_callback, f"当前在线人数获取失败 AppID={item.appid}: {exc}")

    achievement_count = ""
    achievement_sample = ""
    if collect_achievements:
        try:
            achievement_count, achievement_sample = fetch_achievement_summary(
                session,
                item.appid,
                timeout=timeout,
                request_delay=request_delay,
                log_callback=log_callback,
                stop_event=stop_event,
                pause_event=pause_event,
            )
        except Exception as exc:
            log_warn(log_callback, f"成就摘要获取失败 AppID={item.appid}: {exc}")

    app_row = _build_app_row(
        item,
        data,
        query_time,
        status="ok",
        review_summary=review_summary,
        current_players=current_players,
        achievement_count=achievement_count,
        achievement_sample=achievement_sample,
        news_count=len(news_rows),
        review_count=len(review_rows),
    )
    return SteamAppBundle(
        app_row,
        review_rows,
        news_rows,
        {"app_count": 1, "review_count": len(review_rows), "news_count": len(news_rows)},
    )


def fetch_app_details(
    session: requests.Session,
    appid: int,
    *,
    language: str,
    country: str,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> dict[str, Any]:
    payload = _request_json(
        session,
        STORE_APPDETAILS_URL,
        params={"appids": appid, "l": language, "cc": country},
        timeout=timeout,
        request_delay=request_delay,
        log_callback=log_callback,
        stop_event=stop_event,
        pause_event=pause_event,
    )
    item = payload.get(str(appid), {}) if isinstance(payload, dict) else {}
    return item if isinstance(item, dict) else {}


def fetch_reviews(
    session: requests.Session,
    appid: int,
    *,
    max_reviews: int,
    language: str,
    filter_mode: str,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    reviews: list[dict[str, Any]] = []
    summary: dict[str, Any] = {}
    cursor = "*"
    first = True
    while first or len(reviews) < max_reviews:
        first = False
        if wait_if_paused(pause_event, stop_event) or should_stop(stop_event):
            break
        remaining = max(0, max_reviews - len(reviews))
        page_size = min(100, remaining if remaining > 0 else 1)
        payload = _request_json(
            session,
            STORE_REVIEWS_URL.format(appid=appid),
            params={
                "json": 1,
                "filter": filter_mode,
                "language": language,
                "review_type": "all",
                "purchase_type": "all",
                "num_per_page": page_size,
                "cursor": cursor,
            },
            timeout=timeout,
            request_delay=request_delay,
            log_callback=log_callback,
            stop_event=stop_event,
            pause_event=pause_event,
        )
        if isinstance(payload, dict) and isinstance(payload.get("query_summary"), dict):
            summary = payload.get("query_summary", {})
        page_reviews = payload.get("reviews", []) if isinstance(payload, dict) else []
        if max_reviews <= 0:
            break
        if not page_reviews:
            break
        for review in page_reviews:
            if isinstance(review, dict):
                reviews.append(review)
                if len(reviews) >= max_reviews:
                    break
        next_cursor = str(payload.get("cursor") or "") if isinstance(payload, dict) else ""
        if not next_cursor or next_cursor == cursor:
            break
        cursor = next_cursor
    return summary, reviews


def fetch_news(
    session: requests.Session,
    appid: int,
    *,
    max_news: int,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> list[dict[str, Any]]:
    payload = _request_json(
        session,
        STEAM_NEWS_URL,
        params={"appid": appid, "count": max_news, "maxlength": 1200, "format": "json"},
        timeout=timeout,
        request_delay=request_delay,
        log_callback=log_callback,
        stop_event=stop_event,
        pause_event=pause_event,
    )
    items = ((payload or {}).get("appnews") or {}).get("newsitems", []) if isinstance(payload, dict) else []
    return [item for item in items if isinstance(item, dict)][:max_news]


def fetch_current_players(
    session: requests.Session,
    appid: int,
    *,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> str:
    payload = _request_json(
        session,
        CURRENT_PLAYERS_URL,
        params={"appid": appid},
        timeout=timeout,
        request_delay=request_delay,
        log_callback=log_callback,
        stop_event=stop_event,
        pause_event=pause_event,
    )
    response = payload.get("response", {}) if isinstance(payload, dict) else {}
    if int(response.get("result", 0) or 0) != 1:
        return ""
    return str(response.get("player_count", ""))


def fetch_achievement_summary(
    session: requests.Session,
    appid: int,
    *,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> tuple[str, str]:
    payload = _request_json(
        session,
        GLOBAL_ACHIEVEMENTS_URL,
        params={"gameid": appid, "format": "json"},
        timeout=timeout,
        request_delay=request_delay,
        log_callback=log_callback,
        stop_event=stop_event,
        pause_event=pause_event,
    )
    achievements = ((payload or {}).get("achievementpercentages") or {}).get("achievements", []) if isinstance(payload, dict) else []
    if not isinstance(achievements, list):
        return "", ""
    rows = []
    for item in achievements[:50]:
        if not isinstance(item, dict):
            continue
        name = _clean_text(item.get("name"), max_length=160)
        percent = item.get("percent", "")
        rows.append(f"{name}: {percent}%".strip())
    return str(len(achievements)), _join(rows)


def fetch_achievement_count(
    session: requests.Session,
    appid: int,
    *,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> str:
    count, _ = fetch_achievement_summary(
        session,
        appid,
        timeout=timeout,
        request_delay=request_delay,
        log_callback=log_callback,
        stop_event=stop_event,
        pause_event=pause_event,
    )
    return count


def fetch_player_summaries(
    session: requests.Session,
    api_key: str,
    steamids: list[str],
    *,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for chunk in _chunks(steamids, 100):
        payload = _request_json(
            session,
            PLAYER_SUMMARIES_URL,
            params={"key": api_key, "steamids": ",".join(chunk)},
            timeout=timeout,
            request_delay=request_delay,
            log_callback=log_callback,
            stop_event=stop_event,
            pause_event=pause_event,
        )
        players = ((payload or {}).get("response") or {}).get("players", []) if isinstance(payload, dict) else []
        for player in players:
            if isinstance(player, dict) and player.get("steamid"):
                result[str(player.get("steamid"))] = player
    return result


def fetch_player_bans(
    session: requests.Session,
    api_key: str,
    steamids: list[str],
    *,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for chunk in _chunks(steamids, 100):
        payload = _request_json(
            session,
            PLAYER_BANS_URL,
            params={"key": api_key, "steamids": ",".join(chunk)},
            timeout=timeout,
            request_delay=request_delay,
            log_callback=log_callback,
            stop_event=stop_event,
            pause_event=pause_event,
        )
        players = payload.get("players", []) if isinstance(payload, dict) else []
        for player in players:
            if isinstance(player, dict) and player.get("SteamId"):
                result[str(player.get("SteamId"))] = player
    return result


def fetch_friend_list(
    session: requests.Session,
    api_key: str,
    steamid: str,
    *,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> list[dict[str, Any]]:
    payload = _request_json(
        session,
        FRIEND_LIST_URL,
        params={"key": api_key, "steamid": steamid, "relationship": "friend", "format": "json"},
        timeout=timeout,
        request_delay=request_delay,
        log_callback=log_callback,
        stop_event=stop_event,
        pause_event=pause_event,
    )
    friends = ((payload or {}).get("friendslist") or {}).get("friends", []) if isinstance(payload, dict) else []
    return [friend for friend in friends if isinstance(friend, dict)] if isinstance(friends, list) else []


def fetch_owned_games(
    session: requests.Session,
    api_key: str,
    steamid: str,
    *,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> tuple[int | None, list[dict[str, Any]]]:
    payload = _request_json(
        session,
        OWNED_GAMES_URL,
        params={
            "key": api_key,
            "steamid": steamid,
            "include_appinfo": 1,
            "include_played_free_games": 1,
            "format": "json",
        },
        timeout=timeout,
        request_delay=request_delay,
        log_callback=log_callback,
        stop_event=stop_event,
        pause_event=pause_event,
    )
    response = payload.get("response", {}) if isinstance(payload, dict) else {}
    games = response.get("games", [])
    return _safe_int(response.get("game_count")), [game for game in games if isinstance(game, dict)] if isinstance(games, list) else []


def fetch_recent_games(
    session: requests.Session,
    api_key: str,
    steamid: str,
    *,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> list[dict[str, Any]]:
    payload = _request_json(
        session,
        RECENT_GAMES_URL,
        params={"key": api_key, "steamid": steamid, "count": 0, "format": "json"},
        timeout=timeout,
        request_delay=request_delay,
        log_callback=log_callback,
        stop_event=stop_event,
        pause_event=pause_event,
    )
    games = ((payload or {}).get("response") or {}).get("games", []) if isinstance(payload, dict) else []
    return [game for game in games if isinstance(game, dict)] if isinstance(games, list) else []


def fetch_steam_level(
    session: requests.Session,
    api_key: str,
    steamid: str,
    *,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> str:
    payload = _request_json(
        session,
        STEAM_LEVEL_URL,
        params={"key": api_key, "steamid": steamid, "format": "json"},
        timeout=timeout,
        request_delay=request_delay,
        log_callback=log_callback,
        stop_event=stop_event,
        pause_event=pause_event,
    )
    response = payload.get("response", {}) if isinstance(payload, dict) else {}
    return str(response.get("player_level", ""))


def fetch_player_badges(
    session: requests.Session,
    api_key: str,
    steamid: str,
    *,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> list[dict[str, Any]]:
    payload = _request_json(
        session,
        BADGES_URL,
        params={"key": api_key, "steamid": steamid, "format": "json"},
        timeout=timeout,
        request_delay=request_delay,
        log_callback=log_callback,
        stop_event=stop_event,
        pause_event=pause_event,
    )
    badges = ((payload or {}).get("response") or {}).get("badges", []) if isinstance(payload, dict) else []
    return [badge for badge in badges if isinstance(badge, dict)] if isinstance(badges, list) else []


def fetch_player_achievements(
    session: requests.Session,
    api_key: str,
    steamid: str,
    appid: int,
    *,
    language: str,
    timeout: float,
    request_delay: float,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> list[dict[str, Any]]:
    payload = _request_json(
        session,
        PLAYER_ACHIEVEMENTS_URL,
        params={"key": api_key, "steamid": steamid, "appid": appid, "l": language},
        timeout=timeout,
        request_delay=request_delay,
        log_callback=log_callback,
        stop_event=stop_event,
        pause_event=pause_event,
    )
    stats = payload.get("playerstats", {}) if isinstance(payload, dict) else {}
    achievements = stats.get("achievements", []) if isinstance(stats, dict) else []
    return [achievement for achievement in achievements if isinstance(achievement, dict)] if isinstance(achievements, list) else []


def _base_app_row(item: SteamWorkItem, query_time: str) -> dict[str, Any]:
    return {
        "来源类型": item.source_type,
        "搜索词": item.source,
        "AppID": item.appid,
        "查询时间": query_time,
    }


def _build_app_row(
    item: SteamWorkItem,
    data: dict[str, Any],
    query_time: str,
    *,
    status: str,
    review_summary: dict[str, Any] | None = None,
    current_players: Any = "",
    achievement_count: Any = "",
    achievement_sample: Any = "",
    news_count: int = 0,
    review_count: int = 0,
) -> dict[str, Any]:
    review_summary = review_summary or {}
    price = data.get("price_overview") if isinstance(data.get("price_overview"), dict) else {}
    release = data.get("release_date") if isinstance(data.get("release_date"), dict) else {}
    platforms = data.get("platforms") if isinstance(data.get("platforms"), dict) else {}
    metacritic = data.get("metacritic") if isinstance(data.get("metacritic"), dict) else {}
    recommendations = data.get("recommendations") if isinstance(data.get("recommendations"), dict) else {}
    screenshots = data.get("screenshots") if isinstance(data.get("screenshots"), list) else []
    movies = data.get("movies") if isinstance(data.get("movies"), list) else []
    pc_requirements = data.get("pc_requirements")
    mac_requirements = data.get("mac_requirements")
    linux_requirements = data.get("linux_requirements")
    support_info = data.get("support_info") if isinstance(data.get("support_info"), dict) else {}
    achievements = data.get("achievements") if isinstance(data.get("achievements"), dict) else {}
    if not achievement_count:
        achievement_count = achievements.get("total", "")
    if not achievement_sample:
        achievement_sample = _achievement_highlights_text(achievements)
    total_reviews = _to_int(review_summary.get("total_reviews"))
    total_positive = _to_int(review_summary.get("total_positive"))
    total_negative = _to_int(review_summary.get("total_negative"))
    positive_rate = ""
    if total_reviews:
        positive_rate = f"{total_positive / total_reviews:.2%}"
    return {
        "来源类型": item.source_type,
        "搜索词": item.source,
        "AppID": item.appid,
        "游戏名": data.get("name") or item.seed_name,
        "状态": status,
        "类型": data.get("type", ""),
        "是否免费": _yes_no(data.get("is_free")),
        "价格": price.get("final_formatted") or ("免费" if data.get("is_free") else ""),
        "原价": price.get("initial_formatted", ""),
        "折扣": str(price.get("discount_percent", "")),
        "发行日期": release.get("date", ""),
        "是否即将发行": _yes_no(release.get("coming_soon")),
        "开发商": " / ".join(data.get("developers", []) or []),
        "发行商": " / ".join(data.get("publishers", []) or []),
        "类型标签": _list_desc(data.get("genres")),
        "分类标签": _list_desc(data.get("categories")),
        "支持平台": " / ".join(key for key, enabled in platforms.items() if enabled),
        "支持语言": _clean_text(data.get("supported_languages")),
        "年龄限制": str(data.get("required_age", "")),
        "控制器支持": data.get("controller_support", ""),
        "PC最低配置": _requirements_text(pc_requirements, "minimum"),
        "PC推荐配置": _requirements_text(pc_requirements, "recommended"),
        "Mac最低配置": _requirements_text(mac_requirements, "minimum"),
        "Mac推荐配置": _requirements_text(mac_requirements, "recommended"),
        "Linux最低配置": _requirements_text(linux_requirements, "minimum"),
        "Linux推荐配置": _requirements_text(linux_requirements, "recommended"),
        "DLC AppID列表": _join(data.get("dlc", []) or []),
        "Demo列表": _demos_text(data.get("demos")),
        "所属完整游戏": _appid_name(data.get("fullgame")),
        "安装包ID列表": _join(data.get("packages", []) or []),
        "套餐/版本列表": _package_groups_text(data.get("package_groups")),
        "内容描述": _content_descriptors_text(data.get("content_descriptors")),
        "分级信息": _ratings_text(data.get("ratings")),
        "客服网站": support_info.get("url", ""),
        "客服邮箱": support_info.get("email", ""),
        "客服电话": support_info.get("phone", ""),
        "第三方账号提示": _clean_text(data.get("ext_user_account_notice"), max_length=1000),
        "DRM提示": _clean_text(data.get("drm_notice"), max_length=1000),
        "法律声明": _clean_text(data.get("legal_notice"), max_length=1000),
        "短简介": _clean_text(data.get("short_description"), max_length=1000),
        "详细简介": _clean_text(data.get("detailed_description"), max_length=3000),
        "总评描述": review_summary.get("review_score_desc", ""),
        "总评分数": review_summary.get("review_score", ""),
        "总评数": total_reviews or "",
        "好评数": total_positive or "",
        "差评数": total_negative or "",
        "好评率": positive_rate,
        "推荐数": recommendations.get("total", ""),
        "Metacritic": metacritic.get("score", ""),
        "当前在线": current_players,
        "成就数量": achievement_count,
        "成就样本列表": achievement_sample,
        "新闻样本数": news_count,
        "玩家评论样本数": review_count,
        "商店链接": _store_url(item.appid),
        "官网": data.get("website", ""),
        "头图": data.get("header_image", ""),
        "胶囊图": data.get("capsule_image", ""),
        "胶囊图V5": data.get("capsule_imagev5", ""),
        "背景图": data.get("background", ""),
        "原始背景图": data.get("background_raw", ""),
        "截图列表": _join([shot.get("path_full", "") for shot in screenshots if isinstance(shot, dict)]),
        "视频列表": _join([_movie_url(movie) for movie in movies if isinstance(movie, dict)]),
        "查询时间": query_time,
    }


def _build_review_row(item: SteamWorkItem, app_data: dict[str, Any], review: dict[str, Any], query_time: str) -> dict[str, Any]:
    author = review.get("author") if isinstance(review.get("author"), dict) else {}
    steamid = str(author.get("steamid", "") or "")
    return {
        "来源类型": item.source_type,
        "搜索词": item.source,
        "AppID": item.appid,
        "游戏名": app_data.get("name") or item.seed_name,
        "评论ID": review.get("recommendationid", ""),
        "SteamID": steamid,
        "是否推荐": _yes_no(review.get("voted_up")),
        "语言": review.get("language", ""),
        "Steam购买": _yes_no(review.get("steam_purchase")),
        "免费获得": _yes_no(review.get("received_for_free")),
        "抢先体验期间": _yes_no(review.get("written_during_early_access")),
        "主要在SteamDeck": _yes_no(review.get("primarily_steam_deck")),
        "拥有游戏数": author.get("num_games_owned", ""),
        "作者评论数": author.get("num_reviews", ""),
        "总游玩小时": _minutes_to_hours(author.get("playtime_forever")),
        "近两周游玩小时": _minutes_to_hours(author.get("playtime_last_two_weeks")),
        "评价时游玩小时": _minutes_to_hours(author.get("playtime_at_review")),
        "最后游玩时间": _format_timestamp(author.get("last_played")),
        "有价值票数": review.get("votes_up", ""),
        "欢乐票数": review.get("votes_funny", ""),
        "加权评分": review.get("weighted_vote_score", ""),
        "开发者回复": _clean_text(review.get("developer_response"), max_length=3000),
        "开发者回复时间": _format_timestamp(review.get("timestamp_dev_responded")),
        "发布时间": _format_timestamp(review.get("timestamp_created")),
        "最后更新时间": _format_timestamp(review.get("timestamp_updated")),
        "评论链接": f"https://steamcommunity.com/profiles/{steamid}/recommended/{item.appid}/" if steamid else "",
        "评论内容": _clean_text(review.get("review"), max_length=5000),
        "查询时间": query_time,
    }


def _build_news_row(item: SteamWorkItem, app_data: dict[str, Any], news: dict[str, Any], query_time: str) -> dict[str, Any]:
    return {
        "来源类型": item.source_type,
        "搜索词": item.source,
        "AppID": item.appid,
        "游戏名": app_data.get("name") or item.seed_name,
        "新闻ID": news.get("gid", ""),
        "标题": _clean_text(news.get("title")),
        "作者": news.get("author", ""),
        "发布时间": _format_timestamp(news.get("date")),
        "来源名称": news.get("feedlabel", ""),
        "来源标签": news.get("feedname", ""),
        "Feed名称": news.get("feedname", ""),
        "Feed类型": news.get("feed_type", ""),
        "外部链接": _yes_no(news.get("is_external_url")),
        "链接": news.get("url", ""),
        "内容摘要": _clean_text(news.get("contents"), max_length=2000),
        "查询时间": query_time,
    }


def collect_player_bundle(
    context: SteamPlayerContext,
    *,
    api_key: str,
    language: str,
    summary: dict[str, Any] | None = None,
    ban: dict[str, Any] | None = None,
    config: dict[str, Any] | None = None,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> SteamPlayerBundle:
    config = config or {}
    timeout = float(config.get("request_timeout", 30) or 30)
    request_delay = float(config.get("request_delay", 0.2) or 0.0)
    collect_owned_games = _config_bool(config, "collect_player_owned_games", "是")
    collect_recent_games = _config_bool(config, "collect_player_recent_games", "是")
    collect_target_achievements = _config_bool(config, "collect_player_target_achievements", "是")
    collect_badges = _config_bool(config, "collect_player_badges", "是")
    collect_friends = _config_bool(config, "collect_player_friends", "是")
    owned_games_limit = max(0, int(config.get("player_owned_games_limit", 0) or 0))
    session = requests.Session()
    query_time = _now_text()
    summary = summary or {}
    ban = ban or {}
    errors: list[str] = []

    level = ""
    try:
        level = fetch_steam_level(
            session,
            api_key,
            context.steamid,
            timeout=timeout,
            request_delay=request_delay,
            log_callback=log_callback,
            stop_event=stop_event,
            pause_event=pause_event,
        )
    except Exception as exc:
        errors.append(f"level: {exc}")

    game_count: int | None = None
    owned_games: list[dict[str, Any]] = []
    if collect_owned_games:
        try:
            game_count, owned_games = fetch_owned_games(
                session,
                api_key,
                context.steamid,
                timeout=timeout,
                request_delay=request_delay,
                log_callback=log_callback,
                stop_event=stop_event,
                pause_event=pause_event,
            )
        except Exception as exc:
            errors.append(f"owned_games: {exc}")

    recent_games: list[dict[str, Any]] = []
    if collect_recent_games:
        try:
            recent_games = fetch_recent_games(
                session,
                api_key,
                context.steamid,
                timeout=timeout,
                request_delay=request_delay,
                log_callback=log_callback,
                stop_event=stop_event,
                pause_event=pause_event,
            )
        except Exception as exc:
            errors.append(f"recent_games: {exc}")

    badges: list[dict[str, Any]] = []
    if collect_badges:
        try:
            badges = fetch_player_badges(
                session,
                api_key,
                context.steamid,
                timeout=timeout,
                request_delay=request_delay,
                log_callback=log_callback,
                stop_event=stop_event,
                pause_event=pause_event,
            )
        except Exception as exc:
            errors.append(f"badges: {exc}")

    friends: list[dict[str, Any]] = []
    if collect_friends:
        try:
            friends = fetch_friend_list(
                session,
                api_key,
                context.steamid,
                timeout=timeout,
                request_delay=request_delay,
                log_callback=log_callback,
                stop_event=stop_event,
                pause_event=pause_event,
            )
        except Exception as exc:
            errors.append(f"friends: {exc}")

    achievements: list[dict[str, Any]] = []
    if collect_target_achievements and context.target_appid:
        try:
            achievements = fetch_player_achievements(
                session,
                api_key,
                context.steamid,
                context.target_appid,
                language=language,
                timeout=timeout,
                request_delay=request_delay,
                log_callback=log_callback,
                stop_event=stop_event,
                pause_event=pause_event,
            )
        except Exception as exc:
            errors.append(f"achievements: {exc}")

    if collect_owned_games and game_count is None and not owned_games:
        errors.append("游戏库私密或为空")
    if not summary:
        errors.append("玩家公开资料为空")

    if owned_games_limit:
        owned_games = sorted(owned_games, key=lambda game: int(game.get("playtime_forever", 0) or 0), reverse=True)[:owned_games_limit]

    library_total_hours = sum(float(game.get("playtime_forever", 0) or 0) / 60 for game in owned_games)
    profile_row = _build_player_profile_row(
        context,
        summary,
        ban,
        query_time,
        level=level,
        badge_count=len(badges),
        friend_count=len(friends),
        game_count=game_count if game_count is not None else len(owned_games),
        library_total_hours=library_total_hours,
        recent_count=len(recent_games),
        errors=errors,
    )
    return SteamPlayerBundle(
        profile_row=profile_row,
        library_rows=[_build_player_library_row(context, game, query_time) for game in owned_games],
        recent_rows=[_build_player_recent_row(context, game, query_time) for game in recent_games],
        achievement_rows=[_build_player_achievement_row(context, achievement, query_time) for achievement in achievements],
        badge_rows=[_build_player_badge_row(context, badge, query_time) for badge in badges],
        friend_rows=[_build_player_friend_row(context, friend, query_time) for friend in friends],
        meta={
            "library_count": len(owned_games),
            "recent_count": len(recent_games),
            "achievement_count": len(achievements),
            "badge_count": len(badges),
            "friend_count": len(friends),
        },
    )


def _player_base_row(context: SteamPlayerContext, query_time: str) -> dict[str, Any]:
    return {
        "来源类型": context.source_type,
        "搜索词": context.source,
        "目标AppID": context.target_appid or "",
        "目标游戏名": context.target_game_name,
        "SteamID": context.steamid,
        "查询时间": query_time,
    }


def _build_player_profile_row(
    context: SteamPlayerContext,
    summary: dict[str, Any],
    ban: dict[str, Any],
    query_time: str,
    *,
    level: str = "",
    badge_count: int = 0,
    friend_count: int = 0,
    game_count: int = 0,
    library_total_hours: float = 0.0,
    recent_count: int = 0,
    errors: list[str] | None = None,
) -> dict[str, Any]:
    row = _player_base_row(context, query_time)
    row.update(
        {
            "昵称": summary.get("personaname", ""),
            "主页链接": summary.get("profileurl", ""),
            "头像": summary.get("avatarfull") or summary.get("avatarmedium") or summary.get("avatar", ""),
            "国家/地区": summary.get("loccountrycode", ""),
            "省州": summary.get("locstatecode", ""),
            "城市ID": summary.get("loccityid", ""),
            "真实名": summary.get("realname", ""),
            "资料可见性": summary.get("communityvisibilitystate", ""),
            "资料状态": summary.get("profilestate", ""),
            "在线状态": summary.get("personastate", ""),
            "最后在线时间": _format_timestamp(summary.get("lastlogoff")),
            "账号创建时间": _format_timestamp(summary.get("timecreated")),
            "目标游戏是否推荐": context.voted_up,
            "目标游戏总游玩小时": context.target_play_hours,
            "目标游戏近两周小时": context.target_recent_hours,
            "目标游戏评价时小时": context.target_review_hours,
            "目标游戏最后游玩时间": context.target_last_played,
            "Steam等级": level,
            "徽章数": badge_count,
            "公开好友数": friend_count,
            "公开游戏数": game_count or "",
            "公开游戏总小时": round(library_total_hours, 1) if library_total_hours else "",
            "近两周游戏数": recent_count or "",
            "社区封禁": _yes_no(ban.get("CommunityBanned")),
            "VAC封禁": _yes_no(ban.get("VACBanned")),
            "VAC封禁数": ban.get("NumberOfVACBans", ""),
            "游戏封禁数": ban.get("NumberOfGameBans", ""),
            "距上次封禁天数": ban.get("DaysSinceLastBan", ""),
            "经济封禁": ban.get("EconomyBan", ""),
            "隐私/错误": _join(errors or []),
        }
    )
    return row


def _build_player_library_row(context: SteamPlayerContext, game: dict[str, Any], query_time: str) -> dict[str, Any]:
    row = _player_base_row(context, query_time)
    row.update(
        {
            "库游戏AppID": game.get("appid", ""),
            "库游戏名": game.get("name", ""),
            "总游玩小时": _minutes_to_hours(game.get("playtime_forever")),
            "近两周游玩小时": _minutes_to_hours(game.get("playtime_2weeks")),
            "最后游玩时间": _format_timestamp(game.get("rtime_last_played")),
            "有公开社区统计": _yes_no(game.get("has_community_visible_stats")),
            "图标": game.get("img_icon_url", ""),
        }
    )
    return row


def _build_player_recent_row(context: SteamPlayerContext, game: dict[str, Any], query_time: str) -> dict[str, Any]:
    row = _player_base_row(context, query_time)
    row.update(
        {
            "最近游戏AppID": game.get("appid", ""),
            "最近游戏名": game.get("name", ""),
            "近两周游玩小时": _minutes_to_hours(game.get("playtime_2weeks")),
            "总游玩小时": _minutes_to_hours(game.get("playtime_forever")),
        }
    )
    return row


def _build_player_achievement_row(context: SteamPlayerContext, achievement: dict[str, Any], query_time: str) -> dict[str, Any]:
    row = _player_base_row(context, query_time)
    row.update(
        {
            "成就API名": achievement.get("apiname", ""),
            "成就名称": achievement.get("name", ""),
            "成就描述": achievement.get("description", ""),
            "是否解锁": _yes_no(int(achievement.get("achieved", 0) or 0) == 1),
            "解锁时间": _format_timestamp(achievement.get("unlocktime")),
        }
    )
    return row


def _build_player_badge_row(context: SteamPlayerContext, badge: dict[str, Any], query_time: str) -> dict[str, Any]:
    row = _player_base_row(context, query_time)
    row.update(
        {
            "徽章ID": badge.get("badgeid", ""),
            "等级": badge.get("level", ""),
            "XP": badge.get("xp", ""),
            "稀有度": badge.get("scarcity", ""),
            "完成时间": _format_timestamp(badge.get("completion_time")),
            "关联AppID": badge.get("appid", ""),
            "社区物品ID": badge.get("communityitemid", ""),
            "边框颜色": badge.get("border_color", ""),
        }
    )
    return row


def _build_player_friend_row(context: SteamPlayerContext, friend: dict[str, Any], query_time: str) -> dict[str, Any]:
    row = _player_base_row(context, query_time)
    row.update(
        {
            "好友SteamID": friend.get("steamid", ""),
            "关系": friend.get("relationship", ""),
            "加好友时间": _format_timestamp(friend.get("friend_since")),
        }
    )
    return row


def write_player_enrichment_rows(
    contexts: list[SteamPlayerContext],
    *,
    api_key: str,
    language: str,
    writer,
    checkpoint=None,
    config: dict[str, Any] | None = None,
    log_callback=None,
    stop_event=None,
    pause_event=None,
) -> dict[str, int]:
    config = config or {}
    if not contexts:
        return {"players": 0, "library_rows": 0, "recent_rows": 0, "achievement_rows": 0, "badge_rows": 0, "friend_rows": 0}
    if not api_key:
        log_warn(log_callback, "玩家画像增强需要 Steam Web API Key，已跳过。")
        return {"players": 0, "library_rows": 0, "recent_rows": 0, "achievement_rows": 0, "badge_rows": 0, "friend_rows": 0}

    max_players = max(0, int(config.get("max_player_profiles", config.get("max_review_players", 2000)) or 0))
    parallel_workers = max(1, min(8, int(config.get("player_parallel_workers", 2) or 2)))
    timeout = float(config.get("request_timeout", 30) or 30)
    request_delay = float(config.get("request_delay", 0.2) or 0.0)
    contexts = dedupe_player_contexts(contexts, max_players)
    steamids = [context.steamid for context in contexts]
    log_line(log_callback, f"开始增强 Steam 评论玩家画像：{len(contexts)} 人，并发 {parallel_workers}。")

    session = requests.Session()
    try:
        summary_map = fetch_player_summaries(
            session,
            api_key,
            steamids,
            timeout=timeout,
            request_delay=request_delay,
            log_callback=log_callback,
            stop_event=stop_event,
            pause_event=pause_event,
        )
    except Exception as exc:
        log_warn(log_callback, f"玩家公开资料批量获取失败，将继续尝试逐项信息：{exc}")
        summary_map = {}
    try:
        ban_map = fetch_player_bans(
            session,
            api_key,
            steamids,
            timeout=timeout,
            request_delay=request_delay,
            log_callback=log_callback,
            stop_event=stop_event,
            pause_event=pause_event,
        )
    except Exception as exc:
        log_warn(log_callback, f"玩家封禁信息批量获取失败：{exc}")
        ban_map = {}

    stats = {"players": 0, "library_rows": 0, "recent_rows": 0, "achievement_rows": 0, "badge_rows": 0, "friend_rows": 0}
    futures: dict[Any, SteamPlayerContext] = {}
    with ThreadPoolExecutor(max_workers=parallel_workers) as executor:
        for context in contexts:
            if wait_if_paused(pause_event, stop_event) or should_stop(stop_event):
                break
            if checkpoint is not None:
                claimed, claim_status = checkpoint.claim_item(context.checkpoint_key, positive_count_fields=("player_count",))
                if not claimed:
                    if claim_status == "completed":
                        log_line(log_callback, f"断点跳过已完成玩家：{context.checkpoint_key}")
                    continue
            futures[
                executor.submit(
                    collect_player_bundle,
                    context,
                    api_key=api_key,
                    language=language,
                    summary=summary_map.get(context.steamid, {}),
                    ban=ban_map.get(context.steamid, {}),
                    config=config,
                    log_callback=log_callback,
                    stop_event=stop_event,
                    pause_event=pause_event,
                )
            ] = context

        total = len(futures)
        for index, future in enumerate(as_completed(futures), start=1):
            context = futures[future]
            if should_stop(stop_event):
                if checkpoint is not None:
                    checkpoint.release_item(context.checkpoint_key)
                continue
            try:
                bundle = future.result()
            except InterruptedError:
                if checkpoint is not None:
                    checkpoint.release_item(context.checkpoint_key)
                continue
            except Exception as exc:
                if checkpoint is not None:
                    checkpoint.release_item(context.checkpoint_key)
                log_warn(log_callback, f"[{index}/{total}] 玩家画像增强失败 {context.steamid}: {exc}")
                continue

            _write_player_bundle(writer, bundle)
            stats["players"] += 1
            stats["library_rows"] += len(bundle.library_rows)
            stats["recent_rows"] += len(bundle.recent_rows)
            stats["achievement_rows"] += len(bundle.achievement_rows)
            stats["badge_rows"] += len(bundle.badge_rows)
            stats["friend_rows"] += len(bundle.friend_rows)
            if checkpoint is not None:
                checkpoint.mark_completed(context.checkpoint_key, {"player_count": 1, **bundle.meta})
            log_line(
                log_callback,
                f"[{index}/{total}] 完成玩家画像 {context.steamid}："
                f"库 {len(bundle.library_rows)}，最近 {len(bundle.recent_rows)}，"
                f"成就 {len(bundle.achievement_rows)}，徽章 {len(bundle.badge_rows)}，好友 {len(bundle.friend_rows)}。",
            )
    return stats


def _write_player_bundle(writer, bundle: SteamPlayerBundle) -> None:
    writer.writerow("玩家画像", bundle.profile_row)
    for row in bundle.library_rows:
        writer.writerow("玩家游戏库", row)
    for row in bundle.recent_rows:
        writer.writerow("最近游玩", row)
    for row in bundle.achievement_rows:
        writer.writerow("目标游戏成就", row)
    for row in bundle.badge_rows:
        writer.writerow("玩家徽章", row)
    for row in bundle.friend_rows:
        writer.writerow("玩家好友", row)


def _movie_url(movie: dict[str, Any]) -> str:
    webm = movie.get("webm") if isinstance(movie.get("webm"), dict) else {}
    mp4 = movie.get("mp4") if isinstance(movie.get("mp4"), dict) else {}
    return webm.get("max") or webm.get("480") or mp4.get("max") or mp4.get("480") or ""


def _store_url(appid: int) -> str:
    return f"https://store.steampowered.com/app/{appid}/"


def _to_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def _safe_int(value: Any) -> int | None:
    try:
        if value is None or str(value).strip() == "":
            return None
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def _row_value(row: tuple[Any, ...], index: dict[str, int], name: str) -> str:
    pos = index.get(name)
    if pos is None or pos >= len(row):
        return ""
    return str(row[pos] or "").strip()


def _chunks(values: list[str], size: int) -> list[list[str]]:
    return [values[index:index + size] for index in range(0, len(values), max(1, size))]


def dedupe_player_contexts(contexts: list[SteamPlayerContext], limit: int = 0) -> list[SteamPlayerContext]:
    deduped: list[SteamPlayerContext] = []
    seen: set[str] = set()
    for context in contexts:
        if not context.steamid:
            continue
        key = context.checkpoint_key
        if key in seen:
            continue
        seen.add(key)
        deduped.append(context)
        if limit and len(deduped) >= limit:
            break
    return deduped


def run_steam_api_spider(
    api_key: str,
    app_ids_or_urls: str | list[str],
    keywords: str | list[str],
    language: str,
    country: str,
    collect_reviews_choice: str,
    max_reviews: int,
    collect_news_choice: str,
    max_news: int,
    log_callback,
    finish_callback,
    stop_event,
    *,
    pause_event=None,
    config: dict[str, Any] | None = None,
):
    config = config or {}
    api_key = str(api_key or "").strip()
    language = str(language or DEFAULT_LANGUAGE).strip() or DEFAULT_LANGUAGE
    country = str(country or DEFAULT_COUNTRY).strip().upper() or DEFAULT_COUNTRY
    app_ids = parse_app_ids(app_ids_or_urls)
    keyword_list = normalize_keywords(keywords)
    if not app_ids and not keyword_list:
        raise ValueError("至少需要输入一个 Steam AppID/商店链接，或一个关键词。")

    max_apps_per_keyword = max(1, int(config.get("max_apps_per_keyword", 100) or 100))
    collect_reviews = str(collect_reviews_choice or "否") == "是"
    collect_news = str(collect_news_choice or "否") == "是"
    max_reviews = max(0, int(max_reviews or 0))
    max_news = max(0, int(max_news or 0))
    parallel_workers = max(1, min(8, int(config.get("parallel_workers", 1) or 1)))
    save_batch_size = max(1, int(config.get("save_batch_size", 10) or 10))
    collect_current_players = _config_bool(config, "collect_current_players", "是")
    collect_achievements = _config_bool(config, "collect_achievements", "否")
    collect_review_player_profiles = _config_bool(config, "collect_review_player_profiles", "否")
    max_review_players = max(0, int(config.get("max_review_players", config.get("max_player_profiles", 2000)) or 0))
    include_non_games = str(config.get("include_non_games", "否") or "否")

    scope = {
        "app_ids": app_ids,
        "keywords": keyword_list,
        "language": language,
        "country": country,
        "collect_reviews": collect_reviews,
        "max_reviews": max_reviews,
        "collect_news": collect_news,
        "max_news": max_news,
        "max_apps_per_keyword": max_apps_per_keyword,
        "collect_current_players": collect_current_players,
        "collect_achievements": collect_achievements,
        "collect_review_player_profiles": collect_review_player_profiles,
        "max_review_players": max_review_players,
        "collect_player_owned_games": str(config.get("collect_player_owned_games", "是") or "是"),
        "player_owned_games_limit": int(config.get("player_owned_games_limit", 0) or 0),
        "collect_player_recent_games": str(config.get("collect_player_recent_games", "是") or "是"),
        "collect_player_target_achievements": str(config.get("collect_player_target_achievements", "是") or "是"),
        "collect_player_badges": str(config.get("collect_player_badges", "是") or "是"),
        "collect_player_friends": str(config.get("collect_player_friends", "是") or "是"),
        "include_non_games": include_non_games,
    }
    checkpoint = open_task_checkpoint(
        "steam_api_research",
        scope,
        log_callback,
        merge_on_keys=("app_ids", "keywords"),
        merge_keep_keys=(
            "language",
            "country",
            "collect_reviews",
            "max_reviews",
            "collect_news",
            "max_news",
            "collect_current_players",
            "collect_achievements",
            "collect_review_player_profiles",
            "max_review_players",
            "collect_player_owned_games",
            "player_owned_games_limit",
            "collect_player_recent_games",
            "collect_player_target_achievements",
            "collect_player_badges",
            "collect_player_friends",
            "include_non_games",
        ),
    )
    default_output_path = build_output_path(
        "steam",
        f"steam_api_research_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
        channel="api_research",
    )
    output_path, writer = open_checkpointed_multi_sheet_writer(
        checkpoint,
        default_output_path,
        {"游戏信息": APP_FIELDS, "玩家评论": REVIEW_FIELDS, "新闻": NEWS_FIELDS, **PLAYER_SHEETS_FIELDS},
        log_callback,
        autosave_every=save_batch_size,
    )
    checkpoint.add_output_path(output_path)

    items: list[SteamWorkItem] = []
    for appid in app_ids:
        items.append(SteamWorkItem(appid=appid, source_type="直接输入"))
    for keyword in keyword_list:
        if should_stop(stop_event):
            break
        log_line(log_callback, f"发现关键词候选游戏：{keyword}，上限 {max_apps_per_keyword} 个。")
        try:
            items.extend(
                discover_apps_for_keyword(
                    keyword,
                    api_key=api_key,
                    language=language,
                    country=country,
                    max_apps=max_apps_per_keyword,
                    config=config,
                    log_callback=log_callback,
                    stop_event=stop_event,
                    pause_event=pause_event,
                )
            )
        except Exception as exc:
            log_error(log_callback, f"关键词发现失败：{keyword}: {exc}")

    deduped: list[SteamWorkItem] = []
    seen_keys: set[str] = set()
    for item in items:
        if item.checkpoint_key not in seen_keys:
            seen_keys.add(item.checkpoint_key)
            deduped.append(item)
    log_line(log_callback, f"本次 Steam API 任务候选 {len(deduped)} 个采集项，并发数 {parallel_workers}。")

    submitted: dict[Any, SteamWorkItem] = {}
    completed_this_run = 0
    try:
        with ThreadPoolExecutor(max_workers=parallel_workers) as executor:
            for item in deduped:
                if should_stop(stop_event):
                    break
                if wait_if_paused(pause_event, stop_event):
                    break
                claimed, claim_status = checkpoint.claim_item(item.checkpoint_key)
                if not claimed:
                    if claim_status == "completed":
                        log_line(log_callback, f"断点跳过已完成游戏：{item.checkpoint_key}")
                    elif claim_status == "active":
                        log_line(log_callback, f"分流跳过正在采集的游戏：{item.checkpoint_key}")
                    continue
                future = executor.submit(
                    collect_steam_app_bundle,
                    item,
                    api_key=api_key,
                    language=language,
                    country=country,
                    collect_reviews=collect_reviews,
                    max_reviews=max_reviews,
                    collect_news=collect_news,
                    max_news=max_news,
                    config=config,
                    log_callback=log_callback,
                    stop_event=stop_event,
                    pause_event=pause_event,
                )
                submitted[future] = item

            total = len(submitted)
            for index, future in enumerate(as_completed(submitted), start=1):
                item = submitted[future]
                if should_stop(stop_event):
                    checkpoint.release_item(item.checkpoint_key)
                    continue
                try:
                    bundle = future.result()
                except InterruptedError:
                    checkpoint.release_item(item.checkpoint_key)
                    continue
                except Exception as exc:
                    checkpoint.release_item(item.checkpoint_key)
                    log_error(log_callback, f"[{index}/{total}] Steam 采集失败 AppID={item.appid}: {exc}")
                    continue
                writer.writerow("游戏信息", bundle.app_row)
                for row in bundle.review_rows:
                    writer.writerow("玩家评论", row)
                for row in bundle.news_rows:
                    writer.writerow("新闻", row)
                player_stats = {}
                if collect_review_player_profiles and bundle.review_rows:
                    player_config = dict(config)
                    player_config["max_player_profiles"] = max_review_players
                    contexts = player_contexts_from_review_rows(bundle.review_rows, limit=max_review_players)
                    player_stats = write_player_enrichment_rows(
                        contexts,
                        api_key=api_key,
                        language=language,
                        writer=writer,
                        config=player_config,
                        log_callback=log_callback,
                        stop_event=stop_event,
                        pause_event=pause_event,
                    )
                if should_stop(stop_event):
                    checkpoint.release_item(item.checkpoint_key)
                    continue
                checkpoint.mark_completed(item.checkpoint_key, {**bundle.meta, **player_stats})
                completed_this_run += 1
                log_line(
                    log_callback,
                    f"[{index}/{total}] 完成 {bundle.app_row.get('游戏名') or item.appid}，"
                    f"评论 {len(bundle.review_rows)}，新闻 {len(bundle.news_rows)}，"
                    f"画像 {player_stats.get('players', 0) if player_stats else 0}。",
                )
    finally:
        try:
            writer.save()
        except Exception:
            pass
        checkpoint.close_run()

    if should_stop(stop_event):
        log_warn(log_callback, f"Steam API 任务已停止，本轮完成 {completed_this_run} 个采集项。")
    else:
        log_line(log_callback, f"Steam API 任务完成，本轮完成 {completed_this_run} 个采集项。")
    finish_callback(output_path)
    return output_path


def run_steam_player_profiles_spider(
    api_key: str,
    xlsx_path: str,
    steam_ids: str | list[str],
    target_appid_or_url: str,
    target_game_name: str,
    language: str,
    log_callback,
    finish_callback,
    stop_event,
    *,
    pause_event=None,
    config: dict[str, Any] | None = None,
):
    config = config or {}
    api_key = str(api_key or "").strip()
    if not api_key:
        raise ValueError("Steam 评论玩家画像补采需要填写 Steam Web API Key。")
    language = str(language or DEFAULT_LANGUAGE).strip() or DEFAULT_LANGUAGE
    target_appids = parse_app_ids(target_appid_or_url)
    target_appid = target_appids[0] if target_appids else None
    max_players = max(0, int(config.get("max_player_profiles", 2000) or 0))

    contexts: list[SteamPlayerContext] = []
    xlsx_path = str(xlsx_path or "").strip()
    if xlsx_path:
        if not Path(xlsx_path).exists():
            raise ValueError(f"玩家评论 Excel 不存在：{xlsx_path}")
        contexts.extend(read_player_contexts_from_xlsx(xlsx_path))
    contexts.extend(
        player_contexts_from_steam_ids(
            steam_ids,
            target_appid=target_appid,
            target_game_name=str(target_game_name or "").strip(),
        )
    )
    contexts = dedupe_player_contexts(contexts, max_players)
    if not contexts:
        raise ValueError("没有识别到可补采的 SteamID。")

    scope = {
        "player_targets": [context.checkpoint_key for context in contexts],
        "language": language,
        "max_player_profiles": max_players,
        "collect_player_owned_games": str(config.get("collect_player_owned_games", "是") or "是"),
        "player_owned_games_limit": int(config.get("player_owned_games_limit", 0) or 0),
        "collect_player_recent_games": str(config.get("collect_player_recent_games", "是") or "是"),
        "collect_player_target_achievements": str(config.get("collect_player_target_achievements", "是") or "是"),
        "collect_player_badges": str(config.get("collect_player_badges", "是") or "是"),
        "collect_player_friends": str(config.get("collect_player_friends", "是") or "是"),
    }
    checkpoint = open_task_checkpoint("steam_player_profiles", scope, log_callback)
    default_output_path = build_output_path(
        "steam",
        f"steam_player_profiles_{time.strftime('%Y%m%d_%H%M%S')}.xlsx",
        channel="player_profiles",
    )
    output_path, writer = open_checkpointed_multi_sheet_writer(
        checkpoint,
        default_output_path,
        PLAYER_SHEETS_FIELDS,
        log_callback,
        autosave_every=max(1, int(config.get("save_batch_size", 10) or 10)),
    )
    checkpoint.add_output_path(output_path)
    try:
        stats = write_player_enrichment_rows(
            contexts,
            api_key=api_key,
            language=language,
            writer=writer,
            checkpoint=checkpoint,
            config=config,
            log_callback=log_callback,
            stop_event=stop_event,
            pause_event=pause_event,
        )
        log_line(
            log_callback,
            f"Steam 玩家画像补采完成：玩家 {stats.get('players', 0)}，"
            f"游戏库行 {stats.get('library_rows', 0)}，最近游玩 {stats.get('recent_rows', 0)}，"
            f"成就 {stats.get('achievement_rows', 0)}，徽章 {stats.get('badge_rows', 0)}，好友 {stats.get('friend_rows', 0)}。",
        )
    finally:
        try:
            writer.save()
        except Exception:
            pass
        checkpoint.close_run()
    finish_callback(output_path)
    return output_path
