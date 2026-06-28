"""
Excel (XLSX) 文件处理工具，使用 openpyxl 库提供流式行写入支持，具备防注入机制与原子写入保护。
"""

from __future__ import annotations

import os
from collections.abc import Iterable, Mapping
from pathlib import Path
from typing import Any

from openpyxl import Workbook

from src.core.csv_utils import sanitize_csv_cell


def sanitize_xlsx_cell(value: Any) -> Any:
    """
    清洗并格式化单个 Excel 单元格的值。
    防范 CSV/Excel 注入漏洞：在 Excel 中，以 =, +, -, @ 开头的内容会被误识别为公式并执行。
    如果在其头部追加单引号 ' 则强制 Excel 将其作为纯文本显示。
    """
    value = sanitize_csv_cell(value)
    if isinstance(value, str) and value[:1] in {"=", "+", "-", "@"}:
        return "'" + value
    return value


class XlsxRowWriter:
    """
    单表格 Excel 行写入器，适合流式追加单 sheet 数据。
    """

    def __init__(
        self,
        output_path: str,
        fieldnames: Iterable[str],
        sheet_name: str = "数据",
        autosave_every: int = 1,
    ):
        """
        Args:
            output_path: 输出的 Excel 文件路径
            fieldnames: 表头字段列表
            sheet_name: 表格的工作表名称
            autosave_every: 累计写入多少行时执行一次磁盘保存（防数据丢失）
        """
        self.output_path = str(output_path)
        Path(self.output_path).parent.mkdir(parents=True, exist_ok=True)
        self.fieldnames = list(fieldnames)
        self.autosave_every = max(1, int(autosave_every or 1))
        self._rows_since_save = 0
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        # Excel 规定 Sheet 名称最大长度为 31 字符，超出直接被截断
        self.worksheet.title = sheet_name[:31] or "数据"
        self.worksheet.append(self.fieldnames)
        self.save()

    def writerow(self, row: Mapping[str, Any]):
        """
        写入单行数据。
        """
        self._append_row(row)
        self._rows_since_save += 1
        if self._rows_since_save >= self.autosave_every:
            self.save()

    def writerows(self, rows: Iterable[Mapping[str, Any]]):
        """
        批量写入多行数据。
        """
        wrote_rows = False
        for row in rows:
            self._append_row(row)
            wrote_rows = True
        if wrote_rows:
            self.save()

    def _append_row(self, row: Mapping[str, Any]):
        self.worksheet.append([sanitize_xlsx_cell(row.get(field, "")) for field in self.fieldnames])

    def save(self):
        """
        原子式保存当前 Excel 文件，写入临时文件再执行 os.replace，
        防止保存过程中程序被异常终止或断电导致原文件损坏。
        """
        temp_path = f"{self.output_path}.tmp"
        self.workbook.save(temp_path)
        try:
            os.replace(temp_path, self.output_path)
        except OSError:
            # 在某些 Windows 网络共享磁盘或者并发占用下，os.replace 可能会报 OSError，
            # 此时 fallback 回直接保存
            self.workbook.save(self.output_path)
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except OSError:
                    pass
        self._rows_since_save = 0


class MultiSheetXlsxWriter:
    """
    多表格 Excel 写入器，支持同时在单个工作簿下写入多个 Sheet。
    """

    def __init__(
        self,
        output_path: str,
        sheets_fields: dict[str, list[str]],
        autosave_every: int = 1,
    ):
        """
        Args:
            output_path: 输出的 Excel 文件路径
            sheets_fields: 键值对，键为 sheet_name，值为表头字段列表
            autosave_every: 缓存多少行时自动保存
        """
        self.output_path = str(output_path)
        Path(self.output_path).parent.mkdir(parents=True, exist_ok=True)
        self.sheets_fields = sheets_fields
        self.autosave_every = max(1, int(autosave_every or 1))
        self._rows_since_save = 0

        self.workbook = Workbook()
        # 移除 openpyxl 实例化时自动创建的默认 Sheet，以便仅保留用户自定义的 Sheet
        default_sheet = self.workbook.active
        if default_sheet is not None:
            self.workbook.remove(default_sheet)

        self.worksheets = {}
        for sheet_name, fieldnames in sheets_fields.items():
            # 同样遵循 31 字符上限限制
            ws = self.workbook.create_sheet(title=sheet_name[:31] or "Sheet")
            ws.append(list(fieldnames))
            self.worksheets[sheet_name] = ws
        self.save()

    def writerow(self, sheet_name: str, row: Mapping[str, Any]):
        """
        向指定名称的工作表中追加单行数据。
        """
        if sheet_name not in self.worksheets:
            import logging
            logging.getLogger(__name__).warning("writerow: sheet '%s' not registered, row skipped", sheet_name)
            return
        fieldnames = self.sheets_fields[sheet_name]
        ws = self.worksheets[sheet_name]
        ws.append([sanitize_xlsx_cell(row.get(field, "")) for field in fieldnames])
        self._rows_since_save += 1
        if self._rows_since_save >= self.autosave_every:
            self.save()

    def save(self):
        """
        原子式保存当前工作薄文件。
        """
        temp_path = f"{self.output_path}.tmp"
        self.workbook.save(temp_path)
        try:
            os.replace(temp_path, self.output_path)
        except OSError:
            self.workbook.save(self.output_path)
            if os.path.exists(temp_path):
                try:
                    os.remove(temp_path)
                except OSError:
                    pass
        self._rows_since_save = 0


