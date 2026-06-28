"""
任务输出汇总模块。

将单次任务产出的多个 Excel 文件汇总到一个工作簿中，每个源文件作为独立 sheet，
保留原始表头与数据结构，避免不同表头的数据错位。原始分文件保持不变，
汇总文件额外生成并落到任务同目录（output/<平台>/<渠道>/<动作日期>/）。
"""

from __future__ import annotations

import os
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.core.app_logging import log_line, log_warn
from src.core.output import build_output_path

# Excel 工作表名称最大长度限制
_SHEET_NAME_MAX_LEN = 31
# 匹配 8 位连续数字的日期目录名（YYYYMMDD）
_DATE_DIR_RE = re.compile(r"^\d{8}$")
# 匹配文件名末尾的时间戳后缀（_YYYYMMDD_HHMMSS 或 _YYYYMMDD），用于简化 sheet 名
_TRAILING_STAMP_RE = re.compile(r"_\d{8}_\d{6}$|_\d{8}$")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    r"""
    生成合法且不重复的工作表名称。
    Excel 不允许 sheet 名包含 : \ / ? * [ ] 且长度不超过 31 字符。
    """
    cleaned = re.sub(r"[\\/:*?\[\]]", "", name or "").strip()
    if not cleaned:
        cleaned = "sheet"
    base = cleaned[:_SHEET_NAME_MAX_LEN]
    candidate = base
    suffix = 2
    while candidate in used:
        tail = f"_{suffix}"
        candidate = f"{base[: _SHEET_NAME_MAX_LEN - len(tail)]}{tail}"
        suffix += 1
    used.add(candidate)
    return candidate


def _derive_output_path(source_paths: list[str]) -> str:
    """
    从首个源文件路径反推平台/渠道/动作日期，构造同目录下的汇总文件路径。
    反推失败时降级放到首个源文件所在目录，使用通用文件名。
    """
    first = Path(source_paths[0])
    date = first.parent.name if _DATE_DIR_RE.match(first.parent.name) else None
    channel = first.parent.parent.name if date else None
    platform = first.parent.parent.parent.name if channel else None

    if platform and channel and date:
        filename = f"{channel}_summary_{date}.xlsx"
        return build_output_path(platform, filename, channel=channel, run_date=date)

    # 降级：直接落到源文件同目录
    fallback_dir = first.parent
    fallback_dir.mkdir(parents=True, exist_ok=True)
    import time

    return str(fallback_dir / f"summary_{time.strftime('%Y%m%d_%H%M%S')}.xlsx")


def _collect_rows(src_ws) -> list[tuple]:
    """
    读取源 worksheet 的所有非空数据行（保留表头与数据顺序）。
    """
    rows: list[tuple] = []
    for row in src_ws.iter_rows(values_only=True):
        if row is None:
            continue
        if all(value is None or str(value).strip() == "" for value in row):
            continue
        rows.append(row)
    return rows


def summarize_outputs(output_paths, log_callback=None) -> str | None:
    """
    将单次任务产出的多个 Excel 文件汇总到一个工作簿。

    每个源文件作为汇总工作簿里的一个独立 sheet，保留原始表头与数据。
    少于 2 个有效文件时不汇总（返回 None）。原始分文件不会被修改或删除。

    Args:
        output_paths: 本次任务产出的 xlsx 路径列表（可含 None / 重复 / 不存在项）
        log_callback: 日志回调

    Returns:
        str | None: 汇总文件绝对路径；未汇总时返回 None
    """
    # 过滤：去重、去 None、去临时文件、去不存在路径
    seen: set[str] = set()
    valid_paths: list[str] = []
    for raw in output_paths or []:
        if not raw:
            continue
        path_str = str(raw)
        name = Path(path_str).name
        if name.startswith("~$"):
            continue
        if not Path(path_str).exists():
            continue
        if path_str in seen:
            continue
        seen.add(path_str)
        valid_paths.append(path_str)

    if len(valid_paths) < 2:
        return None

    output_path = _derive_output_path(valid_paths)
    # 防自包含：复制阶段跳过路径等于汇总输出路径的文件
    output_path_norm = os.path.normpath(output_path)

    log_line(log_callback, f"开始汇总 {len(valid_paths)} 个文件到：{output_path}")

    summary_wb = Workbook()
    # 删除 Workbook 默认创建的空 sheet
    default_ws = summary_wb.active
    if default_ws is not None:
        summary_wb.remove(default_ws)

    used_sheet_names: set[str] = set()
    merged_files = 0

    for src_path in valid_paths:
        if os.path.normpath(src_path) == output_path_norm:
            continue
        wb = None
        try:
            wb = load_workbook(src_path, read_only=True, data_only=True)
        except Exception as exc:
            log_warn(log_callback, f"  跳过无法读取的文件 {Path(src_path).name}（{exc}）")
            continue

        try:
            # 去掉文件名末尾的时间戳后缀，让 sheet 名更短更有意义（如 tiktok_keyword_关键词）
            base_name = _TRAILING_STAMP_RE.sub("", Path(src_path).stem) or Path(src_path).stem
            src_worksheets = wb.worksheets
            multi_sheet = len(src_worksheets) > 1
            file_data_rows = 0
            for ws in src_worksheets:
                rows = _collect_rows(ws)
                if len(rows) <= 1:
                    # 跳过完全空或只有表头的 sheet，不在汇总中创建空表
                    continue
                # 单 sheet 源文件用文件名；多 sheet 源文件用「文件名__源sheet名」保留来源与结构
                if multi_sheet and ws.title:
                    name_seed = f"{base_name}__{ws.title}"
                else:
                    name_seed = base_name
                sheet_name = _safe_sheet_name(name_seed, used_sheet_names)
                dst_ws = summary_wb.create_sheet(title=sheet_name)
                for row in rows:
                    dst_ws.append(row)
                file_data_rows += len(rows) - 1
            if file_data_rows:
                merged_files += 1
                log_line(log_callback, f"  已并入 {Path(src_path).name}（{file_data_rows} 条数据）")
        except Exception as exc:
            log_warn(log_callback, f"  汇总 {Path(src_path).name} 时出错：{exc}")
        finally:
            try:
                wb.close()
            except Exception:
                pass

    if merged_files == 0:
        log_warn(log_callback, "汇总失败：未能从任何源文件读取到有效数据。")
        return None

    # 原子保存：先写临时文件再替换，避免中途异常留下半成品
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    temp_path = f"{output_path}.tmp"
    try:
        summary_wb.save(temp_path)
        os.replace(temp_path, output_path)
    except OSError:
        # 某些环境下 os.replace 可能失败，退避直接保存目标路径
        try:
            summary_wb.save(output_path)
        except Exception as exc:
            log_warn(log_callback, f"汇总文件保存失败：{exc}")
            return None
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except OSError:
                pass

    log_line(log_callback, f"汇总完成：合并 {merged_files} 个文件 → {output_path}")
    return output_path
