import json
import tempfile
import unittest
from contextlib import contextmanager
from pathlib import Path
from unittest.mock import patch

from src.core.task_checkpoint import (
    load_tool_inputs,
    open_checkpointed_row_writer,
    open_task_checkpoint,
    save_tool_inputs,
    task_fingerprint,
)
from src.core.xlsx import MultiSheetXlsxWriter, XlsxRowWriter


@contextmanager
def _patched_checkpoint_roots(tmp: str):
    root = Path(tmp)
    with patch("src.core.task_checkpoint.get_workspace_root", return_value=root), patch(
        "src.core.task_checkpoint.get_app_state_root",
        return_value=root,
    ):
        yield root


class TestTaskCheckpoint(unittest.TestCase):
    def test_fingerprint_is_stable_for_dict_order(self):
        left = task_fingerprint("tool", {"b": 2, "a": [1, 2]})
        right = task_fingerprint("tool", {"a": [1, 2], "b": 2})
        self.assertEqual(left, right)

    def test_save_and_load_tool_inputs(self):
        with tempfile.TemporaryDirectory() as tmp:
            with _patched_checkpoint_roots(tmp):
                save_tool_inputs("tool_x", {"keywords": "a\nb", "limit_time": "是"})
                self.assertEqual(load_tool_inputs("tool_x")["keywords"], "a\nb")

    def test_checkpoint_migrates_legacy_workspace_cache(self):
        with tempfile.TemporaryDirectory() as tmp:
            workspace = Path(tmp) / "workspace"
            state = Path(tmp) / "state"
            legacy = workspace / "output" / "checkpoints" / "last_inputs"
            legacy.mkdir(parents=True)
            (legacy / "tool_legacy.json").write_text(
                json.dumps({"values": {"keywords": "old"}}),
                encoding="utf-8",
            )

            with patch("src.core.task_checkpoint.get_workspace_root", return_value=workspace), patch(
                "src.core.task_checkpoint.get_app_state_root",
                return_value=state,
            ):
                self.assertEqual(load_tool_inputs("tool_legacy")["keywords"], "old")
                self.assertTrue((state / "checkpoints" / "last_inputs" / "tool_legacy.json").exists())

    def test_checkpoint_marks_completed_items(self):
        with tempfile.TemporaryDirectory() as tmp:
            with _patched_checkpoint_roots(tmp):
                checkpoint = open_task_checkpoint("tool_y", {"links": ["A", "B"]})
                self.assertFalse(checkpoint.is_completed("A"))
                checkpoint.mark_completed("A", {"row": 1})

                reloaded = open_task_checkpoint("tool_y", {"links": ["A", "B"]})
                self.assertTrue(reloaded.is_completed("a"))
                self.assertEqual(reloaded.completed_count(), 1)

    def test_checkpoint_persists_custom_state(self):
        with tempfile.TemporaryDirectory() as tmp:
            with _patched_checkpoint_roots(tmp):
                checkpoint = open_task_checkpoint("tool_state", {"links": ["A"]})
                checkpoint.set_state("seed_cache", {"count": 2, "items": ["a", "b"]})

                reloaded = open_task_checkpoint("tool_state", {"links": ["A"]})
                self.assertEqual(reloaded.get_state("seed_cache")["count"], 2)
                self.assertEqual(reloaded.get_state("seed_cache")["items"], ["a", "b"])

    def test_successful_completion_distinguishes_old_zero_count_records(self):
        with tempfile.TemporaryDirectory() as tmp:
            with _patched_checkpoint_roots(tmp):
                checkpoint = open_task_checkpoint("tool_counts", {"links": ["A", "B"]})
                checkpoint.completed["a"] = {
                    "completed_at": "2026-01-01 00:00:00",
                    "meta": {"tweet_count": 0},
                }
                checkpoint.completed["b"] = {
                    "completed_at": "2026-01-01 00:00:00",
                    "meta": {"tweet_count": 2},
                }
                checkpoint.save()

                reloaded = open_task_checkpoint("tool_counts", {"links": ["A", "B"]})
                self.assertFalse(reloaded.is_successfully_completed("A", positive_count_fields=("tweet_count",)))
                self.assertTrue(reloaded.is_successfully_completed("B", positive_count_fields=("tweet_count",)))

                reloaded.mark_completed("A", {"tweet_count": 0})
                self.assertFalse(reloaded.is_successfully_completed("A", positive_count_fields=("tweet_count",)))

    def test_checkpoint_log_mentions_input_count_and_history_count(self):
        with tempfile.TemporaryDirectory() as tmp:
            with _patched_checkpoint_roots(tmp):
                checkpoint = open_task_checkpoint("tool_log", {"profile_urls": ["A", "B"]})
                checkpoint.mark_completed("A", {"profile_ok": 1})

                messages = []
                open_task_checkpoint("tool_log", {"profile_urls": ["A", "B"]}, log_callback=messages.append)

                self.assertIn("本次输入 2 条博主链接", messages[0])
                self.assertIn("已加载 1 条历史断点记录", messages[0])

    def test_checkpoint_can_merge_same_input_with_different_runtime_limits(self):
        with tempfile.TemporaryDirectory() as tmp:
            with _patched_checkpoint_roots(tmp):
                old_checkpoint = open_task_checkpoint("tool_merge", {"profile_urls": ["A", "B"], "max_scrolls": 50})
                old_output_path = Path(tmp) / "old.xlsx"
                old_output_path.touch()
                old_checkpoint.add_output_path(str(old_output_path))
                old_checkpoint.mark_completed("A", {"tweet_count": 3})

                messages = []
                new_checkpoint = open_task_checkpoint(
                    "tool_merge",
                    {"profile_urls": ["A", "B"], "max_scrolls": 10},
                    log_callback=messages.append,
                    merge_on_keys=("profile_urls",),
                )

                self.assertTrue(new_checkpoint.is_successfully_completed("A", positive_count_fields=("tweet_count",)))
                self.assertIn("已从旧参数任务合并 1 条历史记录", "\n".join(messages))
                self.assertEqual(new_checkpoint.latest_output_path(), str(old_output_path))

    def test_parallel_checkpoints_claim_inputs_without_overlap(self):
        with tempfile.TemporaryDirectory() as tmp:
            with _patched_checkpoint_roots(tmp):
                first = open_task_checkpoint("tool_parallel", {"profile_urls": ["A", "B"]})
                second = open_task_checkpoint("tool_parallel", {"profile_urls": ["A", "B"]})

                self.assertEqual(first.claim_item("A"), (True, "claimed"))
                self.assertEqual(second.claim_item("A"), (False, "active"))
                self.assertEqual(second.claim_item("B"), (True, "claimed"))

                first.mark_completed("A", {"row": 1})
                second.mark_completed("B", {"row": 2})

                reloaded = open_task_checkpoint("tool_parallel", {"profile_urls": ["A", "B"]})
                self.assertTrue(reloaded.is_completed("A"))
                self.assertTrue(reloaded.is_completed("B"))
                self.assertEqual(reloaded.completed_count(), 2)
                first.close_run()
                second.close_run()
                reloaded.close_run()

    def test_parallel_outputs_use_separate_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            with _patched_checkpoint_roots(tmp):
                first = open_task_checkpoint("tool_parallel_output", {"profile_urls": ["A", "B"]})
                second = open_task_checkpoint("tool_parallel_output", {"profile_urls": ["A", "B"]})
                default_output_path = str(Path(tmp) / "rows.xlsx")

                first_path, first_writer = open_checkpointed_row_writer(first, default_output_path, ["name"])
                first.add_output_path(first_path)
                second_path, second_writer = open_checkpointed_row_writer(second, default_output_path, ["name"])
                second.add_output_path(second_path)
                first_writer.writerow({"name": "first"})
                second_writer.writerow({"name": "second"})
                first_writer.save()
                second_writer.save()

                self.assertNotEqual(first_path, second_path)
                self.assertTrue(Path(first_path).exists())
                self.assertTrue(Path(second_path).exists())
                first.close_run()
                second.close_run()

    def test_xlsx_row_writer_can_append_existing_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "rows.xlsx"
            writer = XlsxRowWriter(str(output_path), ["name", "count"])
            writer.writerow({"name": "one", "count": 1})
            writer.save()

            resumed = XlsxRowWriter(str(output_path), ["name", "count"], append=True)
            resumed.writerow({"name": "two", "count": 2})
            resumed.save()

            from openpyxl import load_workbook

            rows = list(load_workbook(output_path).active.iter_rows(values_only=True))
            self.assertEqual(rows, [("name", "count"), ("one", 1), ("two", 2)])

    def test_multi_sheet_writer_can_append_existing_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "multi.xlsx"
            fields = {"profiles": ["url"], "works": ["url", "title"]}
            writer = MultiSheetXlsxWriter(str(output_path), fields)
            writer.writerow("profiles", {"url": "a"})
            writer.save()

            resumed = MultiSheetXlsxWriter(str(output_path), fields, append=True)
            resumed.writerow("works", {"url": "b", "title": "title"})
            resumed.save()

            from openpyxl import load_workbook

            workbook = load_workbook(output_path)
            self.assertEqual(list(workbook["profiles"].iter_rows(values_only=True)), [("url",), ("a",)])
            self.assertEqual(list(workbook["works"].iter_rows(values_only=True)), [("url", "title"), ("b", "title")])

    def test_multi_sheet_writer_rejects_append_to_different_schema_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            output_path = Path(tmp) / "old_single_sheet.xlsx"
            writer = XlsxRowWriter(str(output_path), ["topic", "profile"])
            writer.writerow({"topic": "#old", "profile": "https://example.com"})
            writer.save()

            with self.assertRaises(ValueError):
                MultiSheetXlsxWriter(str(output_path), {"profiles": ["url"], "works": ["url", "title"]}, append=True)


if __name__ == "__main__":
    unittest.main(verbosity=2)
